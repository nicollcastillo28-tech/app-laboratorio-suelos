"""
GEODELTA LAB - App para digitar ensayos de laboratorio de suelos
Estructura: Proyecto -> Perforación (Sondeo/Apique/Fuente-Cantera) -> Muestra -> Ensayo

Cómo correrla en tu computador:
    streamlit run app.py
"""

import html
import json
import math
import os
import re
import zipfile
from datetime import date, datetime, timedelta
from io import BytesIO

import extra_streamlit_components as stx
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

import db

# ════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE LA PÁGINA
# ════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Geodelta Lab", page_icon="🧪", layout="wide", initial_sidebar_state="collapsed")

APP_VERSION = "v5.0.0"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_GRANULOMETRIA = os.path.join(BASE_DIR, "templates", "CLASIFICACION_DE_SUELOS.xlsm")
TEMPLATE_BITACORA_ORDEN = os.path.join(BASE_DIR, "templates", "GDA-FL-003_bitacora_orden.xlsx")
TEMPLATE_HUMEDAD = os.path.join(BASE_DIR, "templates", "GDA-FLC-014_humedad_natural.xlsx")
TEMPLATE_MASA_UNITARIA = os.path.join(BASE_DIR, "templates", "GDA-FLC-004_masa_unitaria.xlsx")
TEMPLATE_CBR = os.path.join(BASE_DIR, "templates", "GDA-FLC-013_cbr.xlsx")

ROLE_LABELS = {"jefe": "Jefe de Laboratorio", "laboratorista": "Laboratorista", "ingeniero": "Director Técnico"}
ROLE_INICIALES = {"jefe": "JL", "laboratorista": "LB", "ingeniero": "DT"}

# ════════════════════════════════════════════════════════════════════
# ESTILOS — paleta "Verdant Precision" (Primary #007A33 · Secondary #4A7862 · Tertiary #D1E8D5 · Neutral #212121)
# ════════════════════════════════════════════════════════════════════
PRIMARY, PRIMARY_DARK, PRIMARY_CONTAINER = "#007A33", "#00591F", "#0B3D22"
SECONDARY, SECONDARY_CONTAINER = "#4A7862", "#D1E8D5"
TERTIARY = "#D1E8D5"
NEUTRAL = "#6B7570"
SUCCESS, SUCCESS_LIGHT = "#16A34A", "#DCFCE7"
WARNING, WARNING_LIGHT = "#D97706", "#FEF3C7"
DANGER, DANGER_LIGHT = "#DC2626", "#FEE2E2"
SURFACE, BG, BORDER, TEXT = "#FFFFFF", "#F7FAF8", "#D6D9D5", "#212121"
MUTED = NEUTRAL

st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;700&family=Material+Symbols+Outlined:wght,FILL@100..700,0..1&display=swap');
    .material-symbols-outlined, [data-testid="stMarkdownContainer"] span.material-symbols-outlined,
    [data-testid="stMarkdownContainer"] p span.material-symbols-outlined {{
        font-family: 'Material Symbols Outlined' !important;
        font-weight: normal; font-style: normal; text-transform: none;
        letter-spacing: normal; word-wrap: normal; white-space: nowrap; direction: ltr;
        -webkit-font-smoothing: antialiased;
        font-variation-settings: 'FILL' 0, 'wght' 400, 'GRAD' 0, 'opsz' 24;
        vertical-align: middle; line-height: 1; font-size: 1.15em; display: inline-block;
    }}
    .msi-fill {{ font-variation-settings: 'FILL' 1, 'wght' 400, 'GRAD' 0, 'opsz' 24; }}
    html, body, [class*="css"] {{ font-family: 'IBM Plex Sans', 'Segoe UI', sans-serif; }}
    [data-testid="stAppViewContainer"], [data-testid="stMarkdownContainer"], [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] span, [data-testid="stMarkdownContainer"] li {{
        font-family: 'IBM Plex Sans', 'Segoe UI', sans-serif !important;
    }}
    /* Fondo de cuadrícula sutil (papel milimetrado) solo en el fondo de la página — las tarjetas
       son opacas ({SURFACE}) así que la cuadrícula solo asoma en los márgenes/espacios entre ellas. */
    .stApp {{
        background-color: {BG};
        background-image:
            linear-gradient(rgba(33,33,33,0.05) 1px, transparent 1px),
            linear-gradient(90deg, rgba(33,33,33,0.05) 1px, transparent 1px);
        background-size: 26px 26px;
    }}
    [data-testid="collapsedControl"] {{ display: none; }}
    section[data-testid="stSidebar"] {{ display: none; }}
    .font-mono {{ font-family: 'JetBrains Mono', monospace; }}

    /* ---- TOP APP BAR (desktop / tablet ancho) ---- */
    .st-key-topbar {{
        position: sticky; top: 0; z-index: 999; background: {SURFACE};
        border-bottom: 1px solid {BORDER}; padding: 10px 4px 6px 4px; margin-bottom: 8px;
    }}
    .st-key-topbar .stButton button {{
        font-family: 'JetBrains Mono', monospace; font-weight: 700;
        letter-spacing: 0.04em; text-transform: uppercase; white-space: nowrap;
        font-size: clamp(10px, 1.1vw, 12px); padding-left: 8px; padding-right: 8px;
    }}
    .topbar-brand {{ display: flex; align-items: center; gap: 10px; height: 38px; }}
    .topbar-brand .brand-title {{
        font-size: clamp(15px, 2vw, 20px); font-weight: 700; color: {PRIMARY}; letter-spacing: -0.02em; white-space: nowrap;
    }}
    .topbar-avatar {{
        width: 36px; height: 36px; border-radius: 999px; background: {PRIMARY_CONTAINER}; color: #FFFFFF;
        display: flex; align-items: center; justify-content: center; font-weight: 700; font-size: 13px;
        border: 1px solid {BORDER}; margin-left: auto; flex-shrink: 0;
    }}

    /* ---- BOTTOM NAV (celular y tablet, ambas orientaciones) ----
       900px solo cubría tablet en vertical; en horizontal (~1024-1194px, iPad/Android típico)
       caía en el layout de escritorio con la nav de arriba apretada contra el padding por
       defecto de Streamlit — de ahí se veía "apretada" incluso en "modo escritorio". */
    .st-key-bottomnav {{ display: none; }}
    @media (max-width: 1180px) {{
        .st-key-topbar-nav {{ display: none; }}
        div[data-testid="stColumn"]:has(.st-key-topbar-nav) {{ display: none; }}
        .st-key-bottomnav {{
            display: block; position: fixed; bottom: 0; left: 0; width: 100%; z-index: 999;
            background: {SURFACE}; border-top: 1px solid {BORDER}; padding: 6px 8px 8px 8px; box-shadow: 0 -2px 8px rgba(0,0,0,0.04);
        }}
        .st-key-bottomnav .stButton button {{
            font-family: 'JetBrains Mono', monospace; font-size: clamp(8px, 2.6vw, 10px); text-transform: uppercase;
            letter-spacing: 0.02em; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
            padding-top: 10px; padding-bottom: 10px; padding-left: 2px; padding-right: 2px;
        }}
        .st-key-bottomnav [data-testid="stHorizontalBlock"] {{ flex-direction: row !important; flex-wrap: nowrap !important; gap: 6px !important; }}
        .st-key-bottomnav [data-testid="stColumn"] {{ width: auto !important; flex: 1 1 0 !important; min-width: 0 !important; }}
        /* Streamlit deja ~5rem de aire a cada lado por defecto (clase ".main" ya no existe en
           esta versión, por eso el selector viejo nunca aplicaba) — en una tablet eso se come
           casi el 20% del ancho útil y hace ver todo más apretado de lo que hace falta. */
        [data-testid="stMainBlockContainer"] {{
            padding-left: 1.25rem !important; padding-right: 1.25rem !important; padding-bottom: 76px !important;
        }}
    }}
    @media (max-width: 420px) {{
        .topbar-brand .brand-title {{ display: none; }}
        .st-key-bottomnav .stButton button {{ font-size: 9px; }}
    }}

    /* Contenedores con borde nativos de Streamlit = nuestras "tarjetas" (sin bugs de HTML suelto) */
    /* OJO: en esta versión de Streamlit ya no existe stVerticalBlockBorderWrapper como wrapper
    aparte — st.container(border=True) marca el propio stVerticalBlock con
    data-test-scroll-behavior="normal" (no lo tienen los stVerticalBlock sin borde). */
    div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"] {{
        border-radius: 12px !important; border: 1px solid {BORDER} !important;
        box-shadow: 0 1px 4px rgba(11,28,48,0.08) !important; background: {SURFACE} !important;
    }}
    /* st.expander (tarjetas de "Perforaciones y muestras", "Historial de Cambios", etc.) trae de
    fábrica un borde casi invisible (20% de opacidad) y sin sombra ni fondo propio — se pierde
    contra el fondo cuadriculado de la app. Se le da el mismo tratamiento de tarjeta que a los
    st.container(border=True) de arriba, para que se distinga igual de bien. */
    [data-testid="stExpander"] details {{
        border-radius: 12px !important; border: 1px solid {BORDER} !important;
        box-shadow: 0 1px 4px rgba(11,28,48,0.08) !important; background: {SURFACE} !important;
    }}
    /* Barra de título del expander (ej. "S1 — Sondeo · 2 muestra(s)") en verde, para que cada
    perforación se distinga de un vistazo dentro de la lista. */
    [data-testid="stExpander"] summary {{
        background: {SECONDARY_CONTAINER} !important; border-radius: 11px 11px 0 0 !important;
    }}
    [data-testid="stExpander"] summary, [data-testid="stExpander"] summary * {{
        color: {PRIMARY} !important;
    }}
    /* Las filas "etiqueta + campo" de los formularios de ensayo (humedad, límites, granulometría...)
       arman el layout con st.columns para que etiqueta y campo queden lado a lado — pero Streamlit
       las apila solo (etiqueta arriba, campo abajo a todo el ancho) en pantallas angostas, que es
       justo lo que se ve "apachurrado"/mal alineado en tablet. Se fuerza a que sigan en fila dentro
       de las tarjetas; en celular muy angosto (420px) se deja el apilado nativo como respaldo.
       min-width:0 es clave: sin eso, la columna de la etiqueta no se encoge por debajo del ancho
       de su texto más largo y la fila se desborda de la tarjeta en vez de acomodarse (la etiqueta
       larga como "Masa suelo seco + recipiente (g) (16 hrs)" empuja el campo fuera de pantalla).
       Ojo: 0 puro también aplastaba las columnas angostas de badges/botones ("Abrir", el badge de
       Estado) en tablas como "Ensayos asignados" — el botón/badge se encogía por debajo de su
       contenido y el texto quedaba amontonado. Esas columnas (con :has) se excluyen del encogido:
       mantienen su tamaño natural y son las columnas de texto/etiqueta las que ceden espacio.
       Los encabezados de tabla (.assigned-th, ej. "ID PROYECTO") NO se excluyen del encogido —
       su columna tiene que encogerse igual que la columna de datos de abajo (misma proporción),
       si no, encabezado y dato quedan desalineados. En vez de eso, se le quita el nowrap más abajo
       para que el texto pase a dos líneas dentro del fondo verde en vez de desbordarse cortado.
       :not(.st-key-home-actions): la fila de tarjetas de Inicio también tiene
       data-test-scroll-behavior="normal" (no es exclusivo de tarjetas con borde, como se asumía
       antes) y quedaba atrapada por estas reglas de "no encoger" pensadas para tablas — con más
       especificidad que las reglas propias de home-actions más abajo, ganaban ellas y la fila de
       Inicio nunca podía hacer wrap (se quedaba desbordada, exigiendo scroll horizontal). */
    div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"]:not(.st-key-home-actions) [data-testid="stHorizontalBlock"] {{
        flex-wrap: nowrap !important;
    }}
    div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"]:not(.st-key-home-actions) [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {{
        min-width: 0 !important;
    }}
    div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"]:not(.st-key-home-actions) [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:has(.stButton),
    div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"]:not(.st-key-home-actions) [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:has(.badge) {{
        min-width: fit-content !important;
        flex-shrink: 0 !important;
    }}
    /* La regla de arriba (:has) protege la columna de un botón/badge en la fila de DATOS, pero la
       fila de ENCABEZADO (.assigned-th, sin botón ni badge) no calzaba con ese :has y se encogía
       distinto → encabezado y dato quedaban desalineados (ej. "ESTADO"/"ACCIÓN" corridos respecto
       a su columna real). Se protegen las últimas 2 columnas por posición, no por contenido, y solo
       dentro de las tarjetas que son tablas (tienen algún .assigned-th) para no tocar los formularios
       de 2 columnas etiqueta+campo. */
    div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"]:has(.assigned-th)
        [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-last-child(-n+2) {{
        min-width: fit-content !important;
        flex-shrink: 0 !important;
    }}
    @media (max-width: 420px) {{
        div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"] [data-testid="stHorizontalBlock"] {{
            flex-wrap: wrap !important;
        }}
    }}
    /* Acento azul a la izquierda solo en tarjetas de contenido — se excluyen la barra de
       navegación, las tarjetas de notificaciones y las cajas donde se digita información
       (esas no son "otra tarjeta más", son campos de captura). home-actions tampoco: es un
       contenedor que envuelve varias tarjetas propias (cada una ya tiene su estilo, gradiente
       o borde), no una tarjeta en sí — el acento le quedaba como una barra verde de borde a
       borde tapando el grupo entero en vez de resaltar una sola tarjeta. bell-alert/bell-quiet
       (el botón-contenedor de la campana en el topbar) tampoco: es un ícono, no una tarjeta —
       se veía como una rayita verde suelta pegada a la campana. */
    div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"]:not(.st-key-topbar):not(.st-key-topbar-nav):not(.st-key-bottomnav):not(.st-key-notif-popover-body):not(.st-key-fab-new-project):not(.st-key-muestra-obs-box):not(.st-key-home-actions):not(.st-key-bell-alert):not(.st-key-bell-quiet):not([class*="st-key-notif-card-"]) {{
        border-left: 4px solid {PRIMARY} !important;
    }}
    /* Tarjetas de "Ensayos asignados" (una por ensayo): más separación entre sí y sombra
       más marcada para que no se confundan entre ellas ni con el fondo de la página. */
    div[data-testid="stVerticalBlock"][data-test-scroll-behavior="normal"][class*="st-key-ensayo-card-"] {{
        margin-bottom: 12px !important;
        box-shadow: 0 2px 6px rgba(11,28,48,0.10) !important;
    }}

    .section-title {{
        font-size: 12px; font-weight: 700; color: {MUTED}; text-transform: uppercase;
        letter-spacing: 0.06em; border-bottom: 1px solid {BORDER}; padding-bottom: 8px;
        margin-bottom: 14px; margin-top: 4px;
    }}
    .badge {{
        display: inline-block; padding: 3px 10px; border-radius: 999px; font-size: 12px; font-weight: 700;
        font-family: 'IBM Plex Sans', sans-serif;
    }}
    .badge-success {{ background: {SUCCESS_LIGHT}; color: {SUCCESS}; }}
    .badge-warning {{ background: {WARNING_LIGHT}; color: {WARNING}; }}
    .badge-danger {{ background: {DANGER_LIGHT}; color: {DANGER}; }}
    .badge-muted {{ background: #EEF1F5; color: {MUTED}; }}
    .status-circle {{
        display: inline-flex; align-items: center; justify-content: center;
        width: 36px; height: 36px; border-radius: 999px; flex-shrink: 0;
    }}
    .status-circle-success {{ background: {SUCCESS_LIGHT}; color: {SUCCESS}; }}
    .status-circle-warning {{ background: {WARNING_LIGHT}; color: {WARNING}; }}
    .status-circle-danger {{ background: {DANGER_LIGHT}; color: {DANGER}; }}
    .status-circle-muted {{ background: #EEF1F5; color: {MUTED}; }}
    .status-circle-primary {{ background: {SECONDARY_CONTAINER}; color: {PRIMARY}; }}

    /* Línea de tiempo del historial de la muestra (marcador + línea conectora + contenido) */
    .timeline-item {{ display: flex; gap: 14px; }}
    .timeline-marker-col {{ display: flex; flex-direction: column; align-items: center; }}
    .timeline-line {{ width: 2px; flex: 1; min-height: 14px; background: {BORDER}; margin: 4px 0; }}
    .timeline-content {{ flex: 1; padding-bottom: 22px; }}
    .timeline-item:last-child .timeline-content {{ padding-bottom: 2px; }}
    .timeline-titulo {{ font-weight: 700; font-size: 14px; color: {TEXT}; }}
    .timeline-actor {{ font-size: 13px; color: {PRIMARY}; margin-top: 1px; }}
    .timeline-fecha {{ font-size: 12px; color: {MUTED}; margin-top: 4px; }}

    /* Tarjeta con acento a la izquierda, para encabezados de detalle (ej. Detalle de Muestra) */
    .st-key-muestra-header-card {{ border-left: 4px solid {PRIMARY} !important; }}
    .st-key-muestra-obs-box .stTextArea textarea {{ background-color: {SECONDARY_CONTAINER} !important; }}
    /* Botones dentro de la campana de notificaciones — por defecto salen blancos/planos y se
       pierden contra el fondo del popover. */
    .st-key-notif-popover-body .stButton button {{
        background-color: {SECONDARY_CONTAINER}; color: {PRIMARY}; border-color: {SECONDARY_CONTAINER};
    }}
    .st-key-notif-popover-body .stButton button:hover {{
        background-color: {PRIMARY}; color: {SURFACE}; border-color: {PRIMARY};
    }}
    /* Campana: roja mientras haya notificaciones sin leer, para que se note de un vistazo. */
    .st-key-bell-alert button {{
        background-color: {DANGER} !important; color: {SURFACE} !important; border-color: {DANGER} !important;
    }}
    .st-key-bell-alert button:hover {{
        background-color: {DANGER_LIGHT} !important; color: {DANGER} !important; border-color: {DANGER} !important;
    }}
    .role-pill {{
        display: inline-block; padding: 4px 12px; border-radius: 999px; font-size: 11px;
        font-family: 'JetBrains Mono', monospace; font-weight: 700; letter-spacing: 0.04em; text-transform: uppercase;
        background: {SECONDARY_CONTAINER}; color: {PRIMARY};
    }}
    .timestamp-caption {{ color: {MUTED}; font-size: 12px; margin-top: 2px; }}
    div.stButton > button[kind="primary"] {{ background-color: {PRIMARY}; border: 1px solid {PRIMARY}; }}
    div.stButton > button[kind="primary"]:hover {{ background-color: {PRIMARY_DARK}; border-color: {PRIMARY_DARK}; }}
    h1, h2, h3, h4, h5, h6 {{ color: {TEXT}; letter-spacing: -0.02em; font-family: 'IBM Plex Sans', sans-serif !important; }}

    /* Campos digitables con fondo distinto al de la página, para que se note qué se puede editar */
    .stTextInput input, .stTextArea textarea, .stNumberInput input,
    .stDateInput input, [data-testid="stDateInputField"],
    .stSelectbox > div > div, .stMultiSelect > div > div {{
        background-color: {SURFACE} !important;
        border: 1px solid {BORDER} !important;
        border-radius: 8px !important;
    }}
    /* El date_input de Streamlit ya no usa un <input> normal, sino un grupo de "spinbuttons"
       (día/mes/año) sin caja propia — sin esto se ve como texto plano sobre el fondo gris. */
    [data-testid="stDateInputField"] {{ padding: 8px 12px !important; }}
    .stTextInput input:focus, .stTextArea textarea:focus, .stNumberInput input:focus,
    [data-testid="stDateInputField"]:focus-within {{
        border-color: {PRIMARY} !important; box-shadow: 0 0 0 1px {PRIMARY} !important;
    }}
    [data-testid="stDataFrameResizable"], [data-testid="stDataEditorGrid"] {{
        background-color: {SURFACE} !important; border: 1px solid {BORDER} !important; border-radius: 8px;
    }}

    .login-icon {{
        width: 56px; height: 56px; border-radius: 16px; background: {SECONDARY_CONTAINER};
        display: flex; align-items: center; justify-content: center; color: {PRIMARY};
        font-size: 26px; margin: 0 auto 10px auto;
    }}
    .login-title {{ text-align: center; color: {TEXT}; font-weight: 600; font-size: 16px; letter-spacing: -0.01em; margin-bottom: 20px; }}
    .login-footer {{ text-align: center; color: {NEUTRAL}; font-size: 12px; margin-top: 18px; }}

    /* Selector de rol en el login, como tarjetas seleccionables */
    .st-key-login-card [data-testid="stRadio"] > div[role="radiogroup"] {{ display: flex; gap: 10px; }}
    .st-key-login-card [data-testid="stRadio"] [role="radiogroup"] label {{
        flex: 1 1 0; border: 1px solid {BORDER}; border-radius: 10px; padding: 8px 12px !important;
        margin: 0 !important; background: {SURFACE}; white-space: nowrap;
    }}
    .st-key-login-card [data-testid="stRadio"] [role="radiogroup"] label p {{ white-space: nowrap; }}
    .st-key-login-card [data-testid="stRadio"] [role="radiogroup"] label:has(input:checked) {{
        border-color: {PRIMARY}; background: {SECONDARY_CONTAINER};
    }}
    .st-key-login-card [data-testid="stRadio"] input[type="radio"] {{ accent-color: {PRIMARY}; }}
    .st-key-login-card [data-testid="stWidgetLabel"] p {{
        font-family: 'JetBrains Mono', monospace; font-size: 11px; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.05em; color: {NEUTRAL};
    }}
    .st-key-login-card .stButton:has(button[kind="secondary"]) button {{
        color: {NEUTRAL}; border: none; background: transparent;
    }}

    /* ---- BENTO CARDS (inspirado en el diseño de Stitch) ---- */
    .bento-primary {{
        background: linear-gradient(135deg, {PRIMARY_CONTAINER} 0%, {PRIMARY} 100%);
        color: #FFFFFF; border-radius: 16px; padding: 26px 28px; min-height: 168px;
        display: flex; flex-direction: column; justify-content: space-between; margin-bottom: 16px;
    }}
    .bento-primary .bento-icon {{
        background: rgba(255,255,255,0.14); width: 44px; height: 44px; border-radius: 10px;
        display: flex; align-items: center; justify-content: center; font-size: 22px; margin-bottom: 14px;
    }}
    .bento-primary .bento-eyebrow {{ font-size: 11px; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase; opacity: 0.65; }}
    .bento-primary h3 {{ color: #FFFFFF; margin: 4px 0 6px 0; }}
    .bento-primary p {{ opacity: 0.75; font-size: 13px; margin: 0; }}

    .bento-light {{
        background: {SURFACE}; border: 1px solid {BORDER}; border-radius: 16px; padding: 24px 26px;
        min-height: 168px; display: flex; flex-direction: column; justify-content: space-between; margin-bottom: 16px;
    }}
    .bento-light .bento-icon {{
        background: {SECONDARY_CONTAINER}; color: {PRIMARY}; width: 44px; height: 44px; border-radius: 10px;
        display: flex; align-items: center; justify-content: center; font-size: 20px; margin-bottom: 14px;
    }}
    .bento-light h3 {{ color: {PRIMARY}; margin: 4px 0 6px 0; font-size: 18px; }}

    /* Tarjetas de la fila de acciones de Inicio, todas del mismo alto y alineadas.
       El wrap nativo de Streamlit para st.columns() se activa según el ancho de la ventana, no
       según el espacio real disponible en esta fila — en una tablet ancha (pero no tan ancha como
       para que quepan cómodas 3 tarjetas con su texto) no llegaba a activarse y la fila se
       desbordaba, obligando a hacer scroll horizontal para ver la tercera tarjeta. Se le da a cada
       columna un ancho mínimo propio y se fuerza el wrap explícitamente: sobran 3 en pantallas
       anchas, se acomodan 2+1 o 1+1+1 en las angostas. */
    .st-key-home-actions [data-testid="stHorizontalBlock"] {{
        align-items: stretch; flex-wrap: wrap !important; row-gap: 16px;
    }}
    .st-key-home-actions [data-testid="stColumn"] {{
        flex: 1 1 260px !important; min-width: 260px !important; width: auto !important;
    }}
    .st-key-home-actions [data-testid="stElementContainer"]:has(.bento-primary),
    .st-key-home-actions [data-testid="stElementContainer"]:has(.bento-light) {{ flex: 1 1 auto; }}
    .st-key-home-actions [data-testid="stElementContainer"]:has(.bento-primary) .stMarkdown,
    .st-key-home-actions [data-testid="stElementContainer"]:has(.bento-light) .stMarkdown,
    .st-key-home-actions [data-testid="stElementContainer"]:has(.bento-primary) .stMarkdown > div,
    .st-key-home-actions [data-testid="stElementContainer"]:has(.bento-light) .stMarkdown > div,
    .st-key-home-actions [data-testid="stElementContainer"]:has(.bento-primary) [data-testid="stMarkdownContainer"],
    .st-key-home-actions [data-testid="stElementContainer"]:has(.bento-light) [data-testid="stMarkdownContainer"] {{
        height: 100%;
    }}
    .st-key-home-actions .bento-primary, .st-key-home-actions .bento-light {{ height: 100%; }}
    .bento-light p {{ color: {MUTED}; font-size: 13px; margin: 0; }}

    .stat-chip {{
        background: {SURFACE}; border: 1px solid {BORDER}; border-radius: 12px; padding: 14px 16px;
        display: flex; align-items: center; gap: 12px; margin-bottom: 12px;
    }}
    .stat-chip .stat-icon {{
        width: 40px; height: 40px; border-radius: 999px; background: {SECONDARY_CONTAINER};
        display: flex; align-items: center; justify-content: center; font-size: 18px;
    }}
    .stat-chip .stat-label {{ font-size: 11px; font-weight: 700; letter-spacing: 0.06em; text-transform: uppercase; color: {MUTED}; }}
    .stat-chip .stat-value {{ font-size: 20px; font-weight: 800; color: {TEXT}; }}

    /* ---- ACTIVITY TABLE (inspirado en el diseño de Stitch) ---- */
    .activity-table-wrap {{ overflow-x: auto; }}
    .activity-table {{ width: 100%; border-collapse: collapse; font-family: 'IBM Plex Sans', sans-serif; }}
    .activity-table thead th {{
        background: {SECONDARY_CONTAINER}; color: {PRIMARY}; font-family: 'JetBrains Mono', monospace;
        font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em;
        padding: 10px 14px; text-align: left; white-space: nowrap; border-bottom: 1px solid {BORDER};
    }}
    .activity-table tbody td {{
        padding: 12px 14px; border-bottom: 1px solid {BORDER}; font-size: 14px; color: {TEXT}; vertical-align: middle;
    }}
    .activity-table tbody tr:last-child td {{ border-bottom: none; }}
    .activity-table tbody tr:hover {{ background: {BG}; }}
    /* OJO: estas 4 clases se usan sueltas (Actividad reciente, Ensayos asignados, Buscar...)
    fuera de cualquier ".activity-table" — ese wrapper ya no existe en ningún lado del código,
    así que iban sin aplicar (texto a tamaño por defecto, más grande de lo pensado) hasta que
    se les quitó el prefijo muerto. */
    .cell-id {{ font-family: 'JetBrains Mono', monospace; color: {PRIMARY}; font-weight: 800; font-size: 14px; }}
    .cell-title {{ font-weight: 600; color: {TEXT}; font-size: 14px; }}
    .cell-sub {{ font-size: 11.5px; color: {NEUTRAL}; margin-top: 1px; }}
    .cell-muted {{ color: {NEUTRAL}; font-size: 12px; }}
    .activity-footer {{
        display: flex; justify-content: space-between; align-items: center; padding: 10px 14px;
        color: {NEUTRAL}; font-size: 13px;
    }}

    /* ---- ENSAYOS ASIGNADOS (panel de Laboratorista) ---- */
    .assigned-th {{
        background: {SECONDARY_CONTAINER}; color: {PRIMARY}; font-family: 'JetBrains Mono', monospace;
        font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em;
        padding: 8px 10px; border-radius: 6px; margin-bottom: 4px; white-space: normal;
        overflow-wrap: break-word; line-height: 1.3;
    }}
    .assigned-chip {{
        background: {BG}; border: 1px solid {BORDER}; color: {PRIMARY}; font-size: 12px; font-weight: 700;
        padding: 3px 10px; border-radius: 6px; display: inline-block;
    }}
    /* Variantes de color para la columna "Ensayos asignados" de la tabla de muestras — cada
       chip refleja el estado de ESE ensayo puntual (verde finalizado / amarillo en proceso /
       rojo sin iniciar), reemplazando la columna "Estado" aparte que mostraba solo un estado
       agregado de toda la muestra. */
    .assigned-chip-success {{ background: {SUCCESS_LIGHT}; border-color: {SUCCESS}; color: {SUCCESS}; }}
    .assigned-chip-warning {{ background: {WARNING_LIGHT}; border-color: {WARNING}; color: {WARNING}; }}
    .assigned-chip-danger {{ background: {DANGER_LIGHT}; border-color: {DANGER}; color: {DANGER}; }}
    /* Versión compacta para la columna "Ensayos asignados" de la lista de muestras: uno debajo
       del otro (Humedad, Granulometría, Límites...) en vez de lado a lado — así la columna solo
       necesita el ancho del chip más largo, no el de los 3 juntos, y le deja espacio de sobra a
       "Tipo"/"Profundidad" al lado para no partirse en varias líneas. */
    .assigned-chip-sm {{ font-size: 10px; padding: 2px 6px; white-space: nowrap; }}
    .assigned-chip-row {{ display: flex; flex-direction: column; align-items: flex-start; gap: 3px; }}

    /* ---- TARJETAS DE PROYECTO (Proyectos en ejecución) ---- */
    .code-badge {{
        display: inline-block; background: {SECONDARY_CONTAINER}; color: {PRIMARY};
        font-family: 'JetBrains Mono', monospace; font-weight: 800; font-size: 16px;
        letter-spacing: 0.02em; padding: 4px 14px; border-radius: 6px;
    }}
    [class*="st-key-projcard_"] {{
        border-left: 4px solid {PRIMARY} !important;
        background: {SURFACE} !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
    }}

    .perf-code-box {{
        display: inline-flex; align-items: center; justify-content: center;
        width: 30px; height: 30px; border-radius: 7px; background: {PRIMARY_CONTAINER};
        color: #fff; font-family: 'JetBrains Mono', monospace; font-weight: 700; font-size: 12px;
    }}

    /* Botón flotante "+" para crear proyecto desde Proyectos en ejecución */
    .st-key-fab-new-project {{
        position: fixed; right: 24px; bottom: 28px; z-index: 998; width: 56px !important;
    }}
    .st-key-fab-new-project .stButton button {{
        width: 56px; height: 56px; border-radius: 999px !important; padding: 0 !important;
        background: {PRIMARY} !important; border: none !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.28) !important;
    }}
    .st-key-fab-new-project .stButton button span[data-testid="stIconMaterial"] {{
        color: #fff !important; font-size: 26px !important;
    }}
    @media (max-width: 1180px) {{
        .st-key-fab-new-project {{ bottom: 92px; right: 16px; }}
    }}
</style>
""", unsafe_allow_html=True)

# Los campos numéricos de la app (pesos, temperaturas, lecturas de tamiz, etc.) son todos
# st.text_input con placeholder "0.00" / "0.0" / "0" / "001" — Streamlit no tiene un widget de
# texto con teclado numérico nativo, así que se marca por JS (vía un iframe de components.html,
# que sí puede tocar el DOM del documento padre por ser mismo origen) usando esa convención de
# placeholder para poner inputmode="decimal"/"numeric" solo ahí. El resto de campos de texto
# (nombre, dirección, correo...) no calzan el patrón y se quedan con el teclado normal.
# El mismo script hace que Enter salte al siguiente campo de texto en vez de quedarse quieto
# (útil digitando tamiz tras tamiz) — el setTimeout deja que Streamlit primero registre el
# valor tecleado (su propio manejador de Enter) antes de mover el foco.
# También hace que el gesto de "atrás" (botón del navegador, o el gesto nativo de Android)
# funcione igual que el botón "← Atrás" de la app — ver _push_history_entry() más abajo, que es
# quien realmente dispara el pushState (justo después de que navigate()/go_back() terminan de
# sincronizar la URL). Este bloque solo pone el listener de popstate: al presionar atrás, dispara
# un recargo completo de la página, que Python resuelve leyendo la URL ya restaurada por el
# navegador (init_state) — por eso esto necesita ir de la mano con la restauración de sesión por
# cookie, si no cada "atrás" mandaría de vuelta al login.
components.html("""
<script>
(function() {
    if (window.parent.__geodeltaInputModeInit) return;
    window.parent.__geodeltaInputModeInit = true;

    function applyInputMode() {
        var inputs = window.parent.document.querySelectorAll(
            'div[data-testid="stTextInput"] input[type="text"]'
        );
        inputs.forEach(function(input) {
            var ph = (input.getAttribute('placeholder') || '').trim();
            var mode = null;
            if (/^-?\\d+[.,]\\d+$/.test(ph)) {
                mode = 'decimal';
            } else if (/^\\d[\\d\\s]*$/.test(ph)) {
                mode = 'numeric';
            }
            if (mode) {
                if (input.getAttribute('inputmode') !== mode) input.setAttribute('inputmode', mode);
            } else if (input.hasAttribute('inputmode')) {
                input.removeAttribute('inputmode');
            }
        });
    }

    function focusNextOnEnter(e) {
        if (e.key !== 'Enter') return;
        var target = e.target;
        if (!target || !target.matches || !target.matches('div[data-testid="stTextInput"] input')) return;
        var inputs = Array.prototype.slice.call(
            window.parent.document.querySelectorAll('div[data-testid="stTextInput"] input')
        );
        var idx = inputs.indexOf(target);
        if (idx === -1 || idx === inputs.length - 1) return;
        var next = inputs[idx + 1];
        setTimeout(function() {
            next.focus();
            next.select();
        }, 0);
    }

    applyInputMode();
    new MutationObserver(applyInputMode).observe(window.parent.document.body, {childList: true, subtree: true});
    window.parent.document.addEventListener('keydown', focusNextOnEnter, true);
    window.parent.addEventListener('popstate', function() {
        window.parent.location.reload();
    });
})();
</script>
""", height=0)

# ════════════════════════════════════════════════════════════════════
# CONSTANTES DEL DOMINIO
# ════════════════════════════════════════════════════════════════════
SIEVES = [
    ("s_3", '3"', "76.2", "E20"), ("s_2p5", '2 1/2"', "63.5", "E21"), ("s_2", '2"', "50.8", "E22"),
    ("s_1p5", '1 1/2"', "38.1", "E23"), ("s_1", '1"', "25.4", "E24"), ("s_34", '3/4"', "19.05", "E25"),
    ("s_12", '1/2"', "12.7", "E26"), ("s_38", '3/8"', "9.52", "E27"), ("s_4", "No. 4", "4.76", "E28"),
    ("s_10", "No. 10", "2.00", "E29"), ("s_20", "No. 20", "0.841", "E30"), ("s_40", "No. 40", "0.42", "E31"),
    ("s_60", "No. 60", "0.25", "E32"), ("s_100", "No. 100", "0.149", "E33"), ("s_200", "No. 200", "0.075", "E34"),
]

ASSAY_LABELS = {"granulometria": "Granulometría", "humedad": "Contenido de humedad", "masa-unitaria": "Peso unitario", "limites": "Límites de Atterberg", "pasa200": "Pasa 200", "cbr": "CBR"}
NORMAS_ENSAYO = {
    "granulometria": ["INV-214-13", "INV.E-213-13", "INV.E 123-13"],
    "humedad": ["INV E-122", "ASTM D2216"],
    "masa-unitaria": ["INV E-202", "ASTM D1188"],
    "cbr": ["INV E-148", "ASTM D1883"],
}
STATUS_LABELS = {"sin-iniciar": "Sin iniciar", "en-proceso": "En proceso", "finalizado": "Finalizado"}
STATUS_BADGE = {"sin-iniciar": "badge-danger", "en-proceso": "badge-warning", "finalizado": "badge-success"}
STATUS_ICON = {"sin-iniciar": "radio_button_unchecked", "en-proceso": "autorenew", "finalizado": "check_circle"}

TIPO_PERFORACION_PREFIX = {"Sondeo": "S", "Apique": "AP", "Fuente/Cantera": "F"}
# Texto que espera la lista desplegable de "tipo de perforación" en la plantilla
# CLASIFICACION_DE_SUELOS.xlsm (celda D12, validada contra AG6:AG10: SONDEO/APIQUE/TRINCHERA/NQ/N.A.).
TIPO_PERFORACION_EXCEL = {"Sondeo": "SONDEO", "Apique": "APIQUE", "Fuente/Cantera": "CANTERA"}
TIPO_MUESTRA_OPTIONS = ["Shelby", "NQ", "SS", "N/A"]
NORMA_PROYECTO_OPTIONS = ["IDU", "NTC", "INVIAS", "Otro"]

# Lista de equipos del laboratorio. Por ahora sin código — agrega o edita los que tengas aquí.
EQUIPO_LIST = [
    "Balanza digital 0.01g", "Balanza digital 0.1g", "Horno de secado", "Tamices serie gruesa",
    "Tamices serie fina", "Tamizadora mecánica", "Cazuela de Casagrande", "Ranurador", "Copa de Casagrande",
    "Molde Proctor estándar", "Molde Proctor modificado", "Prensa CBR", "Balanza hidrostática",
    "Horno de parafinado", "Cronómetro", "Termómetro", "Extractor de muestras", "Otro",
]

# Equipos reales de laboratorio con su código interno, tal como aparecen en el formato físico
# "EQUIPOS UTILIZADOS" para el ensayo de Granulometría (incluye el lavado por Tamiz No. 200).
EQUIPO_GRANULOMETRIA = [
    "Balanza GDA-E-010", "Balanza GDA-E-011", "Balanza GDA-E-012",
    "Horno GDA-E-007", "Horno GDA-E-404",
    "Tamiz de lavado GDA-E-", "Serie de tamices GDA-E-030 a GDA-E-045",
]

# Equipos reales usados en el ensayo de Contenido de Humedad Natural.
EQUIPO_HUMEDAD = ["Balanza GDA-E-010", "Balanza GDA-E-011", "Balanza GDA-E-014", "Horno GDA-E-007", "Horno GDA-E-404"]

# Método del ensayo de humedad (INV E-122), tal como aparece en la plantilla oficial (celda C28).
METODO_HUMEDAD = ["Método A", "Método B"]

# Equipos reales usados en el ensayo de Límites de Atterberg (Líquido + Plástico).
EQUIPO_LIMITES = [
    "Balanza GDA-E-012", "Balanza GDA-E-011", "Balanza GDA-E-010",
    "Cazuela Casagrande GDA-E-081", "Cazuela Casagrande GDA-E-060", "Cazuela Casagrande GDA-E-400",
    "Horno GDA-E-404", "Horno GDA-E-007", "Tamiz No. 40 GDA-E-054",
]

# Equipos reales usados en el ensayo de Peso Unitario Parafinado.
EQUIPO_MASA_UNITARIA = ["Balanza GDA-E-011", "Termómetro GDA-E-126"]

# CBR (INV E-148 / ASTM D1883) — todavía sin código interno asignado (a diferencia de los otros
# EQUIPO_* de arriba); se deja con nombre genérico igual que EQUIPO_LIST hasta que se tenga el
# código real de cada equipo.
EQUIPO_CBR = [
    "Prensa CBR", "Molde CBR", "Disco espaciador", "Pesas de sobrecarga",
    "Trípode con extensómetro", "Balanza digital 0.01g", "Balanza digital 0.1g",
    "Horno de secado", "Cronómetro",
]

# ════════════════════════════════════════════════════════════════════
# DESCRIPCIÓN VISUAL ESTRUCTURADA (menús desplegables en vez de texto libre) — para poder
# comparar de un vistazo la clasificación que hace el laboratorista a ojo con la clasificación
# USCS/AASHTO que calcula la app a partir de los datos de Granulometría/Límites (ver
# clasificar_uscs/clasificar_aashto). Todas las opciones van en MAYÚSCULA (así se escribe en el
# informe oficial) y siguen la nomenclatura estándar de descripción visual-manual de suelos
# (INV E-102 / ASTM D2488), no una lista inventada. No hay campo de texto libre — la frase final
# se arma solo con lo elegido en estos menús (ver descripcion_visual_estructurada).
DESC_TIPO_SUELO_OPTIONS = ["", "GRAVA", "ARENA", "LIMO", "ARCILLA", "ORGÁNICO", "OTROS"]
# Componente secundario ("grava CON ALGO DE arena", "arcilla CON ALGO DE arena") — mismas
# opciones que el tipo principal (sin la casilla en blanco, se filtra el tipo principal ya
# elegido para no dejar escoger "grava con algo de grava"). Es opcional: la casilla
# "¿tiene un componente secundario?" decide si se muestra o no.
DESC_TIPO_SECUNDARIO_OPTIONS = ["GRAVA", "ARENA", "LIMO", "ARCILLA", "ORGÁNICO", "OTROS"]
# Color principal y subtonalidad van separados (antes venían mezclados como "Café oscuro"): el
# color base es el matiz dominante, la subtonalidad es el matiz que lo modifica (ej. "MARRÓN
# ROJIZO", "GRIS AMARILLENTO") — así se puede combinar cualquier color con cualquier matiz en vez
# de tener que enumerar cada combinación como una opción aparte.
DESC_COLOR_OPTIONS = ["", "GRIS", "MARRÓN", "AMARILLO", "ROJO", "NEGRO", "BLANCO", "BEIGE", "NARANJA", "VERDE"]
DESC_SUBTONALIDAD_OPTIONS = ["", "CLARO", "OSCURO", "ROJIZO", "AMARILLENTO", "VERDOSO", "GRISÁCEO", "BLANQUECINO"]
# Suelos gruesos (grava/arena) se describen por cementación; suelos finos (limo/arcilla/orgánico)
# por consistencia — son dos propiedades distintas, no dos escalas de lo mismo, por eso solo se
# despliega una u otra según el tipo de grano elegido (ver _es_grueso).
DESC_CEMENTACION_OPTIONS = ["", "DÉBIL", "MODERADA", "FUERTE"]
DESC_CONSISTENCIA_OPTIONS = ["", "MUY BLANDA", "BLANDA", "FIRME", "DURA", "MUY DURA"]
# Forma solo aplica a grava (partículas lo bastante grandes para juzgar su forma a ojo);
# angulosidad aplica a cualquier suelo grueso (grava o arena).
DESC_FORMA_OPTIONS = ["", "PLANAS", "ALARGADAS", "PLANAS Y ALARGADAS"]
DESC_ANGULOSIDAD_OPTIONS = ["", "ANGULOSA", "SUB ANGULOSA", "SUB REDONDEADA", "REDONDEADA"]
DESC_HUMEDAD_OPTIONS = ["", "SECA", "HÚMEDA", "SATURADA"]


def _es_grueso(tipo_suelo):
    """Grava/Arena = grano grueso (cementación, angulosidad); Limo/Arcilla/Orgánico = grano fino
    (consistencia). Ver INV E-102."""
    return tipo_suelo in ("GRAVA", "ARENA")


def descripcion_visual_estructurada(muestra, tipo_override=None):
    """Arma la frase legible en mayúscula (ej. 'LIMO DE COLOR MARRÓN ROJIZO DE CONSISTENCIA DURA
    EN CONDICIÓN SECA') a partir de los menús desplegables de la muestra — se usa tanto en la
    vista de solo lectura como en el Excel oficial. None si todavía no se ha elegido ninguna
    opción.

    `tipo_override`: reemplaza SOLO la palabra inicial (ej. "GRAVA ARCILLOSA" en vez de "GRAVA")
    sin tocar el resto de la frase — para armar la versión "según la clasificación calculada" (ver
    descripcion_visual_calculada) sin duplicar toda esta lógica. El resto de la frase (forma,
    angulosidad, cementación o consistencia) se sigue decidiendo con el tipo de grano que
    realmente eligió el laboratorista a ojo, no con el override: son justo los campos que él
    mismo llenó bajo ese tipo, y cambiar de escala a mitad de la frase dejaría campos vacíos."""
    # .upper(): datos de antes de este cambio (migración 0015) se guardaron en minúscula/mixta
    # (ej. "Limo", "Café oscuro") — se normalizan al leer para que la frase salga toda en
    # mayúscula igual que los datos nuevos, sin tener que migrar filas viejas en la base de datos.
    tipo = (muestra.get("desc_tipo_suelo") or "").upper() or None
    partes = [tipo_override or tipo] if (tipo_override or tipo) else []

    # Componente secundario (ej. "GRAVA CON ALGO DE ARENA", "ARCILLA CON ALGO DE ARENA") — un
    # segundo tipo de grano que también está presente en la muestra, sin ser el dominante.
    secundario = (muestra.get("desc_tipo_secundario") or "").upper() or None
    if secundario:
        partes.append(f"CON ALGO DE {secundario}")

    angulosidad = (muestra.get("desc_angulosidad") or "").upper() if _es_grueso(tipo) else None
    if angulosidad:
        partes.append(angulosidad)
    forma = (muestra.get("desc_forma") or "").upper() if tipo == "GRAVA" else None
    if forma:
        partes.append(f"DE FORMA {forma}")

    color = (muestra.get("desc_color") or "").upper() or None
    subtonalidad = (muestra.get("desc_subtonalidad") or "").upper() or None
    if color:
        partes.append(f"DE COLOR {color}" + (f" {subtonalidad}" if subtonalidad else ""))
    elif subtonalidad:
        partes.append(f"DE COLOR {subtonalidad}")

    if _es_grueso(tipo):
        cementacion = (muestra.get("desc_cementacion") or "").upper() or None
        if cementacion:
            partes.append(f"CON CEMENTACIÓN {cementacion}")
    else:
        consistencia = (muestra.get("desc_consistencia") or "").upper() or None
        if consistencia:
            partes.append(f"DE CONSISTENCIA {consistencia}")

    # "CONDICIÓN {ESTADO}", no "CONDICIÓN DE HUMEDAD {ESTADO}" — el estado (SECA/HÚMEDA/
    # SATURADA) ya funciona como adjetivo de "condición", "condición de humedad húmeda" suena
    # redundante.
    humedad = (muestra.get("desc_humedad") or "").upper() or None
    if humedad:
        partes.append(f"EN CONDICIÓN {humedad}")

    return " ".join(partes) if partes else None


def descripcion_visual_para_excel(muestra):
    """Texto que va al campo "DESCRIPCIÓN VISUAL" del Excel oficial: la frase armada de los
    menús desplegables — None si todavía no hay ninguna opción elegida, para que el llamador
    pueda caer al respaldo de siempre (observaciones del ensayo, o el tipo de muestra). Si la
    muestra viene de antes de quitar el campo de notas libres (ver migración 0016), esas notas
    viejas se siguen mostrando como respaldo en vez de perderse."""
    return descripcion_visual_estructurada(muestra) or (muestra.get("descripcion_visual") or "").strip() or None


def descripcion_visual_calculada(muestra):
    """Segunda versión de la descripción visual, con el tipo de suelo que de verdad salió en la
    clasificación USCS (calculada con los datos ya digitados de Granulometría/Límites, ver
    clasificar_uscs) en vez del tipo que el laboratorista eligió a ojo antes de tener esos datos
    — combinado con el resto de características, que sí siguen siendo juicio visual del
    laboratorista (color, cementación o consistencia, humedad...). Se muestra AL LADO de
    descripcion_visual_estructurada, no en su lugar — la inicial a ojo se conserva tal cual.
    None mientras no haya datos suficientes para calcular la USCS."""
    gran_assay = get_assay(muestra["id_unico"], "granulometria")
    if not gran_assay:
        return None
    lim_assay = get_assay(muestra["id_unico"], "limites")
    resultado = clasificar_uscs(gran_assay.get("data"), lim_assay.get("data") if lim_assay else None)
    simbolo = resultado.get("simbolo")
    if not simbolo:
        return None
    nombre = USCS_NOMBRES.get(simbolo, simbolo).upper()
    return descripcion_visual_estructurada(muestra, tipo_override=nombre)

# Filas de Límite Líquido (INV. E-125-13) y Límite Plástico (INV. E-126-13), con las celdas
# reales de la plantilla CLASIFICACION_DE_SUELOS.xlsm (sección "LIMITES DE ATTERBERG", a la
# derecha de la tabla de Granulometría en la hoja "MUESTRA"): 3 columnas de ensayo para el
# Líquido (S/T/U) y 2 para el Plástico (Y/Z) — el propio Excel calcula LL, LP e IP.
LIMITE_LIQUIDO_FILAS = [
    ("lim_ll_recipiente", "Recipiente No.", ["S20", "T20", "U20"]),
    ("lim_ll_golpes", "No. de Golpes", ["S19", "T19", "U19"]),
    ("lim_ll_humedo", "Masa suelo húmedo + rec. (g)", ["S21", "T21", "U21"]),
    ("lim_ll_seco", "Masa suelo seco + rec. (g)", ["S22", "T22", "U22"]),
    # Las lecturas a 14/15/16 horas son solo un dato de apoyo/verificación del laboratorista —
    # no tienen celda propia en el Excel (la plantilla solo usa "lim_ll_seco").
    ("lim_ll_seco_14h", "Masa suelo seco + rec. (g) (14 hrs)", []),
    ("lim_ll_seco_15h", "Masa suelo seco + rec. (g) (15 hrs)", []),
    ("lim_ll_seco_16h", "Masa suelo seco + rec. (g) (16 hrs)", []),
    ("lim_ll_recip_masa", "Masa recipiente (g)", ["S23", "T23", "U23"]),
]
LIMITE_LIQUIDO_N = 3

LIMITE_PLASTICO_FILAS = [
    ("lim_lp_recipiente", "Recipiente No.", ["Y20", "Z20"]),
    ("lim_lp_humedo", "Masa suelo húmedo + rec. (g)", ["Y21", "Z21"]),
    ("lim_lp_seco", "Masa suelo seco + rec. (g)", ["Y22", "Z22"]),
    ("lim_lp_seco_14h", "Masa suelo seco + rec. (g) (14 hrs)", []),
    ("lim_lp_seco_15h", "Masa suelo seco + rec. (g) (15 hrs)", []),
    ("lim_lp_seco_16h", "Masa suelo seco + rec. (g) (16 hrs)", []),
    ("lim_lp_recip_masa", "Masa recipiente (g)", ["Y23", "Z23"]),
]
LIMITE_PLASTICO_N = 2

# "Pasa 200" sí aparece como ensayo aparte (se puede solicitar sin Granulometría), pero
# comparte plantilla y datos con Granulometría: si ambos se solicitan para la misma muestra,
# los dos leen y escriben el mismo diccionario de datos (ver render_assay_form), así que lo que
# se digite en cualquiera de los dos se refleja igual en ambos.
BITACORA_ENSAYOS = [
    "Granulometría", "Pasa 200", "Humedad", "Límites de Atterberg", "Límite de contracción",
    "Materia orgánica", "Proctor", "CBR", "Compresión inconfinada", "Compresión en roca",
    "Peso unitario", "Gravedad específica", "Consolidación", "Corte CD", "Corte CU", "Corte UU", "Otro",
]
SUPPORTED_ASSAY_MAP = {
    "Granulometría": "granulometria", "Humedad": "humedad", "Peso unitario": "masa-unitaria",
    "Límites de Atterberg": "limites", "Pasa 200": "pasa200", "CBR": "cbr",
}

BITACORA_BASE_COLS = ["Número", "Prof. De", "Prof. A", "Tipo de muestra"] + BITACORA_ENSAYOS + ["Observaciones"]

# Celdas de la plantilla oficial GDA-FL-003 (hoja "S1"): columna por ensayo, checkbox de
# norma y checkbox de tipo de perforación. Ensayos sin columna en la plantilla (p. ej.
# Carga puntual, Desgaste) simplemente no se marcan.
BITACORA_XLSX_ENSAYO_COL = {
    "Granulometría": "F", "Pasa 200": "G", "Humedad": "H", "Límites de Atterberg": "I",
    "Límite de contracción": "J", "Materia orgánica": "K", "Proctor": "L", "CBR": "M",
    "Compresión inconfinada": "N", "Compresión en roca": "O", "Peso unitario": "P",
    "Gravedad específica": "Q", "Consolidación": "R", "Corte CD": "S", "Corte CU": "T",
    "Corte UU": "U", "Otro": "AG",
}
BITACORA_XLSX_NORMA_CELL = {"IDU": "AG10", "INVIAS": "AI10", "NTC": "AG12", "Otro": "AI12"}
BITACORA_XLSX_TIPO_CELL = {"Sondeo": "H14", "Apique": "D14", "Fuente/Cantera": "AH14"}
BITACORA_XLSX_MAX_ROWS = 14  # la plantilla trae 14 filas fijas (18 a 31)


# ════════════════════════════════════════════════════════════════════
# CARGA DE DATOS DESDE SUPABASE (una vez por rerun, en el enrutador principal)
#
# El resto de la app sigue leyendo st.session_state.projects/perforaciones/
# muestras/assays exactamente con la misma forma que tenían en el store en
# memoria (perforaciones/muestras siguen agrupadas por código interno, los
# ensayos siguen usando el id_unico de la muestra como "muestra_id") — así
# que ninguno de los ~50 sitios de lectura repartidos por la app tuvo que
# cambiar. Lo único que cambia es de dónde viene el dato (Supabase, no un
# dict en memoria) y que ahora SÍ sobrevive un reinicio del servidor.
# Los sitios de ESCRITURA sí cambiaron: mutan vía db.py y luego hacen
# st.rerun(), que dispara una nueva llamada a _load_data() con el dato fresco.
# ════════════════════════════════════════════════════════════════════
def _load_data():
    projects = db.list_projects()
    proj_by_id = {p["id"]: p for p in projects}
    st.session_state.projects = projects

    perfs_raw = db.list_all_perforaciones()
    perf_by_id = {p["id"]: p for p in perfs_raw}
    perforaciones = {}
    for p in perfs_raw:
        proj = proj_by_id.get(p["project_id"])
        if proj:
            perforaciones.setdefault(proj["codigo_interno"], []).append(p)
    st.session_state.perforaciones = perforaciones

    muestras_raw = db.list_all_muestras()
    muestra_by_id = {m["id"]: m for m in muestras_raw}
    muestras = {}
    for m in muestras_raw:
        perf = perf_by_id.get(m["perforacion_id"])
        proj = proj_by_id.get(perf["project_id"]) if perf else None
        if perf and proj:
            key = f"{proj['codigo_interno']}::{perf['codigo']}"
            muestras.setdefault(key, []).append(m)
    st.session_state.muestras = muestras

    assays = []
    for a in db.list_all_assays():
        muestra = muestra_by_id.get(a["muestra_id"])
        perf = perf_by_id.get(muestra["perforacion_id"]) if muestra else None
        proj = proj_by_id.get(perf["project_id"]) if perf else None
        if not (muestra and perf and proj):
            continue
        a = dict(a)
        a["muestra_id"] = muestra["id_unico"]
        a["codigo_interno"] = proj["codigo_interno"]
        a["perforacion_codigo"] = perf["codigo"]
        a["muestra_numero"] = muestra["numero"]
        a["lastModified"] = a["updated_at"]
        a["createdAt"] = a["created_at"]
        assays.append(a)
    st.session_state.assays = assays

    notifications = db.list_notifications(st.session_state.role) if st.session_state.role else []
    for n in notifications:
        n["role"] = n["target_role"]
        n["muestra_id"] = n.get("muestra_id_unico")
    st.session_state.notifications = notifications


# ════════════════════════════════════════════════════════════════════
# ESTADO INICIAL
# ════════════════════════════════════════════════════════════════════
SESSION_COOKIE_MAX_AGE_DAYS = 30

# OJO: Streamlit Community Cloud filtra casi todas las cookies en su capa de proxy antes de que
# lleguen al backend de la app — st.context.cookies (lo que se usaba antes acá) da un dict vacío
# una vez desplegado, aunque funcione perfecto corriendo local. Por eso la sesión nunca se
# restauraba al recargar la página en producción, solo en las pruebas locales. CookieManager
# (extra_streamlit_components) evita el problema porque lee la cookie con JS en el navegador y
# se la manda a Python como el valor de un componente — nunca pasa por el proxy del servidor.
# NO envolver esto en @st.fragment: cuando el componente entrega su valor real de forma asíncrona,
# el rerun automático que dispara Streamlit quedaría acotado al fragmento y el resto del script
# (_try_restore_session_from_cookie, más abajo) nunca se enteraría del valor nuevo.
cookie_manager = stx.CookieManager(key="gdl_cookie_manager")


def _set_session_cookie(access_token, refresh_token):
    """Guarda los tokens de la sesión de Supabase en una cookie del navegador (nunca en la
    URL — la pantalla actual sí va en la URL, ver _sync_query_params, pero el token no) para
    poder restaurar el login solo tras un recargo de página o una reconexión, en vez de
    mandar siempre a la persona de vuelta a pedirle código+clave (ver init_state)."""
    expira = datetime.now() + timedelta(days=SESSION_COOKIE_MAX_AGE_DAYS)
    cookie_manager.batch_set({"gdl_at": access_token, "gdl_rt": refresh_token}, expires_at=expira)


def _delete_cookie_safe(nombre, key):
    """cookie_manager.delete() hace `del self.cookies[nombre]` sin verificar que exista, y
    truena con KeyError si ya no está (p. ej. la cookie nunca llegó a existir en este navegador,
    o el componente todavía no había entregado su valor real cuando se llamó) — se verifica antes
    para que cerrar sesión no reviente si no hay nada que borrar."""
    if nombre in cookie_manager.cookies:
        cookie_manager.delete(nombre, key=key)


def _clear_session_cookie():
    _delete_cookie_safe("gdl_at", key="del_gdl_at")
    _delete_cookie_safe("gdl_rt", key="del_gdl_rt")


def _set_remember_user_cookie(codigo):
    """Guarda el código de usuario (no la clave) en una cookie aparte de la de sesión, para
    precargar el campo "Código de usuario" del login la próxima vez que haga falta iniciar
    sesión — checkbox "Recordar mi usuario" en render_login()."""
    expira = datetime.now() + timedelta(days=SESSION_COOKIE_MAX_AGE_DAYS)
    cookie_manager.set("gdl_user", codigo, key="set_gdl_user", expires_at=expira)


def _clear_remember_user_cookie():
    _delete_cookie_safe("gdl_user", key="del_gdl_user")


def _push_history_entry():
    """Convierte la navegación que acaba de terminar en una entrada de historial de verdad
    (ver navigate()/go_back() y el router principal, que llaman a esto en el rerun SIGUIENTE
    al que cambió de pantalla — no en el mismo, por la misma razón que _set_session_cookie no
    se llama justo antes de un st.rerun(): el iframe no alcanza a montarse y correr su script
    antes de que el próximo rerun lo reemplace). Para cuando esto corre, _sync_query_params()
    ya terminó de aplicar sus 6 asignaciones en un rerun previo y completo, así que la URL del
    navegador ya es la definitiva — no hace falta ninguna espera ni detección de estabilidad.
    Se vio en pruebas que esto se termina llamando 2 veces por una sola navegación (probable
    doble rerun del click del botón) — se compara contra la última URL empujada para no dejar
    una entrada de historial duplicada, que obligaría a presionar "atrás" dos veces seguidas
    para moverse una sola pantalla."""
    components.html("""
    <script>
    (function() {
        var href = window.parent.location.href;
        if (window.parent.__geodeltaLastPushedHref === href) return;
        window.parent.__geodeltaLastPushedHref = href;
        window.parent.history.pushState({geodelta: true}, '', href);
    })();
    </script>
    """, height=0)


def _sync_query_params():
    """Guarda la pantalla actual en la URL (nunca el token de sesión — eso vive en una
    cookie aparte, ver _set_session_cookie). Al cerrar sesión se limpia todo, así que la
    sesión solo termina cuando el usuario le da a "Cerrar sesión" o la cookie vence (30 días).
    Se actualizan los 6 parámetros de una sola vez con .update(): asignarlos uno por uno
    (st.query_params["x"] = y) manda un ForwardMsg — y dispara una re-sincronización del
    navegador — por cada asignación; con 6 asignaciones seguidas eso se notó como hasta 7
    entradas de historial por cada navegación (ver _push_history_entry). .update() está
    documentado en el propio Streamlit para mandar un solo mensaje."""
    if st.session_state.role:
        st.query_params.update({
            "screen": st.session_state.screen,
            "codigo": st.session_state.selected_codigo or "",
            "perf": st.session_state.selected_perforacion or "",
            "muestra": st.session_state.selected_muestra_id or "",
            "assay": st.session_state.selected_assay_id or "",
            "atipo": st.session_state.selected_assay_type or "",
        })
    else:
        st.query_params.clear()


def init_state():
    if "initialized" in st.session_state:
        return
    st.session_state.initialized = True
    st.session_state.role = None
    st.session_state.profile = None
    st.session_state.screen = "home"

    st.session_state.projects = []
    st.session_state.perforaciones = {}
    st.session_state.muestras = {}
    st.session_state.assays = []
    st.session_state.notifications = []

    st.session_state.nav_stack = []
    st.session_state.bitacora_draft = {}
    st.session_state.draft_perforaciones = []
    st.session_state.draft_muestras = {}
    st.session_state.selected_codigo = ""
    st.session_state.selected_perforacion = ""
    st.session_state.selected_muestra_id = ""
    st.session_state.selected_assay_id = None
    st.session_state.selected_assay_type = None
    st.session_state.read_only_view = False

    # Restaura la posición de navegación desde la URL tras un recargo o reconexión (ver
    # _sync_query_params) — la persona vuelve a la misma pantalla en vez de a Inicio.
    st.session_state.screen = st.query_params.get("screen") or "home"
    st.session_state.selected_codigo = st.query_params.get("codigo") or ""
    st.session_state.selected_perforacion = st.query_params.get("perf") or ""
    st.session_state.selected_muestra_id = st.query_params.get("muestra") or ""
    st.session_state.selected_assay_id = st.query_params.get("assay") or None
    st.session_state.selected_assay_type = st.query_params.get("atipo") or None


def _try_restore_session_from_cookie():
    """Restaura el login desde la cookie del navegador (ver _set_session_cookie) — antes,
    cualquier recargo de página (F5, reconexión) mandaba de vuelta al login aunque la sesión de
    Supabase siguiera siendo válida. Si el refresh_token ya no sirve (venció, se cerró sesión en
    otro dispositivo), restore_session devuelve None y simplemente se queda en la pantalla de
    login, como antes.

    OJO: a diferencia del resto de init_state(), esto NO puede correr una sola vez al principio
    de la sesión — CookieManager lee la cookie de forma asíncrona con JS, así que en el primer
    rerun todavía no tiene el valor real (solo un default vacío) y recién lo entrega un par de
    reruns después. Por eso esta función se llama en CADA rerun mientras no haya sesión, hasta
    que la cookie real llegue. Una vez se intenta con un token real (funcione o no), no se
    reintenta más en esta sesión de navegador, para no golpear a Supabase en cada rerun si de
    verdad venció."""
    if st.session_state.role is not None or st.session_state.get("_cookie_restore_attempted"):
        return
    at = cookie_manager.get("gdl_at")
    rt = cookie_manager.get("gdl_rt")
    if not (at and rt):
        return
    st.session_state["_cookie_restore_attempted"] = True
    profile = db.restore_session(at, rt)
    if profile:
        st.session_state.profile = profile
        st.session_state.role = profile["role"]
        # restore_session() refresca el token si el access_token de la cookie ya había
        # vencido — Supabase rota el refresh_token en ese refresh, así que hay que
        # reescribir la cookie con los tokens nuevos (ver _tokens_rotated en el router
        # principal) o el siguiente recargo fallaría con un refresh_token ya inválido.
        st.session_state["_tokens_rotated"] = True


init_state()
_try_restore_session_from_cookie()


def navigate(screen):
    actual = st.session_state.get("screen")
    if actual and actual != screen:
        st.session_state.nav_stack.append(actual)
    st.session_state.screen = screen
    _sync_query_params()
    st.session_state._pending_history_push = True
    st.rerun()


def go_back(fallback="home"):
    """Vuelve a la pantalla realmente anterior (pila de navegación) en vez de un destino fijo."""
    if st.session_state.nav_stack:
        st.session_state.screen = st.session_state.nav_stack.pop()
    else:
        st.session_state.screen = fallback
    _sync_query_params()
    st.session_state._pending_history_push = True
    st.rerun()


def to_float(v, default=None):
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return default


def fmt_num(v, decimals=3):
    """Formatea un derivado numérico permitiendo hasta `decimals` decimales, sin ceros de más."""
    if v is None:
        return None
    s = f"{v:.{decimals}f}".rstrip("0").rstrip(".")
    return s if s not in ("", "-") else "0"


def calcular_humedad_pct(data):
    """% de humedad a partir de los datos digitados en el ensayo de Humedad, con la misma
    fórmula que trae la plantilla oficial (GDA-FLC-014, celda I24):
    (masa húmeda - masa seca) / (masa seca - masa recipiente) * 100."""
    masa_humedo = to_float(data.get("hum_masa_humedo_mas_recipiente"))
    masa_seco = to_float(data.get("hum_seco_mas_recipiente"))
    masa_recip = to_float(data.get("hum_masa_recipiente"))
    if masa_humedo is None or masa_seco is None or masa_recip is None:
        return None
    masa_suelo_seco = masa_seco - masa_recip
    if not masa_suelo_seco:
        return None
    return (masa_humedo - masa_seco) / masa_suelo_seco * 100


def icon(name, size=18, fill=False, color=None):
    """Ícono de Material Symbols para insertar dentro de HTML propio (st.markdown con unsafe_allow_html)."""
    cls = "material-symbols-outlined msi-fill" if fill else "material-symbols-outlined"
    style = f"font-size:{size}px;"
    if color:
        style += f"color:{color};"
    return f'<span class="{cls}" style="{style}">{name}</span>'


def status_badge_html(status, font_size=None):
    style = f' style="font-size:{font_size}px;"' if font_size else ""
    return (f'<span class="badge {STATUS_BADGE[status]}"{style}>'
            f'{icon(STATUS_ICON[status], size=13)} {STATUS_LABELS[status]}</span>')


def status_circle_html(status, size=20):
    circle_class = STATUS_BADGE[status].replace("badge-", "status-circle-")
    return (f'<span class="status-circle {circle_class}" title="{html.escape(STATUS_LABELS[status])}">'
            f'{icon(STATUS_ICON[status], size=size, fill=True)}</span>')


APROBACION_INFO = {
    None: ("No Confirmada", "badge-danger"),
    "pendiente_ing": ("En Proceso", "badge-warning"),
    "aprobado": ("Confirmada", "badge-success"),
}


def aprobacion_badge_html(etapa):
    """Estado de aprobación (Jefe → Director Técnico) de un ensayo individual — independiente
    del semáforo de laboratorio (sin-iniciar/en-proceso/finalizado)."""
    label, clase = APROBACION_INFO.get(etapa, APROBACION_INFO[None])
    return f'<span class="badge {clase}">DT: {label}</span>'


def card_header_html(icon_name, title, extra_html=""):
    """Encabezado de tarjeta con ícono + título (y opcionalmente un badge a la derecha),
    usado en las tarjetas de los formularios de ensayo (Norma, Equipos, Pasa 200, etc.)."""
    return (f'<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:14px;">'
            f'<div style="display:flex;align-items:center;gap:9px;font-weight:800;color:{PRIMARY};font-size:19px;">'
            f'{icon(icon_name, size=22)} {title}</div>{extra_html}</div>')


def param_table_html(rows, header_left="PARÁMETRO", header_right="VALOR REGISTRADO"):
    """Tabla de 2 columnas (etiqueta/valor) para la vista de solo lectura ('Resultados de Ensayo')."""
    body = "".join(
        f'<tr><td style="padding:10px 14px;border-bottom:1px solid {BORDER};color:{TEXT};">{html.escape(str(label))}</td>'
        f'<td style="padding:10px 14px;border-bottom:1px solid {BORDER};text-align:right;font-weight:600;color:{PRIMARY};">'
        f'{html.escape(str(value)) if value not in (None, "") else "—"}</td></tr>'
        for label, value in rows
    )
    return (f'<table style="width:100%;border-collapse:collapse;font-size:14px;">'
            f'<thead><tr style="background:{SECONDARY_CONTAINER};">'
            f'<th style="padding:10px 14px;text-align:left;font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">{header_left}</th>'
            f'<th style="padding:10px 14px;text-align:right;font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">{header_right}</th>'
            f'</tr></thead><tbody>{body}</tbody></table>')


def param_table_3col_html(rows, headers=("PARÁMETRO", "ANTES DEL LAVADO (g)", "DESPUÉS DEL LAVADO (g)")):
    """Tabla de 3 columnas (etiqueta + dos valores), usada para el Pasa No. 200 de Granulometría
    en la vista de solo lectura ('Resultados de Ensayo')."""
    def celda(v):
        return html.escape(str(v)) if v not in (None, "") else "—"
    body = "".join(
        f'<tr><td style="padding:10px 14px;border-bottom:1px solid {BORDER};color:{TEXT};">{html.escape(str(label))}</td>'
        f'<td style="padding:10px 14px;border-bottom:1px solid {BORDER};text-align:center;font-weight:600;color:{PRIMARY};">{celda(v1)}</td>'
        f'<td style="padding:10px 14px;border-bottom:1px solid {BORDER};text-align:center;font-weight:600;color:{PRIMARY};">{celda(v2)}</td></tr>'
        for label, v1, v2 in rows
    )
    head_left, head_mid, head_right = headers
    return (f'<table style="width:100%;border-collapse:collapse;font-size:14px;">'
            f'<thead><tr style="background:{SECONDARY_CONTAINER};">'
            f'<th style="padding:10px 14px;text-align:left;font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">{head_left}</th>'
            f'<th style="padding:10px 14px;text-align:center;font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">{head_mid}</th>'
            f'<th style="padding:10px 14px;text-align:center;font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">{head_right}</th>'
            f'</tr></thead><tbody>{body}</tbody></table>')


def param_table_ncol_html(headers, rows):
    """Tabla con cantidad arbitraria de columnas de valores (primer elemento de cada fila es la
    etiqueta, el resto son valores) — usada para Límite Líquido/Plástico, que tienen 3 y 2
    columnas de ensayo respectivamente."""
    def celda(v):
        return html.escape(str(v)) if v not in (None, "") else "—"
    body = "".join(
        '<tr>' + f'<td style="padding:10px 14px;border-bottom:1px solid {BORDER};color:{TEXT};">{html.escape(str(row[0]))}</td>'
        + "".join(f'<td style="padding:10px 14px;border-bottom:1px solid {BORDER};text-align:center;font-weight:600;color:{PRIMARY};">{celda(v)}</td>' for v in row[1:])
        + '</tr>'
        for row in rows
    )
    head_cells = "".join(
        f'<th style="padding:10px 14px;text-align:{"left" if i == 0 else "center"};font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">{h}</th>'
        for i, h in enumerate(headers)
    )
    return (f'<table style="width:100%;border-collapse:collapse;font-size:14px;">'
            f'<thead><tr style="background:{SECONDARY_CONTAINER};">{head_cells}</tr></thead><tbody>{body}</tbody></table>')


def condicion_table_html(muestra):
    """Tabla 'CONDICIÓN / TEMPERATURA °C / HUMEDAD %' para la vista de solo lectura."""
    def fila(cond_key, label):
        temp = muestra.get(f"cond_{cond_key}_temp") or "—"
        hum = muestra.get(f"cond_{cond_key}_hum") or "—"
        return (f'<tr><td style="padding:10px 14px;border-bottom:1px solid {BORDER};font-weight:600;color:{TEXT};">{label}</td>'
                f'<td style="padding:10px 14px;border-bottom:1px solid {BORDER};text-align:center;">{html.escape(str(temp))}</td>'
                f'<td style="padding:10px 14px;border-bottom:1px solid {BORDER};text-align:center;">{html.escape(str(hum))}</td></tr>')
    body = fila("inicial", "Inicial") + fila("final", "Final")
    return (f'<table style="width:100%;border-collapse:collapse;font-size:14px;">'
            f'<thead><tr style="background:{SECONDARY_CONTAINER};">'
            f'<th style="padding:10px 14px;text-align:left;font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">CONDICIÓN</th>'
            f'<th style="padding:10px 14px;text-align:center;font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">TEMPERATURA °C</th>'
            f'<th style="padding:10px 14px;text-align:center;font-size:11px;letter-spacing:0.04em;color:{PRIMARY};">HUMEDAD %</th>'
            f'</tr></thead><tbody>{body}</tbody></table>')


def split_equipo_codigo(equipo):
    """Separa 'Balanza GDA-E-011' en ('Balanza', 'GDA-E-011') para mostrarlo en dos líneas."""
    idx = equipo.find("GDA-E")
    if idx == -1:
        return equipo, ""
    return equipo[:idx].strip(), equipo[idx:].strip()


def equipos_readonly_html(equipos):
    """Lista de equipos utilizados (ícono + nombre + código) para la vista de solo lectura."""
    if not equipos:
        return f'<div class="cell-muted">Ningún equipo seleccionado.</div>'
    items = "".join(
        f'<div style="display:flex;align-items:center;gap:10px;padding:8px 0;border-bottom:1px solid {BORDER};">'
        f'{icon("construction", size=18, color=PRIMARY)}'
        f'<div><div style="font-weight:600;">{html.escape(nombre)}</div>'
        f'<div class="cell-muted" style="font-size:12px;">{html.escape(codigo) if codigo else "—"}</div></div></div>'
        for nombre, codigo in (split_equipo_codigo(e) for e in equipos)
    )
    return f'<div>{items}</div>'


def now_iso():
    return datetime.now().isoformat()


def add_notification(role, mensaje, codigo=None, perf=None, muestra_id=None):
    """Notificación compartida entre todas las sesiones logueadas con `role` (tabla
    notifications en Supabase — ver _load_data)."""
    db.add_notification(role, mensaje, codigo_interno=codigo, perforacion_codigo=perf, muestra_id_unico=muestra_id)


def add_historial(obj, titulo, subtitulo="", icono="history", tono="muted"):
    """Registro de auditoría de un ensayo: cuándo se entregó, cuándo confirmó el Jefe, cuándo
    aprobó (o devolvió) el Director Técnico, etc. Se muestra como línea de tiempo (ver
    historial_timeline_html). `obj` es el dict de assay tal como lo devuelve get_assay."""
    db.add_historial(obj["id"], titulo, subtitulo, icono, tono)


def historial_timeline_html(historial):
    items = sorted(historial, key=lambda h: h["fecha"], reverse=True)
    filas = []
    for i, h in enumerate(items):
        linea = '<div class="timeline-line"></div>' if i < len(items) - 1 else ""
        marcador = (f'<div class="timeline-marker-col">'
                    f'<div class="status-circle status-circle-{h.get("tono", "muted")}">'
                    f'{icon(h.get("icono", "history"), size=17, fill=True)}</div>{linea}</div>')
        # Compatibilidad con entradas antiguas (formato previo: solo {"fecha","texto"}) que
        # puedan seguir en el store en memoria de una sesión anterior al rediseño del historial.
        titulo = h.get("titulo") or h.get("texto") or "Cambio registrado"
        actor_html = (f'<div class="timeline-actor">{html.escape(h["subtitulo"])}</div>'
                      if h.get("subtitulo") else "")
        filas.append(
            f'<div class="timeline-item">{marcador}'
            f'<div class="timeline-content"><div class="timeline-titulo">{html.escape(titulo)}</div>'
            f'{actor_html}<div class="timeline-fecha">{format_dt(h["fecha"])}</div></div></div>'
        )
    return "".join(filas)


def format_dt(iso_str):
    try:
        return datetime.fromisoformat(iso_str).strftime("%d/%m/%Y %H:%M")
    except (ValueError, TypeError):
        return "—"


def require_role(*allowed):
    if st.session_state.role not in allowed:
        st.warning("No tienes permiso para ver esta sección.")
        if st.button("← Volver al inicio"):
            navigate("home")
        st.stop()


def get_project(codigo):
    return next((p for p in st.session_state.projects if p["codigo_interno"] == codigo), None)


@st.cache_data(ttl=30)
def list_laboratoristas():
    """Nombres de los laboratoristas activos, para el selector de asignación —
    reemplaza el texto libre de antes, que aceptaba cualquier cosa."""
    return sorted(
        p["full_name"] for p in db.list_profiles()
        if p["role"] == "laboratorista" and p.get("active", True)
    )


def get_muestra(codigo, perforacion_codigo, muestra_id):
    for m in st.session_state.muestras.get(f"{codigo}::{perforacion_codigo}", []):
        if m["id_unico"] == muestra_id:
            return m
    return None


def get_perforacion(codigo, perforacion_codigo):
    return next((p for p in st.session_state.perforaciones.get(codigo, []) if p["codigo"] == perforacion_codigo), None)


def get_assay(muestra_id, tipo_interno):
    return next((a for a in st.session_state.assays if a["muestra_id"] == muestra_id and a["tipo"] == tipo_interno), None)


def compute_muestra_estado(muestra):
    """El estado de la muestra se calcula solo, a partir del estado de cada ensayo solicitado."""
    statuses = []
    for label, activo in muestra["ensayos"].items():
        if not activo:
            continue
        tipo_interno = SUPPORTED_ASSAY_MAP.get(label)
        if not tipo_interno:
            continue
        a = get_assay(muestra["id_unico"], tipo_interno)
        statuses.append(a["status"] if a else "sin-iniciar")
    if not statuses:
        return "sin-iniciar"
    if all(s == "finalizado" for s in statuses):
        return "finalizado"
    if any(s in ("en-proceso", "finalizado") for s in statuses):
        return "en-proceso"
    return "sin-iniciar"


def project_progress(codigo):
    counts = {"sin-iniciar": 0, "en-proceso": 0, "finalizado": 0}
    for perf in st.session_state.perforaciones.get(codigo, []):
        for m in st.session_state.muestras.get(f"{codigo}::{perf['codigo']}", []):
            counts[compute_muestra_estado(m)] += 1
    return counts


def project_status(codigo):
    """'ejecutado' solo si el proyecto tiene al menos una muestra, TODAS están finalizadas
    (el laboratorista terminó) Y TODOS sus ensayos tienen el visto bueno final del Director
    Técnico (aprobación por ensayo individual) — no basta con que el laboratorio haya
    terminado, cada ensayo tiene que estar aprobado para poder entregarse."""
    counts = project_progress(codigo)
    total = sum(counts.values())
    if total == 0 or counts["finalizado"] != total:
        return "ejecucion"
    for a in st.session_state.assays:
        if a["codigo_interno"] == codigo and a.get("etapa_revision") != "aprobado":
            return "ejecucion"
    return "ejecutado"


def desarchivar_proyecto(codigo):
    """Reabre un proyecto ejecutado: revierte a 'en-proceso' todos sus ensayos finalizados,
    para que el laboratorista pueda volver a digitar o el Jefe agregar nuevas muestras/
    perforaciones. También limpia la aprobación de cada ensayo (etapa_revision y demás) —
    si se reabre, tiene que volver a pasar por Jefe y Director Técnico antes de poder archivarse
    de nuevo. El estado 'ejecutado' se recalcula solo (ver project_status). El llamador hace
    st.rerun() después, que recarga todo fresco desde Supabase (ver _load_data)."""
    for a in st.session_state.assays:
        if a["codigo_interno"] == codigo:
            db.reset_confirmacion(a["id"], reset_status=(a["status"] == "finalizado"))


def confirm_delete(action_key, label):
    """Botón de eliminar con confirmación en dos pasos. Devuelve True solo cuando se confirma."""
    flag = f"confirm_{action_key}"
    if st.session_state.get(flag):
        st.warning(f"¿Eliminar {label}? Esta acción no se puede deshacer.")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Sí, eliminar", key=f"yes_{action_key}", type="primary", use_container_width=True):
                st.session_state[flag] = False
                return True
        with c2:
            if st.button("Cancelar", key=f"no_{action_key}", use_container_width=True):
                st.session_state[flag] = False
                st.rerun()
        return False
    if st.button("Eliminar", key=f"del_{action_key}", use_container_width=True, icon=":material/delete:"):
        st.session_state[flag] = True
        st.rerun()
    return False


# ════════════════════════════════════════════════════════════════════
# LOGIN
# ════════════════════════════════════════════════════════════════════
def render_login():
    st.markdown("<br>", unsafe_allow_html=True)
    col = st.columns([1, 1.3, 1])[1]
    with col:
        st.markdown(f'<div class="login-icon">{icon("biotech", size=26)}</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-title">Geodelta Lab</div>', unsafe_allow_html=True)
        with st.container(border=True, key="login-card"):
            st.markdown("#### Bienvenido de nuevo")
            st.caption("Ingresa tu código de usuario y tu clave para acceder al sistema.")
            codigo_recordado = cookie_manager.get("gdl_user") or ""
            codigo = st.text_input("Código de usuario", value=codigo_recordado, placeholder="ej. jperez", autocomplete="off")
            password = st.text_input("Clave de acceso", type="password", placeholder="••••••••")
            recordar = st.checkbox("Recordar mi usuario", value=True)
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("INGRESAR", type="primary", use_container_width=True):
                if not codigo or not password:
                    st.error("Ingresa tu código de usuario y tu clave.")
                else:
                    try:
                        profile = db.sign_in(codigo, password)
                    except db.AuthError as e:
                        st.error(str(e))
                    else:
                        st.session_state.profile = profile
                        st.session_state.role = profile["role"]
                        # No se guarda la cookie aquí mismo: el st.rerun() de abajo corta la
                        # ejecución antes de que el iframe de components.html llegue a montarse
                        # y correr su script en el navegador (se probó y la cookie nunca quedaba
                        # puesta). Se guarda el token pendiente y se escribe la cookie en el
                        # siguiente rerun, cuando ya no hay un rerun inmediato después que lo corte.
                        st.session_state._pending_cookie_tokens = db.get_session_tokens()
                        st.session_state._pending_remember_user = codigo.strip() if recordar else ""
                        st.rerun()
            st.markdown('<hr style="margin:16px 0 4px 0;">', unsafe_allow_html=True)
            if st.button("¿Olvidaste tu clave?", key="forgot_pwd", type="secondary", use_container_width=True):
                st.info("Contacta al Jefe de laboratorio para restablecer tu clave de acceso.")
        st.markdown(f'<div class="login-footer">{icon("build", size=14)} Geodelta Lab Engineering</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════
# NAVEGACIÓN — TopAppBar + BottomNav (reemplaza el sidebar)
# ════════════════════════════════════════════════════════════════════
NAV_ITEMS = [
    ("home", "Inicio", "home"), ("projects-active", "Proyectos", "folder"), ("search", "Buscar", "search"),
]
ACTIVE_MAP = {
    "home": "home",
    "projects-active": "projects-active", "projects-done": "projects-active", "new-project": "projects-active",
    "project-detail": "projects-active", "edit-project": "projects-active",
    "perforacion-detail": "projects-active", "muestra-detail": "projects-active",
    "bitacora": "projects-active", "continue": "projects-active", "assay-form": "projects-active",
    "search": "search",
}


def render_topbar():
    active = ACTIVE_MAP.get(st.session_state.screen)
    with st.container(key="topbar"):
        c_brand, c_nav, c_bell, c_avatar, c_logout = st.columns([2.2, 4.2, 0.7, 0.7, 0.7])
        with c_brand:
            st.markdown(f'<div class="topbar-brand">{icon("biotech", size=24)}'
                        f'<span class="brand-title">Geodelta Lab</span></div>', unsafe_allow_html=True)
        with c_nav:
            with st.container(key="topbar-nav"):
                cols = st.columns(len(NAV_ITEMS))
                for col, (key, label, icono) in zip(cols, NAV_ITEMS):
                    with col:
                        if st.button(label, key=f"nav_{key}", use_container_width=True, icon=f":material/{icono}:",
                                     type="primary" if active == key else "secondary"):
                            navigate(key)
        with c_bell:
            mis_notifs = sorted(
                (n for n in st.session_state.notifications if n["role"] == st.session_state.role),
                key=lambda n: n["fecha"], reverse=True)
            no_leidas = sum(1 for n in mis_notifs if not n["leida"])
            with st.container(key="bell-alert" if no_leidas else "bell-quiet"):
                with st.popover(str(no_leidas) if no_leidas else "", icon=":material/notifications:", use_container_width=True):
                    st.markdown("**Notificaciones**")
                    if not mis_notifs:
                        st.caption("No tienes notificaciones.")
                    else:
                        with st.container(key="notif-popover-body"):
                            if no_leidas and st.button("Marcar todas como leídas", key="notif_marcar_todas", use_container_width=True):
                                db.mark_all_notifications_read(st.session_state.role)
                                st.rerun()
                            for n in mis_notifs[:15]:
                                with st.container(border=True, key=f"notif-card-{n['id']}"):
                                    estilo_msg = "font-weight:700;" if not n["leida"] else f"font-weight:400;font-size:13px;color:{MUTED};"
                                    st.markdown(f'<div style="{estilo_msg}">{html.escape(n["mensaje"])}</div>'
                                                f'<div class="timestamp-caption">{format_dt(n["fecha"])}</div>', unsafe_allow_html=True)
                                    if n.get("muestra_id") and st.button("Ir a la muestra →", key=f"notif_go_{n['id']}", use_container_width=True):
                                        db.mark_notification_read(n["id"])
                                        st.session_state.selected_codigo = n["codigo_interno"]
                                        st.session_state.selected_perforacion = n["perforacion_codigo"]
                                        st.session_state.selected_muestra_id = n["muestra_id"]
                                        navigate("muestra-detail")
        with c_avatar:
            iniciales = ROLE_INICIALES.get(st.session_state.role, "LB")
            st.markdown(f'<div class="topbar-avatar">{iniciales}</div>', unsafe_allow_html=True)
        with c_logout:
            if st.button("", key="logout_top", help="Cerrar sesión", use_container_width=True, icon=":material/logout:"):
                db.sign_out()
                st.session_state.role = None
                st.session_state.profile = None
                st.session_state.nav_stack = []
                st.session_state._pending_logout_cookie_clear = True
                navigate("home")


def render_bottomnav():
    active = ACTIVE_MAP.get(st.session_state.screen)
    with st.container(key="bottomnav"):
        cols = st.columns(len(NAV_ITEMS))
        for col, (key, label, icono) in zip(cols, NAV_ITEMS):
            with col:
                if st.button(label, key=f"bnav_{key}", use_container_width=True, icon=f":material/{icono}:",
                             type="primary" if active == key else "secondary"):
                    navigate(key)


# ════════════════════════════════════════════════════════════════════
# INICIO
# ════════════════════════════════════════════════════════════════════
def _ensayos_pendientes_dt():
    """Ensayos ya finalizados por el laboratorista, confirmados por el Jefe y esperando el
    visto bueno del Director Técnico (aprobación por ensayo individual). Devuelve tuplas
    (codigo, perf_codigo, muestra, ensayo_label)."""
    pendientes = []
    for p in st.session_state.projects:
        codigo = p["codigo_interno"]
        for perf in st.session_state.perforaciones.get(codigo, []):
            for m in st.session_state.muestras.get(f"{codigo}::{perf['codigo']}", []):
                for ensayo_label, activo in m["ensayos"].items():
                    if not activo or ensayo_label not in BITACORA_ENSAYOS:
                        continue
                    tipo_i = SUPPORTED_ASSAY_MAP.get(ensayo_label)
                    a = get_assay(m["id_unico"], tipo_i) if tipo_i else None
                    if a and a.get("etapa_revision") == "pendiente_ing":
                        pendientes.append((codigo, perf["codigo"], m, ensayo_label))
    return pendientes


def render_home():
    es_jefe = st.session_state.role == "jefe"
    es_ingeniero = st.session_state.role == "ingeniero"
    es_supervisor = es_jefe or es_ingeniero
    if es_jefe:
        st.markdown("## Bienvenido, Jefe de Laboratorio")
        st.caption("Resumen de operaciones y control de calidad geotécnica para hoy.")
    elif es_ingeniero:
        st.markdown("## Bienvenido, Director Técnico")
        st.caption("Revisión final y aprobación de muestras antes de entregarlas al cliente.")
    else:
        st.markdown("## Panel de Laboratorista")
        st.caption("Gestiona tus proyectos asignados y registra los resultados de los ensayos de suelo.")

    st.markdown("<br>", unsafe_allow_html=True)

    with st.container(key="home-actions"):
        if es_jefe:
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f'<div class="bento-primary"><div class="bento-icon">{icon("add")}</div>'
                             '<div><h3>Crear nuevo proyecto</h3><p>Registrar nuevo cliente y parámetros de sitio.</p></div></div>',
                             unsafe_allow_html=True)
                if st.button("Crear proyecto →", key="cta_new_project", use_container_width=True):
                    navigate("new-project")
            with c2:
                st.markdown(f'<div class="bento-light"><div class="bento-icon">{icon("sync")}</div>'
                             f'<div><h3>Proyectos en ejecución</h3><p>{sum(1 for p in st.session_state.projects if project_status(p["codigo_interno"])=="ejecucion")} proyecto(s) activos en laboratorio.</p></div></div>',
                             unsafe_allow_html=True)
                if st.button("Ver proyectos →", key="cta_active", use_container_width=True):
                    navigate("projects-active")
            with c3:
                st.markdown(f'<div class="bento-light"><div class="bento-icon">{icon("archive")}</div>'
                             '<div><h3>Proyectos ejecutados</h3><p>Revisar reportes finales y resultados certificados.</p></div></div>',
                             unsafe_allow_html=True)
                if st.button("Explorar archivo →", key="cta_done", use_container_width=True):
                    navigate("projects-done")
        elif es_ingeniero:
            pendientes_ing = _ensayos_pendientes_dt()
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f'<div class="bento-primary"><div class="bento-icon">{icon("fact_check")}</div>'
                             f'<div><span class="bento-eyebrow">Tareas prioritarias</span>'
                             f'<h3>Ensayos pendientes de tu aprobación</h3><p>{len(pendientes_ing)} ensayo(s) '
                             f'confirmados por el Jefe de Laboratorio, esperando tu visto bueno final.</p></div></div>',
                             unsafe_allow_html=True)
            with c2:
                st.markdown(f'<div class="bento-light"><div class="bento-icon">{icon("sync")}</div>'
                             f'<div><h3>Proyectos en ejecución</h3><p>{sum(1 for p in st.session_state.projects if project_status(p["codigo_interno"])=="ejecucion")} proyecto(s) activos en laboratorio.</p></div></div>',
                             unsafe_allow_html=True)
                if st.button("Ver proyectos →", key="cta_active_ing", use_container_width=True):
                    navigate("projects-active")
            with c3:
                st.markdown(f'<div class="bento-light"><div class="bento-icon">{icon("archive")}</div>'
                             '<div><h3>Proyectos ejecutados</h3><p>Consulta el historial certificado.</p></div></div>',
                             unsafe_allow_html=True)
                if st.button("Explorar archivo →", key="cta_done_ing", use_container_width=True):
                    navigate("projects-done")
            if pendientes_ing:
                st.markdown("<br>", unsafe_allow_html=True)
                for codigo, perf_codigo, m, ensayo_label in pendientes_ing[:5]:
                    proyecto = get_project(codigo)
                    with st.container(border=True):
                        cols = st.columns([3, 1])
                        cols[0].markdown(
                            f'<div class="cell-title">{html.escape(proyecto["nombre"] if proyecto else codigo)} · {html.escape(ensayo_label)}</div>'
                            f'<div class="cell-sub">{html.escape(codigo)} · {html.escape(perf_codigo)} · Muestra {m["numero"]}</div>',
                            unsafe_allow_html=True)
                        with cols[1]:
                            if st.button("Revisar →", key=f"revisar_ing_{m['id_unico']}_{ensayo_label}", use_container_width=True):
                                st.session_state.selected_codigo = codigo
                                st.session_state.selected_perforacion = perf_codigo
                                st.session_state.selected_muestra_id = m["id_unico"]
                                navigate("muestra-detail")
        else:
            c1, c2 = st.columns([2, 1])
            with c1:
                activos = sum(1 for p in st.session_state.projects if project_status(p["codigo_interno"]) == "ejecucion")
                st.markdown(f'<div class="bento-primary"><div class="bento-icon">{icon("assignment")}</div>'
                             f'<div><span class="bento-eyebrow">Tareas prioritarias</span>'
                             f'<h3>Proyectos en ejecución</h3><p>Accede a los proyectos activos para registrar granulometría, humedad y peso unitario.</p></div></div>',
                             unsafe_allow_html=True)
                if st.button(f"Ver proyectos → ({activos} activos)", key="cta_active_aux", use_container_width=True):
                    navigate("projects-active")
            with c2:
                st.markdown(f'<div class="bento-light"><div class="bento-icon">{icon("archive")}</div>'
                             '<div><h3>Proyectos ejecutados</h3><p>Consulta el historial. Solo lectura.</p></div></div>',
                             unsafe_allow_html=True)
                if st.button("Explorar archivo →", key="cta_done_aux", use_container_width=True):
                    navigate("projects-done")

    st.markdown("<br>", unsafe_allow_html=True)
    todos_los_ensayos = sorted(st.session_state.assays, key=lambda a: a["lastModified"], reverse=True)

    if es_supervisor:
        recientes = todos_los_ensayos[:5]
        with st.container(border=True):
            h1, h2 = st.columns([4, 1])
            with h1:
                st.markdown(f'<div class="section-title" style="border-bottom:none;margin-bottom:0;padding-bottom:0;">'
                            f'{icon("history", size=15)} Actividad reciente</div>', unsafe_allow_html=True)
            with h2:
                if st.button("Ver todo →", key="cta_ver_todo_actividad", use_container_width=True):
                    navigate("search")

            if not recientes:
                st.info("Todavía no hay actividad registrada.")
            else:
                col_ratios = [1.2, 2.0, 1.9, 1.5, 1.1, 0.9]
                headers = st.columns(col_ratios)
                for col, label in zip(headers, ["ID proyecto", "Nombre del proyecto", "Sondeo / Muestra", "Última actualización", "Estado", "Acción"]):
                    col.markdown(f'<div class="assigned-th">{label}</div>', unsafe_allow_html=True)
                for i, a in enumerate(recientes):
                    if i:
                        st.markdown(f'<hr style="margin:8px 0;border-color:{BORDER};">', unsafe_allow_html=True)
                    proyecto = get_project(a["codigo_interno"])
                    titulo = html.escape(proyecto["nombre"] if proyecto else a["codigo_interno"])
                    subtitulo = html.escape(f'{a["perforacion_codigo"]} · Muestra {a["muestra_numero"]} · {ASSAY_LABELS[a["tipo"]]}')
                    actualizacion = format_dt(a["lastModified"])
                    if a.get("laboratorist"):
                        actualizacion += f' · {html.escape(a["laboratorist"])}'
                    cols = st.columns(col_ratios, vertical_alignment="center")
                    cols[0].markdown(f'<span class="cell-id">{html.escape(a["codigo_interno"])}</span>', unsafe_allow_html=True)
                    cols[1].markdown(f'<div class="cell-title">{titulo}</div>', unsafe_allow_html=True)
                    cols[2].markdown(f'<div class="cell-sub">{subtitulo}</div>', unsafe_allow_html=True)
                    cols[3].markdown(f'<span class="cell-muted">{html.escape(actualizacion)}</span>', unsafe_allow_html=True)
                    with cols[4]:
                        st.markdown(f'<div style="text-align:center;">{status_circle_html(a["status"], size=16)}</div>', unsafe_allow_html=True)
                    with cols[5]:
                        if st.button("Abrir", key=f"open_recent_{a['id']}", use_container_width=True):
                            st.session_state.selected_codigo = a["codigo_interno"]
                            navigate("project-detail")
                st.markdown(f'<div class="activity-footer">Mostrando {len(recientes)} de {len(todos_los_ensayos)} ensayo(s)</div>',
                            unsafe_allow_html=True)
    else:
        # Cualquier laboratorista puede abrir cualquier ensayo pendiente — no hay asignación
        # previa del Jefe, cada quien indica su propio nombre en el campo "Laboratorista" del
        # formulario del ensayo al hacerlo. Por eso se recorren las muestras y su checklist de
        # ensayos directamente, en vez de filtrar st.session_state.assays (que solo trae ensayos
        # que YA tienen una fila creada, es decir que alguien ya abrió al menos una vez), para que
        # un ensayo recién marcado en la bitácora aparezca aquí de inmediato aunque nadie lo haya
        # abierto todavía.
        pendientes = []
        for key, muestras_perf in st.session_state.muestras.items():
            codigo_proyecto, perf_codigo = key.split("::", 1)
            for m in muestras_perf:
                for ensayo_label, marcado in m["ensayos"].items():
                    if not marcado:
                        continue
                    tipo_interno = SUPPORTED_ASSAY_MAP.get(ensayo_label)
                    if not tipo_interno:
                        continue  # sin formulario propio, no hay nada que "Abrir"
                    existing = get_assay(m["id_unico"], tipo_interno)
                    status = existing["status"] if existing else "sin-iniciar"
                    if status == "finalizado":
                        continue
                    pendientes.append({
                        "id": existing["id"] if existing else None,
                        "codigo_interno": codigo_proyecto, "perforacion_codigo": perf_codigo,
                        "muestra_id": m["id_unico"], "muestra_db_id": m["id"], "muestra_numero": m["numero"],
                        "tipo": tipo_interno, "status": status,
                        "lastModified": existing["lastModified"] if existing else m.get("updated_at", ""),
                    })
        pendientes.sort(key=lambda a: a["lastModified"], reverse=True)
        with st.container(border=True):
            h1, h2 = st.columns([4, 1])
            with h1:
                st.markdown(f'<div class="section-title" style="border-bottom:none;margin-bottom:0;padding-bottom:0;">'
                            f'{icon("assignment", size=15)} Ensayos pendientes</div>', unsafe_allow_html=True)
            with h2:
                st.markdown(f'<div style="text-align:right;"><span class="badge badge-muted">Total: {len(pendientes)}</span></div>',
                            unsafe_allow_html=True)

            if not pendientes:
                st.info("No hay ensayos pendientes por ahora.")
            else:
                col_ratios = [1.2, 2.0, 1.9, 1.5, 1.1, 0.9]
                headers = st.columns(col_ratios)
                for col, label in zip(headers, ["ID ensayo", "Proyecto", "Tipo de ensayo", "Última actualización", "Estado", "Acción"]):
                    col.markdown(f'<div class="assigned-th">{label}</div>', unsafe_allow_html=True)
                for i, a in enumerate(pendientes):
                    if i:
                        st.markdown(f'<hr style="margin:8px 0;border-color:{BORDER};">', unsafe_allow_html=True)
                    proyecto = get_project(a["codigo_interno"])
                    cols = st.columns(col_ratios, vertical_alignment="center")
                    ensayo_id = f'{a["codigo_interno"]}-{a["perforacion_codigo"]}-M{a["muestra_numero"]}'
                    cols[0].markdown(f'<span class="cell-id">{html.escape(ensayo_id)}</span>', unsafe_allow_html=True)
                    titulo = html.escape(proyecto["nombre"] if proyecto else a["codigo_interno"])
                    subtitulo = html.escape(proyecto.get("localizacion", "")) if proyecto else ""
                    cols[1].markdown(f'<div class="cell-title">{titulo}</div><div class="cell-sub">{subtitulo}</div>',
                                      unsafe_allow_html=True)
                    cols[2].markdown(f'<span class="assigned-chip">{ASSAY_LABELS[a["tipo"]]}</span>', unsafe_allow_html=True)
                    cols[3].markdown(f'<span class="cell-muted">{html.escape(format_dt(a["lastModified"]))}</span>',
                                      unsafe_allow_html=True)
                    with cols[4]:
                        st.markdown(f'<div style="text-align:center;">{status_circle_html(a["status"], size=16)}</div>', unsafe_allow_html=True)
                    with cols[5]:
                        if st.button("Abrir", key=f"open_assigned_{a['muestra_id']}_{a['tipo']}", use_container_width=True):
                            if a["id"]:
                                st.session_state.selected_assay_id = a["id"]
                            else:
                                nuevo = db.create_assay(a["muestra_db_id"], a["tipo"])
                                st.session_state.selected_assay_id = nuevo["id"]
                            st.session_state.selected_codigo = a["codigo_interno"]
                            st.session_state.selected_perforacion = a["perforacion_codigo"]
                            st.session_state.selected_muestra_id = a["muestra_id"]
                            st.session_state.selected_assay_type = a["tipo"]
                            navigate("assay-form")


def _render_project_list(codes, empty_msg, allow_delete, mark_read_only=False, allow_unarchive=False):
    if not codes:
        st.info(empty_msg)
        return
    for p in st.session_state.projects:
        if p["codigo_interno"] not in codes:
            continue
        counts = project_progress(p["codigo_interno"])
        with st.container(border=True):
            cols = st.columns([3, 2, 1, 1] if allow_delete else [3, 2, 1])
            with cols[0]:
                st.markdown(f"**{p['codigo_interno']}**")
                st.caption(p["nombre"])
            with cols[1]:
                st.markdown(f'<span class="cell-muted">{icon(STATUS_ICON["sin-iniciar"], size=14)} {counts["sin-iniciar"]}'
                            f'&nbsp;&nbsp;·&nbsp;&nbsp;{icon(STATUS_ICON["en-proceso"], size=14)} {counts["en-proceso"]}'
                            f'&nbsp;&nbsp;·&nbsp;&nbsp;{icon(STATUS_ICON["finalizado"], size=14)} {counts["finalizado"]}</span>',
                            unsafe_allow_html=True)
            with cols[2]:
                if st.button("Abrir", key=f"openlist_{p['codigo_interno']}", use_container_width=True):
                    st.session_state.selected_codigo = p["codigo_interno"]
                    navigate("project-detail")
            if allow_delete:
                with cols[3]:
                    if confirm_delete(f"project_{p['codigo_interno']}", f"el proyecto {p['codigo_interno']}"):
                        codigo = p["codigo_interno"]
                        db.archive_project(p["id"])
                        st.session_state.bitacora_draft = {k: v for k, v in st.session_state.bitacora_draft.items() if not k.startswith(codigo + "::")}
                        st.rerun()
            if allow_unarchive:
                if st.button("Desarchivar proyecto", icon=":material/unarchive:", key=f"unarchive_{p['codigo_interno']}", use_container_width=True):
                    desarchivar_proyecto(p["codigo_interno"])
                    st.rerun()


def _resumen_tecnico_perforaciones(codigo):
    """Una línea por perforación: sus muestras y los ensayos solicitados en ellas."""
    lineas = []
    for perf in st.session_state.perforaciones.get(codigo, []):
        muestras = st.session_state.muestras.get(f"{codigo}::{perf['codigo']}", [])
        if not muestras:
            lineas.append(f"<strong>{perf['codigo']}</strong>: sin muestras")
            continue
        ids = ", ".join(f"M-{m['numero']}" for m in muestras)
        ensayos = sorted({e for m in muestras for e, activo in m["ensayos"].items() if activo and e in BITACORA_ENSAYOS})
        linea = f"<strong>{perf['codigo']}</strong>: {ids}"
        if ensayos:
            linea += f" · {', '.join(ensayos)}"
        lineas.append(linea)
    return lineas


def render_projects_active():
    if st.button("← Atrás"):
        go_back()
    st.markdown("## Proyectos en ejecución")
    st.caption("Monitoreo técnico de sondeos y análisis geotécnico.")

    proyectos = [p for p in st.session_state.projects if project_status(p["codigo_interno"]) == "ejecucion"]

    en_ensayo = sum(1 for p in proyectos if project_progress(p["codigo_interno"])["en-proceso"] > 0)
    c1, c2 = st.columns(2)
    for col, icono, label, valor in [
        (c1, "folder", "Activos", len(proyectos)), (c2, "science", "En ensayo", en_ensayo),
    ]:
        with col:
            st.markdown(f'<div class="stat-chip"><div class="stat-icon">{icon(icono, size=20)}</div>'
                        f'<div><div class="stat-label">{label}</div><div class="stat-value">{valor}</div></div></div>',
                        unsafe_allow_html=True)

    busqueda = st.text_input("Buscar", placeholder="Buscar por código o nombre...", label_visibility="collapsed", icon=":material/search:")
    if busqueda:
        q = busqueda.lower()
        proyectos = [p for p in proyectos if q in p["codigo_interno"].lower() or q in p["nombre"].lower()]

    if st.session_state.role == "jefe":
        with st.container(key="fab-new-project"):
            if st.button("", icon=":material/add:", key="fab_new_project_btn", help="Crear nuevo proyecto"):
                navigate("new-project")

    if not proyectos:
        st.info("No hay proyectos en ejecución en este momento.")
        return

    for p in proyectos:
        codigo = p["codigo_interno"]
        counts = project_progress(codigo)
        total = sum(counts.values())
        if counts["en-proceso"] > 0:
            estado_badge, estado_label = "badge-warning", "En ensayo"
        elif total == 0:
            estado_badge, estado_label = "badge-muted", "Sin muestras"
        else:
            estado_badge, estado_label = "badge-muted", "Por iniciar"

        with st.container(border=True, key=f"projcard_{codigo}"):
            top = st.columns([3, 1])
            top[0].markdown(f'<span class="code-badge">{html.escape(codigo)}</span>', unsafe_allow_html=True)
            top[1].markdown(f'<div style="text-align:right;"><span class="badge {estado_badge}">{estado_label}</span></div>',
                             unsafe_allow_html=True)
            st.markdown(f"**{p['nombre']}**")
            st.markdown(f'<span class="cell-muted">{icon("location_on", size=14)} {html.escape(p.get("localizacion") or "—")}</span>',
                        unsafe_allow_html=True)
            st.markdown(f'<div class="section-title" style="margin-bottom:4px;">Norma</div>'
                        f'<div style="margin-bottom:10px;">{html.escape(p.get("norma") or "—")}</div>', unsafe_allow_html=True)
            m1, m2 = st.columns(2)
            m1.markdown(f'<span class="cell-muted">Fecha bitácora:</span><br><span style="font-weight:600;">'
                        f'{html.escape(p.get("fecha_bitacora") or "—")}</span>', unsafe_allow_html=True)
            m2.markdown(f'<span class="cell-muted">Fecha ingreso:</span><br><span style="font-weight:600;">'
                        f'{html.escape(p.get("fecha_ingreso_muestra") or "—")}</span>', unsafe_allow_html=True)
            resumen = _resumen_tecnico_perforaciones(codigo)
            if resumen:
                st.markdown(f'<div class="section-title" style="margin-bottom:6px;">'
                            f'Resumen técnico de perforaciones ({len(resumen)})</div>', unsafe_allow_html=True)
                for linea in resumen:
                    st.markdown(f'<div class="cell-sub" style="margin-bottom:4px;">{linea}</div>', unsafe_allow_html=True)
            if st.button("Ver proyecto →", key=f"veractivo_{codigo}", type="primary", use_container_width=True):
                st.session_state.selected_codigo = codigo
                navigate("project-detail")


def render_projects_done():
    if st.button("← Atrás"):
        go_back()
    st.markdown("## Proyectos ejecutados")
    if st.session_state.role == "laboratorista":
        st.info("Modo consulta: puedes ver los resultados, pero no editarlos.")
    codes = [p["codigo_interno"] for p in st.session_state.projects if project_status(p["codigo_interno"]) == "ejecutado"]
    _render_project_list(codes, "Todavía no hay proyectos completamente finalizados.",
                          allow_delete=(st.session_state.role == "jefe"), mark_read_only=True,
                          allow_unarchive=(st.session_state.role == "jefe"))


# ════════════════════════════════════════════════════════════════════
# NUEVO PROYECTO (solo Jefe)
# ════════════════════════════════════════════════════════════════════
def _str_excel(v):
    """str() de una celda de Excel, sin el '.0' feo que deja Python en números enteros
    guardados como float (típico en teléfonos digitados en una celda numérica)."""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v or "").strip()


def _fecha_excel(v):
    """Convierte el valor de una celda de fecha de Excel a un date de Python, o None si
    no se pudo leer como fecha (celda vacía, texto libre, etc.)."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _leer_bitacora_cliente_xlsx(file_obj):
    """Lee el formato GDA-FL-021 (Bitácora de Proyecto) que envía el cliente y devuelve
    los campos que sirven para precargar Nuevo Proyecto."""
    wb = load_workbook(file_obj, data_only=True)
    ws = wb["HOJA1"] if "HOJA1" in wb.sheetnames else wb.active
    return {
        "cliente": _str_excel(ws["E11"].value),
        "nombre": _str_excel(ws["E12"].value),
        "localizacion": _str_excel(ws["E13"].value),
        "direccion_cliente": _str_excel(ws["E14"].value),
        "telefono_contacto": _str_excel(ws["E15"].value),
        "correo_cliente": _str_excel(ws["Q15"].value),
        "nombre_contacto": _str_excel(ws["E16"].value),
        "fecha_inicio_proyecto": _fecha_excel(ws["E17"].value),
        "fecha_final_proyecto": _fecha_excel(ws["Q17"].value),
    }


def render_new_project():
    require_role("jefe")
    if st.button("← Atrás"):
        go_back()
    st.markdown("## Bitácora de proyecto")

    st.markdown('<div class="section-title">Código interno</div>', unsafe_allow_html=True)
    # Sugerencia de consecutivo: año actual (2 dígitos) y, dentro de los proyectos ya creados con
    # ese año, el número más alto + 1 — igual que armaría el consecutivo el Jefe a mano. Se
    # precargan como valor editable (no solo placeholder) para no tener que digitarlos siempre;
    # el Jefe puede borrarlos y poner otros si el proyecto es de un año distinto.
    anio_sugerido = str(date.today().year)[-2:]
    if "new_anio" not in st.session_state:
        st.session_state["new_anio"] = anio_sugerido
    if "new_numero" not in st.session_state:
        numeros_mismo_anio = [
            int(p["numero"]) for p in st.session_state.projects
            if str(p.get("anio", "")) == st.session_state["new_anio"] and str(p.get("numero", "")).isdigit()
        ]
        st.session_state["new_numero"] = f"{(max(numeros_mismo_anio) + 1):03d}" if numeros_mismo_anio else "001"

    c1, c2, c3 = st.columns([1, 1, 1])
    with c1:
        st.text_input("Prefijo", value="GDA", disabled=True, autocomplete="off")
    with c2:
        numero = st.text_input("Número", key="new_numero", autocomplete="off")
    with c3:
        anio = st.text_input("Año", key="new_anio", autocomplete="off")

    codigo_interno = f"GDA-{numero}-{anio}" if numero and anio else ""
    existing_codes = [p["codigo_interno"] for p in st.session_state.projects]
    codigo_valido = bool(codigo_interno) and codigo_interno not in existing_codes
    if codigo_interno:
        if not codigo_valido:
            st.error(f"El código **{codigo_interno}** ya existe.")
        else:
            st.success(f"Código interno: **{codigo_interno}**")

    st.markdown('<div class="section-title">Cargar bitácora de proyecto del cliente (opcional)</div>', unsafe_allow_html=True)
    uploaded_cliente_xlsx = st.file_uploader(
        "Bitácora de proyecto del cliente (Excel)", type=["xlsx"], key="cliente_xlsx_uploader",
        help="Si el cliente te envió el formato GDA-FL-021 (Bitácora de Proyecto), súbelo aquí para "
             "precargar Cliente, Nombre del proyecto, Localización, Dirección, Teléfono, Correo, "
             "Nombre de contacto y fechas de inicio/fin del proyecto.",
    )
    if uploaded_cliente_xlsx is not None and st.session_state.get("_cliente_xlsx_last") != uploaded_cliente_xlsx.name:
        try:
            datos_cliente = _leer_bitacora_cliente_xlsx(uploaded_cliente_xlsx)
            st.session_state["new_nombre"] = datos_cliente["nombre"]
            st.session_state["new_localizacion"] = datos_cliente["localizacion"]
            st.session_state["new_cliente"] = datos_cliente["cliente"]
            st.session_state["new_correo_cliente"] = datos_cliente["correo_cliente"]
            st.session_state["new_direccion_cliente"] = datos_cliente["direccion_cliente"]
            st.session_state["new_telefono_contacto"] = datos_cliente["telefono_contacto"]
            st.session_state["new_nombre_contacto"] = datos_cliente["nombre_contacto"]
            if datos_cliente["fecha_inicio_proyecto"]:
                st.session_state["new_fecha_inicio_proyecto"] = datos_cliente["fecha_inicio_proyecto"]
            if datos_cliente["fecha_final_proyecto"]:
                st.session_state["new_fecha_final_proyecto"] = datos_cliente["fecha_final_proyecto"]
            st.session_state["_cliente_xlsx_last"] = uploaded_cliente_xlsx.name
            st.success("Datos del cliente cargados desde el Excel. Revísalos abajo antes de guardar.")
        except Exception:
            st.error("No se pudo leer el archivo. Verifica que sea el formato GDA-FL-021 (Bitácora de Proyecto).")

    st.markdown('<div class="section-title">Información del proyecto</div>', unsafe_allow_html=True)
    nombre = st.text_input("Nombre del proyecto", key="new_nombre", placeholder="Estudio de suelos vía Bogotá-Medellín")
    localizacion = st.text_input("Localización", key="new_localizacion", placeholder="Km 14+200")
    norma = st.radio("Norma", NORMA_PROYECTO_OPTIONS, horizontal=True)

    c1, c2 = st.columns(2)
    with c1:
        fecha_bitacora = st.date_input("Fecha de bitácora", value=date.today(), format="DD/MM/YYYY")
    with c2:
        fecha_ingreso = st.date_input("Fecha de ingreso de muestra", value=date.today(), format="DD/MM/YYYY")

    ec1, ec2, ec3 = st.columns(3)
    with ec1:
        fecha_recepcion = st.date_input("Fecha de recepción", value=date.today(), format="DD/MM/YYYY")
    with ec2:
        fecha_ejecucion = st.date_input("Fecha de ejecución", value=date.today(), format="DD/MM/YYYY")
    with ec3:
        fecha_emision = st.date_input("Fecha de emisión", value=date.today(), format="DD/MM/YYYY")

    st.markdown('<div class="section-title">Datos del cliente (para el encabezado de los informes — solo el Jefe los ve)</div>', unsafe_allow_html=True)
    cliente = st.text_input("Cliente", key="new_cliente", placeholder="Nombre del cliente")
    direccion_cliente = st.text_input("Dirección cliente", key="new_direccion_cliente", placeholder="Dirección del cliente")
    dc1, dc2 = st.columns(2)
    with dc1:
        telefono_contacto = st.text_input("Teléfono de contacto", key="new_telefono_contacto", placeholder="300 000 0000")
    with dc2:
        correo_cliente = st.text_input("Correo electrónico", key="new_correo_cliente", placeholder="correo@cliente.com")
    nombre_contacto = st.text_input("Nombre de contacto", key="new_nombre_contacto", placeholder="Nombre de quien coordina con el cliente")
    muestra_tomada_por = st.text_input("Muestra tomada por", placeholder="Nombre de quien tomó la muestra")

    dc3, dc4 = st.columns(2)
    with dc3:
        if "new_fecha_inicio_proyecto" not in st.session_state:
            st.session_state["new_fecha_inicio_proyecto"] = date.today()
        fecha_inicio_proyecto = st.date_input("Fecha inicio proyecto", key="new_fecha_inicio_proyecto", format="DD/MM/YYYY")
    with dc4:
        if "new_fecha_final_proyecto" not in st.session_state:
            st.session_state["new_fecha_final_proyecto"] = date.today()
        fecha_final_proyecto = st.date_input("Fecha final proyecto", key="new_fecha_final_proyecto", format="DD/MM/YYYY")
    # A diferencia de las demás fechas, esta normalmente no se sabe todavía al crear el
    # proyecto (es cuándo terminó de verdad, no la fecha final planeada) — arranca vacía en
    # vez de con la fecha de hoy, y se puede completar después desde "Editar proyecto".
    fecha_final_real = st.date_input(
        "Fecha final real", key="new_fecha_final_real", format="DD/MM/YYYY", value=None,
        help="Cuándo terminó de verdad el proyecto — distinta de la fecha final planeada. "
             "Déjala vacía si todavía no ha terminado.",
    )

    # La asignación ya no es por proyecto entero — se asigna ensayo por ensayo desde el
    # detalle de cada muestra (ver render_muestra_detail), una vez que la bitácora existe.

    # Perforaciones y muestras se arman aquí mismo, antes de crear el proyecto formalmente —
    # en estado de borrador propio (no en st.session_state.perforaciones/muestras, que ahora
    # se recargan desde Supabase en cada rerun y no existen todavía para un proyecto sin crear).
    perforaciones = st.session_state.setdefault("draft_perforaciones", [])
    edited_frames = {}
    if codigo_valido and nombre:
        st.markdown('<div class="section-title">Perforación</div>', unsafe_allow_html=True)
        pc1, pc2 = st.columns([2, 1])
        with pc1:
            tipo = st.selectbox("Tipo de perforación", list(TIPO_PERFORACION_PREFIX.keys()))
        with pc2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Nueva perforación", use_container_width=True, icon=":material/add:"):
                prefix = TIPO_PERFORACION_PREFIX[tipo]
                consecutivo = len([p for p in perforaciones if p["tipo"] == tipo]) + 1
                codigo_perf = f"{prefix}{consecutivo}"
                perforaciones.append({"tipo": tipo, "consecutivo": consecutivo, "codigo": codigo_perf})
                st.session_state.setdefault("draft_muestras", {})[codigo_perf] = []
                st.rerun()

        if perforaciones:
            st.markdown('<div class="section-title">Perforaciones y muestras</div>', unsafe_allow_html=True)
        for perf in perforaciones:
            key = f"{codigo_interno}::{perf['codigo']}"
            muestras = st.session_state.draft_muestras.setdefault(perf["codigo"], [])
            with st.expander(f"**{perf['codigo']}** — {perf['tipo']}  ·  {len(muestras)} muestra(s)", expanded=True):
                df_source = _bitacora_draft_df(key, muestras)

                column_config = {
                    "Número": st.column_config.TextColumn(default=""),
                    "Prof. De": st.column_config.NumberColumn(default=0.0, step=0.01),
                    "Prof. A": st.column_config.NumberColumn(default=0.0, step=0.01),
                    "Tipo de muestra": st.column_config.SelectboxColumn(options=TIPO_MUESTRA_OPTIONS, default=TIPO_MUESTRA_OPTIONS[0]),
                }
                for e in BITACORA_ENSAYOS:
                    column_config[e] = st.column_config.CheckboxColumn(e, default=False)
                column_config["Observaciones"] = st.column_config.TextColumn(
                    default="", width="medium", help="Cómo llegó la muestra o cualquier condición que impida el ensayo.")

                st.caption("Usa el ícono para agregar fila sobre la tabla para sumar una muestra nueva. Para eliminar una, selecciona el cuadro a la izquierda de su fila y usa el ícono de basura que aparece sobre la tabla.")
                edited = st.data_editor(
                    df_source, num_rows="dynamic", use_container_width=True,
                    column_config=column_config, key=f"neweditor_{key}",
                )
                edited_frames[key] = edited

                # Cada perforación se descarga en su propio Excel: la plantilla oficial
                # representa UN sondeo (hoja "S1"), así que no se mezclan varias en un archivo.
                filas_perf_preview = []
                for row in edited.to_dict("records"):
                    numero_m = str(row.get("Número", "")).strip()
                    if not numero_m or numero_m.lower() == "none" or numero_m == "nan":
                        continue
                    filas_perf_preview.append({
                        "perf_codigo": perf["codigo"], "numero": numero_m,
                        "tipo_muestra": row.get("Tipo de muestra") or TIPO_MUESTRA_OPTIONS[0],
                        "profundidad_de": row.get("Prof. De") or 0.0, "profundidad_hasta": row.get("Prof. A") or 0.0,
                        "ensayos": {e: bool(row.get(e, False)) for e in BITACORA_ENSAYOS},
                        "observaciones": row.get("Observaciones") or "",
                    })
                project_preview = {
                    "codigo_interno": codigo_interno, "numero": numero, "anio": anio, "nombre": nombre,
                    "localizacion": localizacion, "norma": norma, "fecha_bitacora": str(fecha_bitacora),
                }
                excel_bytes, truncado = generar_excel_bitacora_orden(project_preview, filas_perf_preview, {perf["tipo"]})
                st.download_button(f"Descargar bitácora — {perf['codigo']}", data=excel_bytes, icon=":material/download:",
                                    file_name=f"{codigo_interno} Bitacora de orden {perf['codigo']}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True, key=f"dl_newperf_{key}")
                if truncado:
                    st.caption(f"El formato oficial admite hasta {BITACORA_XLSX_MAX_ROWS} muestras; se incluyeron las primeras {BITACORA_XLSX_MAX_ROWS}.")

                if confirm_delete(f"newperf_{key}", f"la perforación {perf['codigo']}"):
                    st.session_state.draft_perforaciones = [p for p in perforaciones if p["codigo"] != perf["codigo"]]
                    st.session_state.draft_muestras.pop(perf["codigo"], None)
                    st.session_state.bitacora_draft.pop(key, None)
                    st.rerun()
    elif nombre or numero or anio:
        st.info("Completa un código interno válido y el nombre del proyecto para agregar perforaciones y muestras.")

    def _limpiar_borrador():
        st.session_state.draft_perforaciones = []
        st.session_state.draft_muestras = {}
        st.session_state.bitacora_draft = {k: v for k, v in st.session_state.bitacora_draft.items()
                                            if not k.startswith(f"{codigo_interno}::")}
        # El número sugerido se calcula una sola vez por visita a esta pantalla (ver arriba) — hay
        # que olvidarlo al salir para que la próxima vez se recalcule contra la lista de proyectos
        # ya actualizada (si no, seguiría sugiriendo el mismo número recién usado).
        st.session_state.pop("new_numero", None)

    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Cancelar", use_container_width=True):
            _limpiar_borrador()
            navigate("home")
    with col2:
        if st.button("Guardar bitácora", type="primary", use_container_width=True, icon=":material/save:",
                      disabled=not codigo_valido or not nombre):
            perforaciones_payload = []
            for perf in perforaciones:
                key = f"{codigo_interno}::{perf['codigo']}"
                df_rows = edited_frames.get(key)
                rows = df_rows.to_dict("records") if df_rows is not None else []
                nuevas = []
                for row in rows:
                    numero_m = str(row.get("Número", "")).strip()
                    if not numero_m or numero_m.lower() == "none" or numero_m == "nan":
                        continue
                    id_unico = f"{codigo_interno}-{perf['codigo']}-M{numero_m}"
                    nuevas.append({
                        "numero": numero_m, "id_unico": id_unico,
                        "profundidad_de": row.get("Prof. De") or 0.0, "profundidad_hasta": row.get("Prof. A") or 0.0,
                        "tipo_muestra": row.get("Tipo de muestra") or TIPO_MUESTRA_OPTIONS[0],
                        "ensayos": {e: bool(row.get(e, False)) for e in BITACORA_ENSAYOS},
                        "observaciones": row.get("Observaciones") or "",
                    })
                perforaciones_payload.append({**perf, "muestras": nuevas})

            db.commit_new_project({
                "codigo_interno": codigo_interno, "numero": numero, "anio": anio, "nombre": nombre,
                "localizacion": localizacion, "norma": norma,
                "fecha_bitacora": str(fecha_bitacora), "fecha_ingreso_muestra": str(fecha_ingreso),
                "cliente": cliente, "correo_cliente": correo_cliente, "muestra_tomada_por": muestra_tomada_por,
                "direccion_cliente": direccion_cliente, "telefono_contacto": telefono_contacto,
                "nombre_contacto": nombre_contacto,
                "fecha_inicio_proyecto": str(fecha_inicio_proyecto), "fecha_final_proyecto": str(fecha_final_proyecto),
                "fecha_final_real": str(fecha_final_real) if fecha_final_real else "",
                "fecha_recepcion": str(fecha_recepcion), "fecha_ejecucion": str(fecha_ejecucion), "fecha_emision": str(fecha_emision),
            }, perforaciones_payload)
            _limpiar_borrador()
            st.session_state.selected_codigo = codigo_interno
            navigate("project-detail")


# ════════════════════════════════════════════════════════════════════
# DETALLE DE PROYECTO → PERFORACIONES + PROGRESO
# ════════════════════════════════════════════════════════════════════
def render_project_detail():
    codigo = st.session_state.selected_codigo
    project = get_project(codigo)
    if not project:
        navigate("home")
        return

    if st.button("← Atrás"):
        go_back()

    progreso = project_progress(codigo)
    total = sum(progreso.values())
    pct_general = round(progreso["finalizado"] / total * 100) if total else 0
    perforaciones = st.session_state.perforaciones.get(codigo, [])
    sondeos_txt = f'{len(perforaciones)} sondeo(s) registrado(s)'

    st.markdown(f'''
        <div class="bento-primary" style="margin-bottom:16px;">
            <span class="bento-eyebrow">Proyecto ID</span>
            <h2 style="color:#fff;margin:4px 0 2px 0;">{html.escape(project["codigo_interno"])}</h2>
            <p style="opacity:0.85;font-size:15px;margin:0 0 12px 0;">{html.escape(project["nombre"])}</p>
            <span class="badge" style="background:rgba(255,255,255,0.15);color:#fff;">{sondeos_txt}</span>
        </div>
    ''', unsafe_allow_html=True)

    # Alerta de plazo: solo Jefe y Director Técnico la ven, y solo mientras el proyecto sigue sin
    # entregarse del todo — project_status devuelve "ejecutado" solo cuando TODAS las muestras
    # están finalizadas Y TODOS los ensayos tienen el visto bueno del Director Técnico; si ya se
    # entregó todo, seguir advirtiendo sobre la fecha límite no aporta nada, aunque haya pasado.
    # Se basa en "Fecha final real" (el plazo que de verdad hay que cumplir, ajustado a
    # contratiempos) y no en "Fecha final proyecto" (la referencia inicial que se le da al cliente).
    if st.session_state.role in ("jefe", "ingeniero") and project_status(codigo) != "ejecutado":
        deadline_raw = project.get("fecha_final_real")
        dias_restantes = None
        if deadline_raw:
            try:
                dias_restantes = (date.fromisoformat(deadline_raw) - date.today()).days
            except ValueError:
                dias_restantes = None
        if dias_restantes is not None and dias_restantes <= 0:
            # Vencida (o vence hoy) y todavía faltan entregas: se resalta mucho más fuerte que
            # los demás estados — encabezado propio en mayúsculas + detalle, en vez de la línea
            # sencilla que usan los casos "faltan X días".
            sub = (f"Pasaron {abs(dias_restantes)} día(s) desde la fecha final real y todavía "
                   f"faltan ensayos por entregar." if dias_restantes < 0 else
                   "Hoy es la fecha final real y todavía faltan ensayos por entregar.")
            st.markdown(f'''
                <div style="display:flex;align-items:flex-start;gap:10px;background:{DANGER_LIGHT};
                            color:{DANGER};border-radius:10px;border:1.5px solid {DANGER};
                            padding:12px 14px;margin-bottom:16px;">
                    {icon("report", size=22)}
                    <div>
                        <div style="font-weight:800;font-size:15px;">
                            {"¡FECHA LÍMITE VENCIDA!" if dias_restantes < 0 else "¡VENCE HOY!"}
                        </div>
                        <div style="font-size:13px;margin-top:2px;">{sub} ({deadline_raw})</div>
                    </div>
                </div>
            ''', unsafe_allow_html=True)
        elif dias_restantes is not None:
            if dias_restantes <= 5:
                tono, fondo, icono_alerta = WARNING, WARNING_LIGHT, "schedule"
            else:
                tono, fondo, icono_alerta = SUCCESS, SUCCESS_LIGHT, "event_available"
            texto = f"Faltan {dias_restantes} día(s)"
            st.markdown(f'''
                <div style="display:flex;align-items:center;gap:10px;background:{fondo};color:{tono};
                            border-radius:10px;padding:10px 14px;margin-bottom:16px;font-weight:600;font-size:14px;">
                    {icon(icono_alerta, size=18)}
                    <span>{texto} para la fecha final real del proyecto ({deadline_raw})</span>
                </div>
            ''', unsafe_allow_html=True)

    with st.container(border=True):
        info_rows = [
            ("location_on", "Ubicación", project.get("localizacion")),
            ("rule", "Norma", project.get("norma")),
        ]
        # Datos del cliente: los ven Jefe y Director Técnico (roles de consulta/revisión), nunca el laboratorista.
        if st.session_state.role in ("jefe", "ingeniero"):
            info_rows.insert(1, ("badge", "Cliente", project.get("cliente")))
            info_rows.insert(2, ("mail", "Correo electrónico", project.get("correo_cliente")))
            info_rows.insert(3, ("home_pin", "Dirección cliente", project.get("direccion_cliente")))
            info_rows.insert(4, ("call", "Teléfono de contacto", project.get("telefono_contacto")))
            info_rows.insert(5, ("contact_page", "Nombre de contacto", project.get("nombre_contacto")))

        def _info_row(icono, label, valor, primera):
            margen = "" if primera else "margin-top:14px;"
            return (f'<div class="cell-muted" style="{margen}text-transform:uppercase;letter-spacing:0.04em;font-size:11px;">'
                    f'{icon(icono, size=14)} {label}</div>'
                    f'<div style="font-weight:600;font-size:15px;">{html.escape(valor or "—")}</div>')

        for i, (icono, label, valor) in enumerate(info_rows):
            st.markdown(_info_row(icono, label, valor, i == 0), unsafe_allow_html=True)

        # Fechas agrupadas en paquetes separados (orden/ingreso, plazos del proyecto, ejecución/emisión)
        # para que no se vean como una sola lista larga y confusa.
        date_groups = [
            [("calendar_month", "Fecha de orden", project.get("fecha_bitacora")),
             ("move_to_inbox", "Ingreso de muestras", project.get("fecha_ingreso_muestra"))],
            [("event", "Fecha inicio proyecto", project.get("fecha_inicio_proyecto")),
             ("event_available", "Fecha final proyecto", project.get("fecha_final_proyecto")),
             ("event_available", "Fecha final real", project.get("fecha_final_real"))],
            [("inbox", "Fecha de recepción", project.get("fecha_recepcion")),
             ("science", "Fecha de ejecución", project.get("fecha_ejecucion")),
             ("outbox", "Fecha de emisión", project.get("fecha_emision"))],
        ]
        for gi, group in enumerate(date_groups):
            margen_inferior = "margin-bottom:6px;" if gi == len(date_groups) - 1 else ""
            rows_html = "".join(_info_row(icono, label, valor, i == 0) for i, (icono, label, valor) in enumerate(group))
            st.markdown(
                f'<div style="margin-top:16px;{margen_inferior}padding:10px 12px 12px;border-radius:10px;'
                f'background:rgba(74,120,98,0.06);border:1px solid rgba(74,120,98,0.16);">{rows_html}</div>',
                unsafe_allow_html=True,
            )

    if st.session_state.role == "jefe":
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Editar proyecto", icon=":material/edit:", use_container_width=True):
                navigate("edit-project")
        with c2:
            if confirm_delete(f"project_{codigo}", f"el proyecto {codigo} y todas sus perforaciones y muestras"):
                db.archive_project(project["id"])
                st.session_state.bitacora_draft = {k: v for k, v in st.session_state.bitacora_draft.items() if not k.startswith(codigo + "::")}
                navigate("home")
        if project_status(codigo) == "ejecutado":
            st.caption("Este proyecto ya está completamente ejecutado (todas sus muestras finalizadas).")
            if st.button("Desarchivar proyecto", icon=":material/unarchive:", use_container_width=True):
                desarchivar_proyecto(codigo)
                st.success("Proyecto desarchivado — vuelve a aparecer en Proyectos en ejecución.")
                st.rerun()

    with st.container(border=True):
        st.markdown('<div class="section-title">Progreso general (así avanzan los laboratoristas)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns([1, 3])
        with c1:
            st.markdown(f'<div style="font-size:24px;font-weight:800;color:{PRIMARY};">{pct_general}%</div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="cell-muted" style="margin-top:8px;">{progreso["finalizado"]} de {total} muestras completadas</div>',
                        unsafe_allow_html=True)
        st.progress(pct_general / 100)
        cols = st.columns(3)
        for col, status_key, label in zip(cols, ["sin-iniciar", "en-proceso", "finalizado"],
                                           ["Sin iniciar", "En proceso", "Finalizado"]):
            with col:
                st.markdown(f'''
                    <div style="background:{SECONDARY_CONTAINER};border-radius:8px;
                                padding:6px 6px;text-align:center;margin-top:10px;margin-bottom:10px;">
                        <div style="font-size:9px;text-transform:uppercase;letter-spacing:0.03em;color:{PRIMARY};">{label}</div>
                        <div style="font-size:15px;font-weight:800;color:{PRIMARY};">{progreso[status_key]}</div>
                    </div>
                ''', unsafe_allow_html=True)

    if st.session_state.role == "jefe":
        tiene_bitacora = bool(perforaciones)
        label_bitacora = "Editar bitácora de orden" if tiene_bitacora else "Generar bitácora de orden"
        icon_bitacora = "edit_document" if tiene_bitacora else "assignment"
        if st.button(label_bitacora, type="primary", icon=f":material/{icon_bitacora}:", use_container_width=True):
            navigate("bitacora")

    st.markdown(f'<div class="section-title">Perforaciones realizadas ({len(perforaciones)} elemento(s) identificado(s))</div>',
                unsafe_allow_html=True)
    if not perforaciones:
        st.info("Este proyecto todavía no tiene perforaciones. Usa la Bitácora para agregarlas.")
    for perf in perforaciones:
        muestras = st.session_state.muestras.get(f"{codigo}::{perf['codigo']}", [])
        counts = {"sin-iniciar": 0, "en-proceso": 0, "finalizado": 0}
        for m in muestras:
            counts[compute_muestra_estado(m)] += 1
        perf_total = len(muestras)
        perf_pct = round(counts["finalizado"] / perf_total * 100) if perf_total else 0
        if perf_total == 0:
            estado_badge, estado_label = "badge-muted", "Pendiente"
        elif counts["finalizado"] == perf_total:
            estado_badge, estado_label = "badge-success", "Completado"
        else:
            estado_badge, estado_label = "badge-warning", "En progreso"
        if muestras:
            prof_txt = f'{min(m["profundidad_de"] for m in muestras):.2f}m – {max(m["profundidad_hasta"] for m in muestras):.2f}m'
        else:
            prof_txt = "—"

        with st.container(border=True):
            top = st.columns([1, 3, 1.3])
            with top[0]:
                st.markdown(f'<div class="perf-code-box">{html.escape(perf["codigo"])}</div>', unsafe_allow_html=True)
            with top[1]:
                st.markdown(f'<span style="font-weight:700;">{perf_pct}%</span>&nbsp;&nbsp;'
                            f'<span class="badge {estado_badge}">{estado_label}</span>', unsafe_allow_html=True)
            with top[2]:
                st.markdown(f'<div style="text-align:right;"><span class="assigned-chip">{html.escape(perf["tipo"])}</span></div>',
                            unsafe_allow_html=True)
            st.markdown(f'<div class="cell-muted"><strong>Profundidad</strong> {prof_txt} · {len(muestras)} muestra(s)</div>',
                        unsafe_allow_html=True)
            st.progress(perf_pct / 100)
            bc1, bc2 = st.columns(2)
            with bc1:
                if st.button("Ver muestras →", key=f"open_perf_{perf['codigo']}", use_container_width=True):
                    st.session_state.selected_perforacion = perf["codigo"]
                    navigate("perforacion-detail")
            with bc2:
                filas_perf = _bitacora_filas_perforacion(codigo, perf["codigo"])
                excel_bytes, _truncado = generar_excel_bitacora_orden(project, filas_perf, {perf["tipo"]})
                st.download_button("Descargar bitácora", data=excel_bytes, icon=":material/download:",
                                    file_name=f"{codigo} Bitacora de orden {perf['codigo']}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True, key=f"dl_perf_{perf['codigo']}")


# ════════════════════════════════════════════════════════════════════
# EDITAR PROYECTO (el código interno no se puede editar: es la llave
# que usan perforaciones, muestras y ensayos en session_state)
# ════════════════════════════════════════════════════════════════════
def render_edit_project():
    require_role("jefe")
    codigo = st.session_state.selected_codigo
    project = get_project(codigo)
    if not project:
        navigate("home")
        return

    if st.button("← Atrás"):
        go_back(fallback="project-detail")
    st.markdown("## Editar proyecto")
    st.markdown(f'<span class="code-badge">{html.escape(codigo)}</span>', unsafe_allow_html=True)
    st.caption("El código interno no se puede modificar.")

    def _parse_fecha(valor):
        try:
            return date.fromisoformat(valor)
        except (TypeError, ValueError):
            return date.today()

    # Campos con `key=` (namespaded por código de proyecto) para poder precargarlos desde el
    # Excel del cliente sin pisar los de otro proyecto que se haya editado en la misma sesión.
    for campo, valor_defecto in (
        ("nombre", project.get("nombre", "")), ("localizacion", project.get("localizacion", "")),
        ("cliente", project.get("cliente", "")), ("direccion_cliente", project.get("direccion_cliente", "")),
        ("telefono_contacto", project.get("telefono_contacto", "")), ("correo_cliente", project.get("correo_cliente", "")),
        ("nombre_contacto", project.get("nombre_contacto", "")),
    ):
        wkey = f"edit_{campo}_{codigo}"
        if wkey not in st.session_state:
            st.session_state[wkey] = valor_defecto

    st.markdown('<div class="section-title">Cargar bitácora de proyecto del cliente (opcional)</div>', unsafe_allow_html=True)
    uploaded_cliente_xlsx_edit = st.file_uploader(
        "Bitácora de proyecto del cliente (Excel)", type=["xlsx"], key=f"cliente_xlsx_uploader_edit_{codigo}",
        help="Si el cliente te envió el formato GDA-FL-021 (Bitácora de Proyecto), súbelo aquí para "
             "precargar Cliente, Nombre del proyecto, Localización, Dirección, Teléfono, Correo, "
             "Nombre de contacto y fechas de inicio/fin del proyecto.",
    )
    guard_key = f"_cliente_xlsx_last_edit_{codigo}"
    if uploaded_cliente_xlsx_edit is not None and st.session_state.get(guard_key) != uploaded_cliente_xlsx_edit.name:
        try:
            datos_cliente = _leer_bitacora_cliente_xlsx(uploaded_cliente_xlsx_edit)
            st.session_state[f"edit_nombre_{codigo}"] = datos_cliente["nombre"]
            st.session_state[f"edit_localizacion_{codigo}"] = datos_cliente["localizacion"]
            st.session_state[f"edit_cliente_{codigo}"] = datos_cliente["cliente"]
            st.session_state[f"edit_correo_cliente_{codigo}"] = datos_cliente["correo_cliente"]
            st.session_state[f"edit_direccion_cliente_{codigo}"] = datos_cliente["direccion_cliente"]
            st.session_state[f"edit_telefono_contacto_{codigo}"] = datos_cliente["telefono_contacto"]
            st.session_state[f"edit_nombre_contacto_{codigo}"] = datos_cliente["nombre_contacto"]
            if datos_cliente["fecha_inicio_proyecto"]:
                st.session_state[f"edit_fecha_inicio_proyecto_{codigo}"] = datos_cliente["fecha_inicio_proyecto"]
            if datos_cliente["fecha_final_proyecto"]:
                st.session_state[f"edit_fecha_final_proyecto_{codigo}"] = datos_cliente["fecha_final_proyecto"]
            st.session_state[guard_key] = uploaded_cliente_xlsx_edit.name
            st.success("Datos del cliente cargados desde el Excel. Revísalos abajo antes de guardar.")
        except Exception:
            st.error("No se pudo leer el archivo. Verifica que sea el formato GDA-FL-021 (Bitácora de Proyecto).")

    nombre = st.text_input("Nombre del proyecto", key=f"edit_nombre_{codigo}")
    localizacion = st.text_input("Localización", key=f"edit_localizacion_{codigo}")
    norma_actual = project.get("norma")
    idx = NORMA_PROYECTO_OPTIONS.index(norma_actual) if norma_actual in NORMA_PROYECTO_OPTIONS else 0
    norma = st.radio("Norma", NORMA_PROYECTO_OPTIONS, index=idx, horizontal=True)

    c1, c2 = st.columns(2)
    with c1:
        fecha_bitacora = st.date_input("Fecha de bitácora", value=_parse_fecha(project.get("fecha_bitacora")), format="DD/MM/YYYY")
    with c2:
        fecha_ingreso = st.date_input("Fecha de ingreso de muestra", value=_parse_fecha(project.get("fecha_ingreso_muestra")), format="DD/MM/YYYY")

    ec1, ec2, ec3 = st.columns(3)
    with ec1:
        fecha_recepcion = st.date_input("Fecha de recepción", value=_parse_fecha(project.get("fecha_recepcion")), format="DD/MM/YYYY")
    with ec2:
        fecha_ejecucion = st.date_input("Fecha de ejecución", value=_parse_fecha(project.get("fecha_ejecucion")), format="DD/MM/YYYY")
    with ec3:
        fecha_emision = st.date_input("Fecha de emisión", value=_parse_fecha(project.get("fecha_emision")), format="DD/MM/YYYY")

    st.markdown('<div class="section-title">Datos del cliente (para el encabezado de los informes)</div>', unsafe_allow_html=True)
    cliente = st.text_input("Cliente", key=f"edit_cliente_{codigo}")
    direccion_cliente = st.text_input("Dirección cliente", key=f"edit_direccion_cliente_{codigo}")
    dc1, dc2 = st.columns(2)
    with dc1:
        telefono_contacto = st.text_input("Teléfono de contacto", key=f"edit_telefono_contacto_{codigo}")
    with dc2:
        correo_cliente = st.text_input("Correo electrónico", key=f"edit_correo_cliente_{codigo}")
    nombre_contacto = st.text_input("Nombre de contacto", key=f"edit_nombre_contacto_{codigo}")
    muestra_tomada_por = st.text_input("Muestra tomada por", value=project.get("muestra_tomada_por", ""))

    if f"edit_fecha_inicio_proyecto_{codigo}" not in st.session_state:
        st.session_state[f"edit_fecha_inicio_proyecto_{codigo}"] = _parse_fecha(project.get("fecha_inicio_proyecto"))
    if f"edit_fecha_final_proyecto_{codigo}" not in st.session_state:
        st.session_state[f"edit_fecha_final_proyecto_{codigo}"] = _parse_fecha(project.get("fecha_final_proyecto"))

    dc3, dc4 = st.columns(2)
    with dc3:
        fecha_inicio_proyecto = st.date_input("Fecha inicio proyecto", key=f"edit_fecha_inicio_proyecto_{codigo}", format="DD/MM/YYYY")
    with dc4:
        fecha_final_proyecto = st.date_input("Fecha final proyecto", key=f"edit_fecha_final_proyecto_{codigo}", format="DD/MM/YYYY")
    # A diferencia de las demás, esta fecha se queda vacía (no en hoy) si el proyecto todavía
    # no ha terminado de verdad — por eso no usa _parse_fecha (esa cae a hoy si no hay valor).
    if f"edit_fecha_final_real_{codigo}" not in st.session_state:
        try:
            st.session_state[f"edit_fecha_final_real_{codigo}"] = date.fromisoformat(project.get("fecha_final_real"))
        except (TypeError, ValueError):
            st.session_state[f"edit_fecha_final_real_{codigo}"] = None
    fecha_final_real = st.date_input(
        "Fecha final real", key=f"edit_fecha_final_real_{codigo}", format="DD/MM/YYYY",
        help="Cuándo terminó de verdad el proyecto — distinta de la fecha final planeada.",
    )

    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Cancelar", use_container_width=True):
            go_back(fallback="project-detail")
    with c2:
        if st.button("Guardar cambios", type="primary", use_container_width=True, icon=":material/save:", disabled=not nombre):
            db.update_project(
                project["id"], nombre=nombre, localizacion=localizacion, norma=norma,
                fecha_bitacora=str(fecha_bitacora), fecha_ingreso_muestra=str(fecha_ingreso),
                cliente=cliente, correo_cliente=correo_cliente,
                muestra_tomada_por=muestra_tomada_por, direccion_cliente=direccion_cliente,
                telefono_contacto=telefono_contacto, nombre_contacto=nombre_contacto,
                fecha_inicio_proyecto=str(fecha_inicio_proyecto), fecha_final_proyecto=str(fecha_final_proyecto),
                fecha_final_real=str(fecha_final_real) if fecha_final_real else None,
                fecha_recepcion=str(fecha_recepcion), fecha_ejecucion=str(fecha_ejecucion), fecha_emision=str(fecha_emision),
            )
            navigate("project-detail")


# ════════════════════════════════════════════════════════════════════
# DETALLE DE PERFORACIÓN → LISTA DE MUESTRAS
# ════════════════════════════════════════════════════════════════════
def _perforacion_ensayos_progress(codigo, perf_codigo):
    muestras = st.session_state.muestras.get(f"{codigo}::{perf_codigo}", [])
    total_ensayos, completados = 0, 0
    for m in muestras:
        for label, activo in m["ensayos"].items():
            if not activo or label not in BITACORA_ENSAYOS:
                continue
            total_ensayos += 1
            tipo_interno = SUPPORTED_ASSAY_MAP.get(label)
            a = get_assay(m["id_unico"], tipo_interno) if tipo_interno else None
            if a and a["status"] == "finalizado":
                completados += 1
    return completados, total_ensayos


def render_perforacion_detail():
    codigo = st.session_state.selected_codigo
    perf_codigo = st.session_state.selected_perforacion
    project = get_project(codigo)
    if not project:
        navigate("home")
        return

    if st.button("← Atrás"):
        go_back(fallback="project-detail")

    perf = next((p for p in st.session_state.perforaciones.get(codigo, []) if p["codigo"] == perf_codigo), None)
    muestras = st.session_state.muestras.get(f"{codigo}::{perf_codigo}", [])
    ensayos_completados, ensayos_total = _perforacion_ensayos_progress(codigo, perf_codigo)
    pct = round(ensayos_completados / ensayos_total * 100) if ensayos_total else 0

    st.markdown(f'<div class="cell-muted" style="text-transform:uppercase;letter-spacing:0.04em;font-size:11px;margin-bottom:4px;">Proyecto</div>'
                f'<span class="code-badge">{html.escape(project["codigo_interno"])}</span> '
                f'<span style="font-weight:600;">{html.escape(project["nombre"])}</span>',
                unsafe_allow_html=True)
    st.markdown(f"### Muestras de Perforación {html.escape(perf_codigo)}")

    with st.container(border=True):
        top = st.columns([3, 1])
        with top[0]:
            st.markdown(f"#### Sondeo {html.escape(perf_codigo)}")
            st.markdown(f'<span class="assigned-chip">{html.escape(perf["tipo"] if perf else "—")}</span>',
                        unsafe_allow_html=True)
            st.markdown(f'<div class="cell-muted" style="margin-top:10px;">'
                        f'{icon("location_on", size=13)} {html.escape(project.get("localizacion") or "—")}'
                        f'&nbsp;&nbsp;&nbsp;{icon("calendar_month", size=13)} {html.escape(project.get("fecha_bitacora") or "—")}</div>',
                        unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            filas_perf = _bitacora_filas_perforacion(codigo, perf_codigo)
            excel_bytes, _truncado = generar_excel_bitacora_orden(project, filas_perf, {perf["tipo"]} if perf else set())
            st.download_button("Exportar perfil", data=excel_bytes, icon=":material/download:",
                                file_name=f"{codigo} Bitacora de orden {perf_codigo}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True)
        with c2:
            if st.session_state.role == "jefe":
                if st.button("Nueva muestra", icon=":material/add:", type="primary", use_container_width=True):
                    navigate("bitacora")

    with st.container(border=True):
        st.markdown('<div class="section-title">Avance del sondeo</div>', unsafe_allow_html=True)
        c1, c2 = st.columns([1, 3])
        with c1:
            st.markdown(f'<div style="font-size:32px;font-weight:800;color:{PRIMARY};">{pct}%</div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="cell-muted" style="margin-top:16px;">'
                        f'{icon("check_circle", size=14)} {ensayos_completados} de {ensayos_total} ensayos completados</div>',
                        unsafe_allow_html=True)
        st.progress(pct / 100)

    st.markdown(f'''
        <div style="background:{PRIMARY};color:#fff;border-radius:10px;padding:10px 16px;
                    display:flex;align-items:center;gap:8px;font-weight:700;font-size:14px;margin-bottom:10px;">
            {icon("science", size=16)} Orden de Laboratorio
        </div>
    ''', unsafe_allow_html=True)
    if not muestras:
        st.info("Esta perforación todavía no tiene muestras. Usa la Bitácora para agregarlas.")
    else:
        with st.container(border=True):
            # Los ensayos asignados van apilados uno debajo del otro (ver .assigned-chip-row),
            # así que la columna ya no necesita ancho para 3 chips lado a lado — se le puede
            # devolver espacio a "Tipo"/"Profundidad" sin que sus encabezados se partan en
            # varias líneas (con la columna muy angosta, hasta "Tipo" se partía letra por letra).
            col_ratios = [0.8, 1.3, 1.4, 2.6, 1.0]
            headers = st.columns(col_ratios)
            for col, label in zip(headers, ["ID", "Tipo", "Profundidad", "Ensayos asignados", "Acción"]):
                col.markdown(f'<div class="assigned-th">{label}</div>', unsafe_allow_html=True)
            for i, m in enumerate(muestras):
                if i:
                    st.markdown(f'<hr style="margin:8px 0;border-color:{BORDER};">', unsafe_allow_html=True)
                cols = st.columns(col_ratios, vertical_alignment="center")
                cols[0].markdown(f'<span class="cell-id">M-{html.escape(str(m["numero"]))}</span>', unsafe_allow_html=True)
                cols[1].markdown(f'<span class="cell-muted">{html.escape(m["tipo_muestra"])}</span>', unsafe_allow_html=True)
                cols[2].markdown(f'<span class="cell-muted">{m["profundidad_de"]}–{m["profundidad_hasta"]} m</span>', unsafe_allow_html=True)
                ensayos_sol = [e for e, v in m["ensayos"].items() if v and e in BITACORA_ENSAYOS]
                chip_parts = []
                for e in ensayos_sol:
                    tipo_interno = SUPPORTED_ASSAY_MAP.get(e)
                    existing = get_assay(m["id_unico"], tipo_interno) if tipo_interno else None
                    status = existing["status"] if existing else "sin-iniciar"
                    chip_class = "assigned-chip assigned-chip-sm " + STATUS_BADGE[status].replace("badge-", "assigned-chip-")
                    chip_parts.append(f'<span class="{chip_class}">{html.escape(e)}</span>')
                chips = (f'<div class="assigned-chip-row">{"".join(chip_parts)}</div>' if chip_parts
                         else '<span class="cell-muted">—</span>')
                cols[3].markdown(chips, unsafe_allow_html=True)
                with cols[4]:
                    if st.button("Abrir", key=f"open_muestra_{m['id_unico']}", use_container_width=True):
                        st.session_state.selected_muestra_id = m["id_unico"]
                        navigate("muestra-detail")


# ════════════════════════════════════════════════════════════════════
# BITÁCORA — crea perforaciones y muestras
# ════════════════════════════════════════════════════════════════════
def _bitacora_row_defaults():
    row = {"Número": "", "Prof. De": 0.0, "Prof. A": 0.0, "Tipo de muestra": TIPO_MUESTRA_OPTIONS[0]}
    for e in BITACORA_ENSAYOS:
        row[e] = False
    row["Observaciones"] = ""
    return row


def _muestras_to_rows(muestras):
    rows = []
    for m in muestras:
        row = {"Número": m["numero"], "Prof. De": m["profundidad_de"], "Prof. A": m["profundidad_hasta"], "Tipo de muestra": m["tipo_muestra"]}
        for e in BITACORA_ENSAYOS:
            row[e] = m["ensayos"].get(e, False)
        row["Observaciones"] = m.get("observaciones", "")
        rows.append(row)
    return rows or [_bitacora_row_defaults()]


def _bitacora_draft_df(key, muestras):
    """Inicializa (o repara) el borrador de bitácora de esta perforación en session_state,
    agregando cualquier columna de BITACORA_BASE_COLS que falte —por ejemplo un ensayo nuevo
    que se agregó a BITACORA_ENSAYOS después de que este borrador ya existía en la sesión—
    sin perder lo que el usuario ya haya digitado."""
    if key not in st.session_state.bitacora_draft:
        df = pd.DataFrame(_muestras_to_rows(muestras))
    else:
        df = st.session_state.bitacora_draft[key]
    for col in BITACORA_BASE_COLS:
        if col not in df.columns:
            df[col] = _bitacora_row_defaults()[col]
    st.session_state.bitacora_draft[key] = df[BITACORA_BASE_COLS]
    return st.session_state.bitacora_draft[key]


def _sync_muestras_perforacion(perforacion_id, codigo, perf_codigo, nuevas_rows):
    """Reconcilia la tabla editada de muestras de una perforación existente contra lo que ya
    hay en Supabase: filas con id_unico ya conocido se actualizan, las nuevas se crean, y las
    que ya no aparecen (se borraron en el editor) se archivan."""
    actuales = st.session_state.muestras.get(f"{codigo}::{perf_codigo}", [])
    actuales_by_id_unico = {m["id_unico"]: m for m in actuales}
    vistos = set()
    for row in nuevas_rows:
        id_unico = row["id_unico"]
        vistos.add(id_unico)
        existente = actuales_by_id_unico.get(id_unico)
        campos = {k: v for k, v in row.items() if k != "id_unico"}
        if existente:
            db.update_muestra(existente["id"], **campos)
        else:
            db.create_muestra(perforacion_id, id_unico=id_unico, **campos)
    for id_unico, m in actuales_by_id_unico.items():
        if id_unico not in vistos:
            db.archive_muestra(m["id"])


def render_bitacora():
    require_role("jefe")
    if st.button("← Atrás"):
        go_back()
    st.markdown("## Orden de ensayos para laboratorio")

    codes = [p["codigo_interno"] for p in st.session_state.projects]
    if not codes:
        st.info("Todavía no hay proyectos.")
        return
    default_idx = codes.index(st.session_state.selected_codigo) if st.session_state.selected_codigo in codes else 0
    codigo = st.selectbox("Proyecto", codes, index=default_idx)
    st.session_state.selected_codigo = codigo
    project = get_project(codigo)

    with st.container(border=True):
        st.markdown(f"**{project['nombre']}** · {project.get('localizacion','—')} · Norma {project.get('norma','—')}")

    perforaciones = st.session_state.perforaciones.setdefault(codigo, [])
    es_jefe = st.session_state.role == "jefe"

    if es_jefe:
        st.markdown('<div class="section-title">Agregar perforación</div>', unsafe_allow_html=True)
        c1, c2 = st.columns([2, 1])
        with c1:
            tipo = st.selectbox("Tipo de perforación", list(TIPO_PERFORACION_PREFIX.keys()))
        with c2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Agregar perforación", use_container_width=True, icon=":material/add:"):
                prefix = TIPO_PERFORACION_PREFIX[tipo]
                consecutivo = len([p for p in perforaciones if p["tipo"] == tipo]) + 1
                codigo_perf = f"{prefix}{consecutivo}"
                db.create_perforacion(project["id"], tipo, consecutivo, codigo_perf)
                st.rerun()
    else:
        st.info("Estás viendo la bitácora en modo lectura. Solo el Jefe puede editarla.")

    st.markdown('<div class="section-title">Perforaciones y muestras</div>', unsafe_allow_html=True)
    if not perforaciones:
        st.info("Todavía no hay perforaciones en este proyecto.")

    edited_frames = {}
    for perf in perforaciones:
        key = f"{codigo}::{perf['codigo']}"
        muestras = st.session_state.muestras.setdefault(key, [])

        with st.expander(f"**{perf['codigo']}** — {perf['tipo']}  ·  {len(muestras)} muestra(s)", expanded=True):
            # OJO: el DataFrame se crea UNA sola vez y se reutiliza el mismo objeto en cada rerun.
            # Reconstruirlo desde cero (dict -> DataFrame) en cada actualización es lo que causaba
            # que la primera edición se perdiera y tocara escribir dos veces.
            df_source = _bitacora_draft_df(key, muestras)

            column_config = {
                "Número": st.column_config.TextColumn(default=""),
                # Sin `format`: con NumberColumn + format printf-style, Streamlit reformatea el valor
                # mostrado a mitad de la edición y descarta la primera pulsación, obligando a digitar
                # dos veces. Sin `format` el editor no interfiere y el valor se guarda a la primera.
                "Prof. De": st.column_config.NumberColumn(default=0.0, step=0.01),
                "Prof. A": st.column_config.NumberColumn(default=0.0, step=0.01),
                "Tipo de muestra": st.column_config.SelectboxColumn(options=TIPO_MUESTRA_OPTIONS, default=TIPO_MUESTRA_OPTIONS[0]),
            }
            for e in BITACORA_ENSAYOS:
                column_config[e] = st.column_config.CheckboxColumn(e, default=False)
            column_config["Observaciones"] = st.column_config.TextColumn(
                default="", width="medium", help="Cómo llegó la muestra o cualquier condición que impida el ensayo.")

            if es_jefe:
                st.caption("Usa el ícono para agregar fila sobre la tabla para sumar una muestra nueva. Para eliminar una, selecciona el cuadro a la izquierda de su fila y usa el ícono de basura que aparece sobre la tabla.")
                # OJO: `data` que se le pasa a st.data_editor debe permanecer estable entre reruns
                # (bitacora_draft[key] solo cambia por acciones explícitas nuestras, como "Agregar
                # muestra"). El resultado editado NO se vuelve a guardar ahí — hacerlo generaba el
                # bug de tener que digitar dos veces, porque el editor detectaba la fuente como
                # "cambiada" y descartaba la edición recién hecha.
                edited = st.data_editor(
                    df_source, num_rows="dynamic", use_container_width=True,
                    column_config=column_config, key=f"editor_{key}",
                )
                edited_frames[key] = edited
                filas_perf_rows = edited.to_dict("records")
            else:
                st.dataframe(df_source, use_container_width=True, hide_index=True)
                filas_perf_rows = df_source.to_dict("records")

            # Cada perforación se descarga en su propio Excel: la plantilla oficial
            # representa UN sondeo (hoja "S1"), así que no se mezclan varias en un archivo.
            filas_perf = []
            for row in filas_perf_rows:
                numero_m = str(row.get("Número", "")).strip()
                if not numero_m or numero_m.lower() == "none" or numero_m == "nan":
                    continue
                filas_perf.append({
                    "perf_codigo": perf["codigo"], "numero": numero_m,
                    "tipo_muestra": row.get("Tipo de muestra") or TIPO_MUESTRA_OPTIONS[0],
                    "profundidad_de": row.get("Prof. De") or 0.0, "profundidad_hasta": row.get("Prof. A") or 0.0,
                    "ensayos": {e: bool(row.get(e, False)) for e in BITACORA_ENSAYOS},
                    "observaciones": row.get("Observaciones") or "",
                })
            excel_bytes, truncado = generar_excel_bitacora_orden(project, filas_perf, {perf["tipo"]})
            st.download_button(f"Descargar bitácora — {perf['codigo']}", data=excel_bytes, icon=":material/download:",
                                file_name=f"{codigo} Bitacora de orden {perf['codigo']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True, key=f"dl_bitacora_{key}")
            if truncado:
                st.caption(f"El formato oficial admite hasta {BITACORA_XLSX_MAX_ROWS} muestras; se incluyeron las primeras {BITACORA_XLSX_MAX_ROWS}.")

            if es_jefe and confirm_delete(f"perf_{key}", f"la perforación {perf['codigo']} y todas sus muestras"):
                db.archive_perforacion(perf["id"])
                st.session_state.bitacora_draft.pop(key, None)
                st.rerun()



    if es_jefe:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Guardar bitácora", type="primary", use_container_width=True, icon=":material/save:"):
            for perf in perforaciones:
                key = f"{codigo}::{perf['codigo']}"
                df_rows = edited_frames.get(key)
                rows = df_rows.to_dict("records") if df_rows is not None else []
                nuevas = []
                for row in rows:
                    numero = str(row.get("Número", "")).strip()
                    if not numero or numero.lower() == "none" or numero == "nan":
                        continue
                    id_unico = f"{codigo}-{perf['codigo']}-M{numero}"
                    nuevas.append({
                        "numero": numero, "id_unico": id_unico,
                        "profundidad_de": row.get("Prof. De") or 0.0, "profundidad_hasta": row.get("Prof. A") or 0.0,
                        "tipo_muestra": row.get("Tipo de muestra") or TIPO_MUESTRA_OPTIONS[0],
                        "ensayos": {e: bool(row.get(e, False)) for e in BITACORA_ENSAYOS},
                        "observaciones": row.get("Observaciones") or "",
                    })
                _sync_muestras_perforacion(perf["id"], codigo, perf["codigo"], nuevas)
                # Se descarta el draft cacheado para que, si se vuelve a abrir esta perforación,
                # se reconstruya desde las muestras recién guardadas (evita mostrar/pisar con una
                # tabla vieja lo que ya se guardó).
                st.session_state.bitacora_draft.pop(key, None)
            st.success("Bitácora guardada. Los laboratoristas ya pueden ver y digitar las muestras.")
            st.rerun()


# ════════════════════════════════════════════════════════════════════
# CLASIFICACIÓN USCS (ASTM D2487 / INV E-102), calculada en la app a partir de los datos ya
# digitados de Granulometría y Límites de Atterberg — para no tener que abrir el Excel solo para
# ver a qué grupo pertenece la muestra. OJO: no hay ningún dato digitado en la app que permita
# distinguir un suelo orgánico (color, olor) del inorgánico, así que siempre se asume inorgánico
# (igual que ya asume la descripción visual del laboratorista); por eso conviene que el Jefe/
# Director Técnico verifiquen los primeros resultados contra el Excel antes de confiar en ellos.
# ════════════════════════════════════════════════════════════════════
USCS_NOMBRES = {
    "GW": "Grava bien gradada", "GP": "Grava mal gradada",
    "GM": "Grava limosa", "GC": "Grava arcillosa",
    "GW-GM": "Grava bien gradada con limo", "GW-GC": "Grava bien gradada con arcilla",
    "GP-GM": "Grava mal gradada con limo", "GP-GC": "Grava mal gradada con arcilla",
    "SW": "Arena bien gradada", "SP": "Arena mal gradada",
    "SM": "Arena limosa", "SC": "Arena arcillosa",
    "SW-SM": "Arena bien gradada con limo", "SW-SC": "Arena bien gradada con arcilla",
    "SP-SM": "Arena mal gradada con limo", "SP-SC": "Arena mal gradada con arcilla",
    "CL": "Arcilla de baja plasticidad", "ML": "Limo de baja plasticidad",
    "CL-ML": "Arcilla limosa de baja plasticidad",
    "CH": "Arcilla de alta plasticidad", "MH": "Limo de alta plasticidad",
}


def _calcular_limites_atterberg(data):
    """LL, LP e IP a partir de las lecturas digitadas (INV E-125/E-126, equivalente a ASTM D4318):
    humedad = (masa húmeda - masa seca) / (masa seca - masa recipiente) x 100 por cada ensayo. El
    Límite Líquido es la humedad interpolada a 25 golpes sobre la curva de fluidez (humedad vs.
    log de golpes) — si algún ensayo se hizo exactamente a 25 golpes se usa esa lectura directa en
    vez de la regresión, igual que la fórmula de la plantilla de Excel. Devuelve (None, None, None)
    si no hay lecturas suficientes."""
    puntos = []
    for i in range(1, LIMITE_LIQUIDO_N + 1):
        golpes = to_float(data.get(f"lim_ll_golpes_{i}"))
        humedo = to_float(data.get(f"lim_ll_humedo_{i}"))
        seco = to_float(data.get(f"lim_ll_seco_{i}"))
        recip = to_float(data.get(f"lim_ll_recip_masa_{i}"))
        if None in (golpes, humedo, seco, recip) or golpes <= 0 or (seco - recip) <= 0:
            continue
        puntos.append((golpes, (humedo - seco) / (seco - recip) * 100))

    ll = None
    if puntos:
        exacto25 = [w for g, w in puntos if abs(g - 25) < 0.5]
        if exacto25:
            ll = exacto25[0]
        elif len(puntos) >= 2:
            xs = [math.log10(g) for g, _ in puntos]
            ys = [w for _, w in puntos]
            n = len(xs)
            mean_x, mean_y = sum(xs) / n, sum(ys) / n
            sxx = sum((x - mean_x) ** 2 for x in xs)
            if sxx == 0:
                ll = ys[0]
            else:
                pendiente = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys)) / sxx
                intercepto = mean_y - pendiente * mean_x
                ll = intercepto + pendiente * math.log10(25)
        else:
            ll = puntos[0][1]

    humedades_lp = []
    for i in range(1, LIMITE_PLASTICO_N + 1):
        humedo = to_float(data.get(f"lim_lp_humedo_{i}"))
        seco = to_float(data.get(f"lim_lp_seco_{i}"))
        recip = to_float(data.get(f"lim_lp_recip_masa_{i}"))
        if None in (humedo, seco, recip) or (seco - recip) <= 0:
            continue
        humedades_lp.append((humedo - seco) / (seco - recip) * 100)
    lp = sum(humedades_lp) / len(humedades_lp) if humedades_lp else None

    if ll is None or lp is None:
        return None, None, None
    ll_i, lp_i = int(ll), int(lp)
    return ll_i, lp_i, max(ll_i - lp_i, 0)


def _calcular_curva_granulometrica(gran_data):
    """% que pasa cada tamiz, a partir de las masas retenidas digitadas y la masa inicial seca
    (derivada de las lecturas de Pasa No. 200 — la misma base que ya usa el Excel en D17). Un
    tamiz sin digitar cuenta como 0 g retenido, igual que en la tabla de solo lectura. Devuelve
    None si todavía no hay masa inicial válida (Pasa No. 200 sin digitar)."""
    masa_seco_mas_recip = to_float(gran_data.get("p200_seco_mas_recipiente_antes"))
    masa_recip = to_float(gran_data.get("p200_masa_recipiente_antes"))
    if masa_seco_mas_recip is None or masa_recip is None:
        return None
    masa_inicial = masa_seco_mas_recip - masa_recip
    if masa_inicial <= 0:
        return None

    puntos = []  # (apertura_mm, % que pasa), en orden de tamiz más grande a más chico
    acumulado = 0.0
    for key, _label, apert, _cell in SIEVES:
        retenido = to_float(gran_data.get(key)) or 0.0
        acumulado += retenido
        puntos.append((float(apert), max(0.0, 100 - acumulado / masa_inicial * 100)))

    pct_finos = puntos[-1][1]
    pct_pasa_4 = next((p for d, p in puntos if abs(d - 4.76) < 0.001), None)
    pct_grava = (100 - pct_pasa_4) if pct_pasa_4 is not None else None
    pct_arena = (pct_pasa_4 - pct_finos) if pct_pasa_4 is not None else None
    return {"puntos": puntos, "pct_finos": pct_finos, "pct_grava": pct_grava, "pct_arena": pct_arena}


def _interpolar_diametro(puntos, objetivo):
    """Diámetro (mm) para un % que pasa dado, interpolando log-lineal sobre la curva granulométrica
    (igual que se lee a mano sobre el papel semilogarítmico). None si el % pedido queda fuera del
    rango de tamices digitados."""
    crecientes = sorted(puntos, key=lambda t: t[0])  # de tamiz más chico a más grande
    for (d1, p1), (d2, p2) in zip(crecientes, crecientes[1:]):
        if p1 == p2 == objetivo:
            return d1
        if p1 <= objetivo <= p2 and p2 > p1:
            frac = (objetivo - p1) / (p2 - p1)
            return 10 ** (math.log10(d1) + frac * (math.log10(d2) - math.log10(d1)))
    return None


def clasificar_uscs(gran_data, lim_data):
    """Clasificación USCS (ASTM D2487) a partir de los datos ya digitados. Devuelve un dict con
    el símbolo y los valores intermedios (para verificarlos contra el Excel), o con la lista
    "faltantes" si todavía no hay datos suficientes para completarla."""
    curva = _calcular_curva_granulometrica(gran_data) if gran_data else None
    if curva is None:
        return {"faltantes": ["Faltan las lecturas de Pasa No. 200 (masa inicial de la muestra)."]}

    pct_finos, pct_grava, pct_arena = curva["pct_finos"], curva["pct_grava"], curva["pct_arena"]
    if pct_grava is None:
        return {"faltantes": ["Falta el retenido del tamiz No. 4."]}

    ll = lp = ip = None
    if lim_data:
        ll, lp, ip = _calcular_limites_atterberg(lim_data)

    def _simbolo_fino(ll, ip):
        a_line = 0.73 * (ll - 20)
        if ip < 4 or ip < a_line:
            return "M"
        if ip > 7 and ip >= a_line:
            return "C"
        return "C-M"  # zona rayada CL-ML

    resultado = {"faltantes": [], "pct_grava": pct_grava, "pct_arena": pct_arena, "pct_finos": pct_finos,
                 "ll": ll, "lp": lp, "ip": ip, "cu": None, "cc": None}

    if pct_finos >= 50:
        if ll is None:
            resultado["faltantes"].append("Falta digitar Límites de Atterberg — la muestra tiene 50% o más "
                                           "de finos y la clasificación depende de ellos.")
            return resultado
        base = _simbolo_fino(ll, ip)
        resultado["simbolo"] = "CL-ML" if base == "C-M" else f"{base}{'H' if ll >= 50 else 'L'}"
        return resultado

    d10 = _interpolar_diametro(curva["puntos"], 10)
    d30 = _interpolar_diametro(curva["puntos"], 30)
    d60 = _interpolar_diametro(curva["puntos"], 60)
    cu = (d60 / d10) if (d10 and d60) else None
    cc = ((d30 ** 2) / (d10 * d60)) if (d10 and d30 and d60) else None
    resultado["cu"], resultado["cc"] = cu, cc

    prefijo = "G" if pct_grava >= pct_arena else "S"
    umbral_cu = 4 if prefijo == "G" else 6
    bien_gradada = cu is not None and cc is not None and cu >= umbral_cu and 1 <= cc <= 3
    simbolo_gradacion = f"{prefijo}{'W' if bien_gradada else 'P'}"

    if pct_finos < 5:
        resultado["simbolo"] = simbolo_gradacion
    elif pct_finos > 12:
        if ll is None:
            resultado["faltantes"].append("Falta digitar Límites de Atterberg — la fracción fina de esta "
                                           "muestra supera el 12% y la clasificación depende de ellos.")
            return resultado
        base = _simbolo_fino(ll, ip)
        resultado["simbolo"] = f"{prefijo}{'C' if base == 'C-M' else base}"
    else:
        if ll is None:
            resultado["faltantes"].append("Falta digitar Límites de Atterberg — la fracción fina de esta "
                                           "muestra está entre 5% y 12% y la clasificación depende de ellos.")
            return resultado
        base = _simbolo_fino(ll, ip)
        resultado["simbolo"] = f"{simbolo_gradacion}-{prefijo}{'C' if base == 'C-M' else base}"

    return resultado


AASHTO_NOMBRES = {
    "A-1-a": "Fragmentos de piedra, grava y arena", "A-1-b": "Grava y arena fina",
    "A-3": "Arena fina",
    "A-2-4": "Grava y arena limosa o arcillosa", "A-2-5": "Grava y arena limosa o arcillosa",
    "A-2-6": "Grava y arena limosa o arcillosa", "A-2-7": "Grava y arena limosa o arcillosa",
    "A-4": "Suelo limoso", "A-5": "Suelo limoso",
    "A-6": "Suelo arcilloso", "A-7-5": "Suelo arcilloso", "A-7-6": "Suelo arcilloso",
}


def _grupo_aashto_a2(ll, indice_p):
    if indice_p < 10.5:
        return "A-2-4" if ll < 40.5 else "A-2-5"
    return "A-2-6" if ll < 40.5 else "A-2-7"


def clasificar_aashto(gran_data, lim_data):
    """Clasificación AASHTO (M 145) a partir de los mismos datos de Granulometría y Límites de
    Atterberg que la USCS — reimplementación fiel de la función AASH() de la plantilla oficial
    CLASIFICACION_DE_SUELOS.xlsm (Módulo11 del macro; ver oletools.olevba si hace falta releerla),
    para poder mostrarla también dentro de la app sin tener que abrir el Excel. Devuelve un dict
    con el símbolo y los valores intermedios, o con "faltantes" si aún no hay datos suficientes."""
    curva = _calcular_curva_granulometrica(gran_data) if gran_data else None
    if curva is None:
        return {"faltantes": ["Faltan las lecturas de Pasa No. 200 (masa inicial de la muestra)."]}

    pasa200 = curva["pct_finos"]
    pasa40 = next((p for d, p in curva["puntos"] if abs(d - 0.42) < 0.001), None)
    pasa10 = next((p for d, p in curva["puntos"] if abs(d - 2.00) < 0.001), None)
    if pasa40 is None or pasa10 is None:
        return {"faltantes": ["Faltan los retenidos de los tamices No. 10 y No. 40."]}

    ll = lp = None
    if lim_data:
        ll, lp, _ip = _calcular_limites_atterberg(lim_data)

    resultado = {"faltantes": [], "pasa200": pasa200, "pasa40": pasa40, "pasa10": pasa10, "ll": ll, "lp": lp}
    if ll is None or lp is None:
        resultado["faltantes"].append("Falta digitar Límites de Atterberg — la clasificación AASHTO "
                                       "siempre los necesita, incluso para suelos granulares.")
        return resultado

    indice_p = ll - lp
    if pasa200 <= 35:
        if pasa200 <= 25 and indice_p <= 6:
            if pasa40 <= 30:
                simbolo = ("A-1-a" if pasa10 <= 50 else "A-1-b") if pasa200 <= 15 else "A-1-b"
            elif pasa40 <= 50:
                simbolo = "A-1-b"
            elif pasa200 <= 10 and indice_p == 0:
                simbolo = "A-3"
            else:
                simbolo = _grupo_aashto_a2(ll, indice_p)
        else:
            simbolo = _grupo_aashto_a2(ll, indice_p)
    elif indice_p < 10.5:
        simbolo = "A-4" if ll < 40.5 else "A-5"
    elif ll < 40.5:
        simbolo = "A-6"
    else:
        simbolo = "A-7-5" if lp >= 30 else "A-7-6"

    resultado["simbolo"] = simbolo
    return resultado


# ════════════════════════════════════════════════════════════════════
# DETALLE DE MUESTRA → LISTA DE ENSAYOS SOLICITADOS
# ════════════════════════════════════════════════════════════════════
def render_muestra_detail():
    codigo = st.session_state.selected_codigo
    perf_codigo = st.session_state.selected_perforacion
    muestra_id = st.session_state.selected_muestra_id
    muestra = get_muestra(codigo, perf_codigo, muestra_id)
    if not muestra:
        navigate("home")
        return

    project = get_project(codigo)
    if st.button("← Atrás"):
        go_back(fallback="perforacion-detail")

    estado = compute_muestra_estado(muestra)
    with st.container(border=True, key="muestra-header-card"):
        top = st.columns([3, 1])
        with top[0]:
            st.markdown(f'<div style="display:flex;align-items:center;gap:8px;">'
                        f'<h3 style="margin:0;">Muestra {muestra["numero"]}</h3>'
                        f'{status_badge_html(estado, font_size=13)}</div>', unsafe_allow_html=True)
            st.caption(f"{codigo} · {project['nombre'] if project else ''}")
        with top[1]:
            st.markdown(f'<div style="text-align:right;"><span class="assigned-chip">{html.escape(muestra["tipo_muestra"])}</span></div>',
                        unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        c1.markdown(f'<div class="cell-muted">Identificador</div><div style="font-weight:600;">{html.escape(muestra["id_unico"])}</div>',
                    unsafe_allow_html=True)
        c2.markdown(f'<div class="cell-muted">Profundidad</div><div style="font-weight:600;">'
                    f'{muestra["profundidad_de"]}–{muestra["profundidad_hasta"]} m</div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="cell-muted">Perforación</div><div style="font-weight:600;">{html.escape(perf_codigo)}</div>',
                    unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown('<div class="section-title">Descripción visual de la muestra</div>', unsafe_allow_html=True)
        st.caption("Cómo se ve la muestra a ojo — para comparar con la clasificación USCS calculada de "
                   "los datos de laboratorio, justo abajo.")
        if st.session_state.role == "laboratorista":
            # .upper() al comparar: datos de antes de este cambio (migración 0015) se guardaron
            # en minúscula/mixta (ej. "Limo") — así se siguen preseleccionando en el menú en vez
            # de aparecer en blanco solo porque el case no coincide con las opciones nuevas.
            tipo_actual = (muestra.get("desc_tipo_suelo") or "").upper()
            tipo_idx = DESC_TIPO_SUELO_OPTIONS.index(tipo_actual) if tipo_actual in DESC_TIPO_SUELO_OPTIONS else 0
            desc_tipo = st.selectbox("Tipo de grano", DESC_TIPO_SUELO_OPTIONS, index=tipo_idx,
                                      key=f"desc_tipo_{muestra_id}", format_func=lambda v: v or "— Seleccionar —")
            grueso = _es_grueso(desc_tipo)

            # Componente secundario (ej. "grava con algo de arena", "arcilla con algo de arena")
            # — opcional, se oculta el selectbox si la casilla no está marcada en vez de forzar
            # a elegir "— Seleccionar —" cada vez. Se excluye el tipo principal de la lista para
            # no dejar armar "grava con algo de grava".
            secundario_actual = (muestra.get("desc_tipo_secundario") or "").upper()
            tiene_secundario = st.checkbox(
                "¿Tiene un componente secundario? (ej. grava con algo de arena, arcilla con algo de arena)",
                value=bool(secundario_actual), key=f"desc_tiene_sec_{muestra_id}")
            desc_secundario = None
            if tiene_secundario:
                opciones_sec = [o for o in DESC_TIPO_SECUNDARIO_OPTIONS if o != desc_tipo]
                sec_idx = opciones_sec.index(secundario_actual) if secundario_actual in opciones_sec else 0
                desc_secundario = st.selectbox("Componente secundario", opciones_sec, index=sec_idx,
                                                key=f"desc_sec_{muestra_id}")

            # Qué campos aparecen después depende del tipo de grano elegido arriba — grueso
            # (grava/arena) se describe por angulosidad y cementación, fino (limo/arcilla/
            # orgánico) por consistencia, y forma solo aplica a grava específicamente. Se arman
            # como lista en vez de columnas fijas porque cuáles aparecen cambia, y se van
            # dibujando en parejas de 2 columnas.
            campos = []
            if desc_tipo == "GRAVA":
                campos.append(("Forma", DESC_FORMA_OPTIONS, "desc_forma", "forma"))
            if grueso:
                campos.append(("Angulosidad", DESC_ANGULOSIDAD_OPTIONS, "desc_angulosidad", "angulosidad"))
            campos.append(("Color", DESC_COLOR_OPTIONS, "desc_color", "color"))
            campos.append(("Subtonalidad", DESC_SUBTONALIDAD_OPTIONS, "desc_subtonalidad", "subton"))
            campos.append(("Cementación", DESC_CEMENTACION_OPTIONS, "desc_cementacion", "cem") if grueso
                           else ("Consistencia", DESC_CONSISTENCIA_OPTIONS, "desc_consistencia", "cons"))
            campos.append(("Condición de humedad", DESC_HUMEDAD_OPTIONS, "desc_humedad", "hum"))

            valores = {}
            for i in range(0, len(campos), 2):
                for col, (label, opciones, campo_db, campo_key) in zip(st.columns(2), campos[i:i + 2]):
                    with col:
                        actual = (muestra.get(campo_db) or "").upper()
                        idx = opciones.index(actual) if actual in opciones else 0
                        # La key incluye si el grano es grueso/fino, para forzar un widget nuevo
                        # si la persona cambia de escala — evita arrastrar en pantalla un valor
                        # que ya no aplica (ej. una cementación al pasar de grava a limo).
                        valores[campo_db] = st.selectbox(
                            label, opciones, index=idx, key=f"{campo_key}_{muestra_id}_{grueso}",
                            format_func=lambda v: v or "— Seleccionar —")

            if st.button("Guardar descripción visual", icon=":material/save:", key=f"desc_visual_save_{muestra_id}"):
                guardar = {"desc_tipo_suelo": desc_tipo, "desc_tipo_secundario": desc_secundario, **valores}
                # Los campos que no aparecieron arriba para este tipo de grano (ej. forma si no
                # es grava) se limpian en vez de dejar guardado un valor viejo que ya no se ve.
                for campo_db in ("desc_forma", "desc_angulosidad", "desc_cementacion", "desc_consistencia"):
                    guardar.setdefault(campo_db, None)
                db.update_muestra(muestra["id"], **guardar)
                st.success("Descripción visual guardada.")
                st.rerun()
        else:
            descripcion_val = descripcion_visual_para_excel(muestra)
            if descripcion_val:
                st.markdown(f'<div style="display:flex;gap:10px;align-items:flex-start;background:{BG};'
                             f'border-radius:10px;padding:12px 14px;margin-bottom:4px;">'
                             f'<span style="margin-top:2px;">{icon("visibility", size=18)}</span>'
                             f'<div style="font-weight:600;line-height:1.5;">{html.escape(descripcion_val)}</div></div>',
                             unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="display:flex;align-items:center;gap:6px;color:{NEUTRAL};font-style:italic;">'
                             f'{icon("visibility_off", size=16)} El laboratorista aún no la digita</div>', unsafe_allow_html=True)
            # Segunda versión, con el tipo de suelo que de verdad salió en la clasificación USCS
            # en vez del que se eligió a ojo — se muestra debajo de la inicial, no en su lugar
            # (ver descripcion_visual_calculada). Nada que mostrar hasta que haya datos de
            # Granulometría/Límites suficientes para calcularla.
            descripcion_calc = descripcion_visual_calculada(muestra)
            if descripcion_calc:
                st.markdown(f'<div style="display:flex;gap:10px;align-items:flex-start;background:{SECONDARY_CONTAINER};'
                             f'border-radius:10px;padding:12px 14px;margin-top:8px;">'
                             f'<span style="margin-top:2px;">{icon("science", size=18)}</span>'
                             f'<div><div class="cell-muted" style="margin-bottom:2px;">Según la clasificación USCS calculada</div>'
                             f'<div style="font-weight:600;line-height:1.5;color:{PRIMARY};">{html.escape(descripcion_calc)}</div></div></div>',
                             unsafe_allow_html=True)

    if muestra["ensayos"].get("Granulometría"):
        with st.container(border=True):
            st.markdown('<div class="section-title">Clasificación USCS</div>', unsafe_allow_html=True)
            st.caption("Calculada en la app con los datos ya digitados de Granulometría y Límites de "
                       "Atterberg — verifícala contra el Excel antes de usarla en un informe.")
            gran_assay = get_assay(muestra_id, "granulometria")
            lim_assay = get_assay(muestra_id, "limites")
            resultado = clasificar_uscs(
                gran_assay.get("data") if gran_assay else None,
                lim_assay.get("data") if lim_assay else None,
            )
            simbolo = resultado.get("simbolo")
            if simbolo:
                nombre = USCS_NOMBRES.get(simbolo, "")
                st.markdown(f'''
                    <div style="display:flex;align-items:center;gap:14px;background:{SECONDARY_CONTAINER};
                                border-radius:10px;padding:14px 16px;">
                        <div style="font-size:28px;font-weight:800;color:{PRIMARY};">{simbolo}</div>
                        <div style="font-weight:600;color:{PRIMARY};">{html.escape(nombre)}</div>
                    </div>
                ''', unsafe_allow_html=True)
                detalles = []
                if resultado.get("pct_grava") is not None:
                    detalles.append(f'Grava {resultado["pct_grava"]:.0f}% · Arena {resultado["pct_arena"]:.0f}% '
                                     f'· Finos {resultado["pct_finos"]:.0f}%')
                if resultado.get("ll") is not None:
                    detalles.append(f'LL {resultado["ll"]} · LP {resultado["lp"]} · IP {resultado["ip"]}')
                if resultado.get("cu") is not None and resultado.get("cc") is not None:
                    detalles.append(f'Cu {resultado["cu"]:.1f} · Cc {resultado["cc"]:.1f}')
                if detalles:
                    st.markdown(f'<div class="cell-muted" style="margin-top:10px;">{" · ".join(detalles)}</div>',
                                unsafe_allow_html=True)
            else:
                razones = resultado.get("faltantes") or ["Aún no hay datos suficientes para calcularla."]
                st.markdown(
                    f'<div style="display:flex;flex-direction:column;gap:8px;background:{BG};border-radius:10px;'
                    f'padding:12px 14px;margin-bottom:4px;">' + "".join(
                        f'<div style="display:flex;align-items:center;gap:8px;color:{NEUTRAL};font-style:italic;">'
                        f'{icon("hourglass_empty", size=18)} {html.escape(razon)}</div>' for razon in razones
                    ) + '</div>', unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown('<div class="section-title">Clasificación AASHTO</div>', unsafe_allow_html=True)
            st.caption("Calculada igual que la USCS de arriba, con los mismos datos de Granulometría y "
                       "Límites de Atterberg — verifícala contra el Excel antes de usarla en un informe.")
            resultado_aashto = clasificar_aashto(
                gran_assay.get("data") if gran_assay else None,
                lim_assay.get("data") if lim_assay else None,
            )
            simbolo_aashto = resultado_aashto.get("simbolo")
            if simbolo_aashto:
                nombre_aashto = AASHTO_NOMBRES.get(simbolo_aashto, "")
                st.markdown(f'''
                    <div style="display:flex;align-items:center;gap:14px;background:{SECONDARY_CONTAINER};
                                border-radius:10px;padding:14px 16px;">
                        <div style="font-size:28px;font-weight:800;color:{PRIMARY};">{simbolo_aashto}</div>
                        <div style="font-weight:600;color:{PRIMARY};">{html.escape(nombre_aashto)}</div>
                    </div>
                ''', unsafe_allow_html=True)
                detalles_aashto = []
                if resultado_aashto.get("pasa200") is not None:
                    detalles_aashto.append(f'Pasa No. 200 {resultado_aashto["pasa200"]:.0f}% · '
                                            f'Pasa No. 40 {resultado_aashto["pasa40"]:.0f}% · '
                                            f'Pasa No. 10 {resultado_aashto["pasa10"]:.0f}%')
                if resultado_aashto.get("ll") is not None:
                    detalles_aashto.append(f'LL {resultado_aashto["ll"]} · LP {resultado_aashto["lp"]}')
                if detalles_aashto:
                    st.markdown(f'<div class="cell-muted" style="margin-top:10px;">{" · ".join(detalles_aashto)}</div>',
                                unsafe_allow_html=True)
            else:
                razones_aashto = resultado_aashto.get("faltantes") or ["Aún no hay datos suficientes para calcularla."]
                st.markdown(
                    f'<div style="display:flex;flex-direction:column;gap:8px;background:{BG};border-radius:10px;'
                    f'padding:12px 14px;margin-bottom:4px;">' + "".join(
                        f'<div style="display:flex;align-items:center;gap:8px;color:{NEUTRAL};font-style:italic;">'
                        f'{icon("hourglass_empty", size=18)} {html.escape(razon)}</div>' for razon in razones_aashto
                    ) + '</div>', unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown('<div class="section-title">Observaciones</div>', unsafe_allow_html=True)
        st.caption("El laboratorista la digita si la muestra presenta fisuras o no se puede realizar el ensayo "
                   "por alguna razón.")
        if st.session_state.role == "laboratorista":
            with st.container(key="muestra-obs-box"):
                observacion = st.text_area(
                    "Observaciones", value=muestra.get("observaciones", ""), label_visibility="collapsed",
                    placeholder="Ej: Muestra con fisuras visibles, no fue posible completar el ensayo...", key=f"obs_{muestra_id}",
                )
            if st.button("Guardar observación", icon=":material/save:", key=f"obs_save_{muestra_id}"):
                db.update_muestra(muestra["id"], observaciones=observacion)
                st.success("Observación guardada.")
                st.rerun()
        else:
            observaciones_val = muestra.get("observaciones")
            if observaciones_val:
                st.markdown(f'<div class="cell-muted">{html.escape(observaciones_val)}</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="display:flex;align-items:center;gap:8px;background:{BG};'
                             f'border-radius:10px;padding:12px 14px;margin-bottom:4px;color:{NEUTRAL};font-style:italic;">'
                             f'{icon("inbox", size=18)} Sin observaciones</div>', unsafe_allow_html=True)

    # Filtra por si la muestra guarda un ensayo que ya no es seleccionable (p. ej. "Pasa 200",
    # que quedó incluido dentro de Granulometría) — no se muestra aunque quede marcado en datos viejos.
    solicitados = [e for e, v in muestra["ensayos"].items() if v and e in BITACORA_ENSAYOS]
    finalizados = sum(
        1 for e in solicitados
        if SUPPORTED_ASSAY_MAP.get(e) and (get_assay(muestra_id, SUPPORTED_ASSAY_MAP[e]) or {}).get("status") == "finalizado"
    )
    total_sol = len(solicitados)
    pct_ensayos = round(finalizados / total_sol * 100) if total_sol else 0

    with st.container(border=True):
        st.markdown('<div class="section-title">Avance de ensayos</div>', unsafe_allow_html=True)
        c1, c2 = st.columns([1, 3])
        with c1:
            st.markdown(f'<div style="font-size:32px;font-weight:800;color:{PRIMARY};">{pct_ensayos}%</div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="cell-muted" style="margin-top:16px;">{finalizados} de {total_sol} ensayos programados</div>',
                        unsafe_allow_html=True)
        st.progress(pct_ensayos / 100)

    st.markdown('<div class="section-title">Ensayos asignados</div>', unsafe_allow_html=True)
    if not solicitados:
        st.info("Esta muestra no tiene ensayos marcados en la bitácora.")
    else:
        finalizables = [e for e in solicitados if SUPPORTED_ASSAY_MAP.get(e)]
        todos_finalizados = bool(finalizables) and all(
            (get_assay(muestra_id, SUPPORTED_ASSAY_MAP[e]) or {}).get("status") == "finalizado" for e in finalizables)
        todos_aprobados = todos_finalizados and all(
            (get_assay(muestra_id, SUPPORTED_ASSAY_MAP[e]) or {}).get("etapa_revision") == "aprobado" for e in finalizables)
        if todos_aprobados:
            st.success("Todos los ensayos fueron aprobados por el Director Técnico — esta muestra está lista "
                        "para entregar al cliente.")
        elif todos_finalizados:
            st.info("Verifica que todas las aprobaciones del Director Técnico estén completas antes de dar "
                     "esta muestra por lista para el cliente.")

        for ensayo_label in solicitados:
            tipo_interno = SUPPORTED_ASSAY_MAP.get(ensayo_label)
            existing = get_assay(muestra_id, tipo_interno) if tipo_interno else None
            status = existing["status"] if existing else "sin-iniciar"
            etapa = existing.get("etapa_revision") if existing else None
            mostrar_aprobacion = bool(tipo_interno and existing and status == "finalizado")
            slug = re.sub(r"[^a-z0-9]+", "-", ensayo_label.lower())

            with st.container(border=True, key=f"ensayo-card-{slug}"):
                cols = st.columns([0.5, 2.0, 1.8, 2.6, 0.9], vertical_alignment="center")
                cols[0].markdown(status_circle_html(status), unsafe_allow_html=True)
                with cols[1]:
                    st.markdown(f"**{ensayo_label}**")
                    if existing and existing.get("laboratorist"):
                        st.markdown(f'<div class="timestamp-caption">{icon("history", size=13)} Última actualización: {format_dt(existing["lastModified"])} · {existing["laboratorist"]}</div>', unsafe_allow_html=True)
                    elif existing:
                        st.markdown(f'<div class="timestamp-caption">{icon("history", size=13)} Última actualización: {format_dt(existing["lastModified"])}</div>', unsafe_allow_html=True)

                if not tipo_interno:
                    cols[2].markdown('<span class="badge badge-muted">Sin formulario aún</span>', unsafe_allow_html=True)
                else:
                    with cols[2]:
                        if mostrar_aprobacion:
                            st.markdown(aprobacion_badge_html(etapa), unsafe_allow_html=True)
                        else:
                            st.markdown(status_badge_html(status), unsafe_allow_html=True)

                    with cols[3]:
                        if not mostrar_aprobacion:
                            pass
                        elif etapa == "aprobado":
                            if st.session_state.role == "ingeniero":
                                with st.popover("Desconfirmar", use_container_width=True):
                                    st.caption("Si los resultados no satisfacen al cliente, esto reabre el ensayo "
                                               "para el laboratorista y reinicia el ciclo de confirmación.")
                                    motivo_desconf = st.text_area("Motivo", key=f"desconfirmar_motivo_{ensayo_label}",
                                                                   placeholder="Qué hay que corregir...")
                                    if st.button("Confirmar desconfirmación", key=f"desconfirmar_{ensayo_label}", use_container_width=True):
                                        if motivo_desconf.strip():
                                            db.ing_desconfirmar(existing["id"], st.session_state.profile, motivo_desconf)
                                            add_notification("laboratorista", f"El Director Técnico desconfirmó {ensayo_label} de la Muestra "
                                                                          f"{muestra['numero']} de {codigo}: {motivo_desconf}", codigo, perf_codigo, muestra_id)
                                            add_notification("jefe", f"El Director Técnico desconfirmó {ensayo_label} de la Muestra "
                                                                      f"{muestra['numero']} de {codigo}: {motivo_desconf}", codigo, perf_codigo, muestra_id)
                                            add_historial(existing, "Desconfirmado por el Director Técnico", f"Director Técnico: {motivo_desconf}",
                                                          icono="undo", tono="danger")
                                            st.success("Desconfirmado — vuelve al laboratorista.")
                                            st.rerun()
                                        else:
                                            st.error("Escribe el motivo antes de desconfirmar.")
                            else:
                                st.button("Confirmado", disabled=True, use_container_width=True, key=f"aprobado_ro_{ensayo_label}")
                        elif etapa == "pendiente_ing":
                            if st.session_state.role == "ingeniero":
                                b1, b2 = st.columns(2)
                                with b1:
                                    if st.button("Confirmar", type="primary",
                                                 use_container_width=True, key=f"ing_aprobar_{ensayo_label}"):
                                        db.ing_aprobar(existing["id"], st.session_state.profile)
                                        add_notification("jefe", f"El Director Técnico aprobó {ensayo_label} de la Muestra "
                                                                  f"{muestra['numero']} de {codigo}.", codigo, perf_codigo, muestra_id)
                                        add_historial(existing, "Aprobación Final del Director Técnico", "Director Técnico",
                                                      icono="verified", tono="success")
                                        st.success("Aprobado.")
                                        st.rerun()
                                with b2:
                                    with st.popover("Devolver", use_container_width=True):
                                        motivo_ing = st.text_area("Motivo", key=f"ing_motivo_{ensayo_label}",
                                                                   placeholder="Qué hay que corregir...")
                                        if st.button("Confirmar devolución", key=f"ing_devolver_{ensayo_label}", use_container_width=True):
                                            if motivo_ing.strip():
                                                db.ing_devolver(existing["id"], st.session_state.profile, motivo_ing)
                                                add_notification("jefe", f"El Director Técnico devolvió {ensayo_label} de la Muestra "
                                                                          f"{muestra['numero']} de {codigo}: {motivo_ing}", codigo, perf_codigo, muestra_id)
                                                add_historial(existing, "Devuelto al Jefe de Laboratorio", f"Director Técnico: {motivo_ing}",
                                                              icono="undo", tono="danger")
                                                st.success("Devuelto al Jefe.")
                                                st.rerun()
                                            else:
                                                st.error("Escribe el motivo antes de devolver.")
                            elif st.session_state.role == "jefe":
                                with st.popover("Desconfirmar", use_container_width=True):
                                    st.caption("Retira tu confirmación — el ensayo deja de estar pendiente de "
                                               "revisión del Director Técnico.")
                                    motivo_jefe_desconf = st.text_area("Motivo", key=f"jefe_desconfirmar_motivo_{ensayo_label}",
                                                                        placeholder="Por qué te retractas...")
                                    if st.button("Confirmar desconfirmación", key=f"jefe_desconfirmar_{ensayo_label}", use_container_width=True):
                                        if motivo_jefe_desconf.strip():
                                            db.jefe_desconfirmar(existing["id"], st.session_state.profile, motivo_jefe_desconf)
                                            add_notification("ingeniero", f"El Jefe de Laboratorio desconfirmó {ensayo_label} de la Muestra "
                                                                           f"{muestra['numero']} de {codigo} — ya no está pendiente de tu revisión.",
                                                              codigo, perf_codigo, muestra_id)
                                            add_historial(existing, "Desconfirmado por el Jefe de Laboratorio", f"Jefe de Laboratorio: {motivo_jefe_desconf}",
                                                          icono="undo", tono="danger")
                                            st.success("Desconfirmado.")
                                            st.rerun()
                                        else:
                                            st.error("Escribe el motivo antes de desconfirmar.")
                            else:
                                st.caption("Esperando al Director Técnico")
                        else:
                            if st.session_state.role == "jefe":
                                b1, b2 = st.columns(2)
                                with b1:
                                    if st.button("Confirmar", type="primary",
                                                 use_container_width=True, key=f"jefe_confirmar_{ensayo_label}"):
                                        db.jefe_confirmar(existing["id"], st.session_state.profile)
                                        add_notification("ingeniero", f"El Jefe de Laboratorio envió {ensayo_label} de la Muestra "
                                                                       f"{muestra['numero']} de {codigo} para tu confirmación final.",
                                                          codigo, perf_codigo, muestra_id)
                                        add_historial(existing, "Ensayo Confirmado", "Jefe de Laboratorio",
                                                      icono="check_circle", tono="success")
                                        st.success("Enviado al Director Técnico.")
                                        st.rerun()
                                with b2:
                                    with st.popover("Devolver", use_container_width=True):
                                        motivo_jefe = st.text_area("Motivo", key=f"jefe_motivo_{ensayo_label}",
                                                                    placeholder="Qué hay que corregir...")
                                        if st.button("Confirmar devolución", key=f"jefe_devolver_{ensayo_label}", use_container_width=True):
                                            if motivo_jefe.strip():
                                                db.jefe_devolver(existing["id"], st.session_state.profile, motivo_jefe)
                                                add_notification("laboratorista", f"El Jefe de Laboratorio devolvió {ensayo_label} de la Muestra "
                                                                              f"{muestra['numero']} de {codigo}: {motivo_jefe}", codigo, perf_codigo, muestra_id)
                                                add_historial(existing, "Devuelto al Laboratorista", f"Jefe de Laboratorio: {motivo_jefe}",
                                                              icono="undo", tono="danger")
                                                st.success("Devuelto al laboratorista.")
                                                st.rerun()
                                            else:
                                                st.error("Escribe el motivo antes de devolver.")
                            else:
                                st.caption("Pendiente del Jefe")

                    with cols[4]:
                        if st.button("Abrir", key=f"open_ensayo_{ensayo_label}", use_container_width=True):
                            if existing:
                                st.session_state.selected_assay_id = existing["id"]
                            else:
                                nuevo = db.create_assay(muestra["id"], tipo_interno)
                                st.session_state.selected_assay_id = nuevo["id"]
                            st.session_state.selected_assay_type = tipo_interno
                            navigate("assay-form")

                if existing:
                    motivo = existing.get("motivo_rechazo")
                    if motivo:
                        quien = "el Jefe de Laboratorio" if existing.get("rechazado_por") == "jefe" else "el Director Técnico"
                        st.warning(f"Devuelto por {quien}: {motivo}")
                    # El historial se muestra siempre que exista (no solo cuando el ensayo está
                    # "Finalizado") — un desconfirmar/devolver reabre el ensayo pero no borra su
                    # registro de auditoría, y debe seguir siendo visible.
                    ens_historial = existing.get("historial", [])
                    if ens_historial:
                        with st.expander(f"Historial de Cambios ({len(ens_historial)})", icon=":material/history:"):
                            st.markdown(historial_timeline_html(ens_historial), unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════
# GENERAR EXCEL DE BITÁCORA DE ORDEN (plantilla oficial GDA-FL-003)
# ════════════════════════════════════════════════════════════════════
def _bitacora_filas_perforacion(codigo, perf_codigo):
    """Muestras de UNA perforación ya guardada, en el formato que espera
    generar_excel_bitacora_orden. Cada perforación se exporta a su propio Excel —
    la plantilla oficial (hoja "S1") representa un solo sondeo, no varios a la vez."""
    filas = []
    for m in st.session_state.muestras.get(f"{codigo}::{perf_codigo}", []):
        filas.append({
            "perf_codigo": perf_codigo, "numero": m["numero"], "tipo_muestra": m["tipo_muestra"],
            "profundidad_de": m["profundidad_de"], "profundidad_hasta": m["profundidad_hasta"],
            "ensayos": m["ensayos"], "observaciones": m.get("observaciones", ""),
        })
    return filas


def generar_excel_bitacora_orden(project, filas, tipos_usados):
    """filas: lista de dicts con perf_codigo, numero, tipo_muestra, profundidad_de,
    profundidad_hasta, ensayos (dict) y observaciones. Devuelve (bytes, truncado)."""
    wb = load_workbook(TEMPLATE_BITACORA_ORDEN)
    ws = wb["S1"]

    # La plantilla ya trae "GDA" impreso en su propio recuadro (AC8); solo se llena el
    # recuadro siguiente (AG8) con número-año, tal como se digita en Código interno.
    numero_proy = str(project.get("numero") or "").strip()
    anio_proy = str(project.get("anio") or "").strip()
    if numero_proy or anio_proy:
        ws["AG8"] = f"{numero_proy}-{anio_proy}"
    fecha_partes = str(project.get("fecha_bitacora") or "").split("-")  # "AAAA-MM-DD"
    if len(fecha_partes) == 3:
        anio, mes, dia = fecha_partes
        ws["E8"] = dia
        ws["F8"] = mes
        ws["H8"] = anio
    ws["F10"] = project.get("nombre") or ""
    ws["E12"] = project.get("localizacion") or ""

    norma_cell = BITACORA_XLSX_NORMA_CELL.get(project.get("norma"))
    if norma_cell:
        ws[norma_cell] = "X"
    for tipo in tipos_usados:
        tipo_cell = BITACORA_XLSX_TIPO_CELL.get(tipo)
        if tipo_cell:
            ws[tipo_cell] = "X"

    truncado = len(filas) > BITACORA_XLSX_MAX_ROWS
    for i, fila in enumerate(filas[:BITACORA_XLSX_MAX_ROWS]):
        r = 18 + i
        ws[f"A{r}"] = fila["perf_codigo"]
        ws[f"B{r}"] = fila["numero"]
        ws[f"C{r}"] = fila["tipo_muestra"]
        ws[f"D{r}"] = to_float(fila.get("profundidad_de"))
        ws[f"E{r}"] = to_float(fila.get("profundidad_hasta"))
        for label, activo in fila.get("ensayos", {}).items():
            col = BITACORA_XLSX_ENSAYO_COL.get(label)
            if activo and col:
                ws[f"{col}{r}"] = "X"
        ws[f"AH{r}"] = fila.get("observaciones") or ""

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return _restaurar_imagenes_perdidas(bio.getvalue(), TEMPLATE_BITACORA_ORDEN), truncado


# ════════════════════════════════════════════════════════════════════
# GENERAR EXCEL DE GRANULOMETRÍA Y HUMEDAD (plantillas reales del laboratorio,
# ambas comparten el mismo diseño de encabezado — filas 1 a 13)
# ════════════════════════════════════════════════════════════════════
def _fecha_ddmmaaaa(iso_str):
    """Convierte una fecha guardada en formato ISO ("AAAA-MM-DD") a DD/MM/AAAA para que
    se vea igual que en la app al pasarla al Excel."""
    try:
        return date.fromisoformat(iso_str).strftime("%d/%m/%Y")
    except (TypeError, ValueError):
        return iso_str or ""


def _llenar_encabezado_informe(ws, codigo, perf_codigo, muestra, project, observaciones_ensayo="", perf_numero_cell="F12"):
    ws["D6"] = project.get("cliente", "") if project else ""  # Cliente
    ws["D7"] = project["nombre"] if project else codigo          # Proyecto
    ws["D8"] = project.get("correo_cliente", "") if project else ""  # Correo electrónico
    ws["D9"] = project.get("localizacion", "") if project else ""  # Localización
    ws["D10"] = project.get("muestra_tomada_por", "") if project else ""  # Muestra tomada por
    ws["K6"] = _fecha_ddmmaaaa(project.get("fecha_recepcion", "")) if project else ""  # Fecha de recepción
    ws["K7"] = _fecha_ddmmaaaa(project.get("fecha_ejecucion", "")) if project else ""  # Fecha de ejecución
    ws["K8"] = _fecha_ddmmaaaa(project.get("fecha_emision", "")) if project else ""  # Fecha de emisión
    ws["L9"] = project.get("numero", "") if project else ""  # Código interno — número (K9 ya trae "GDA")
    ws["M9"] = project.get("anio", "") if project else ""  # Código interno — año

    perf = get_perforacion(codigo, perf_codigo)
    ws["D12"] = TIPO_PERFORACION_EXCEL.get(perf["tipo"], "") if perf else ""  # Tipo de perforación (lista desplegable)
    # El número de perforación cae en una celda distinta según la plantilla: F12 en la de
    # Humedad (GDA-FLC-014), E12 en la de Granulometría/Límites (GDA-FLC-001, actualizada 2026).
    ws[perf_numero_cell] = perf["consecutivo"] if perf else ""
    ws["H12"] = muestra["numero"]
    ws["K12"] = to_float(muestra.get("profundidad_de"))
    ws["M12"] = to_float(muestra.get("profundidad_hasta"))
    # Descripción visual: primero la frase armada de los menús desplegables (+ notas
    # adicionales si hay, ver descripcion_visual_para_excel) — independiente de las
    # Observaciones—, si no hay nada de eso, lo que el laboratorista escribió en "Observaciones"
    # del propio ensayo, y solo si ninguna de las dos existe, el tipo de muestra como último recurso.
    ws["D13"] = descripcion_visual_para_excel(muestra) or observaciones_ensayo or f"Tipo de muestra: {muestra.get('tipo_muestra','')}"


def _escribir_limites(ws, data):
    def _escribir(filas):
        for key, _label, cells in filas:
            es_recipiente = key.endswith("recipiente")
            for i, cell in enumerate(cells, start=1):
                valor = data.get(f"{key}_{i}", "")
                if es_recipiente:
                    ws[cell] = valor
                else:
                    num = to_float(valor)
                    if num is not None:
                        ws[cell] = num
    _escribir(LIMITE_LIQUIDO_FILAS)
    _escribir(LIMITE_PLASTICO_FILAS)


def _reparar_graficos_perdidos(xlsm_bytes, template_path):
    """openpyxl no conserva las 'chartUserShapes' de un gráfico (las anotaciones de texto
    dibujadas a mano encima, ej. las etiquetas LÍNEA U/LÍNEA A/CH/CL-ML de la Carta de
    Plasticidad): se pierden tanto los archivos (drawingN.xml, chartN.xml.rels) como —esto
    es lo que de verdad hace que no se vean— la propia referencia <c:userShapes r:id="..."/>
    dentro del chartN.xml, que es lo que le dice a Excel que busque esas formas. El gráfico
    y sus series/ejes sí sobreviven intactos porque no dependen de esa referencia.
    Esta función restaura los archivos faltantes desde la plantilla original Y vuelve a
    insertar la referencia dentro del/de los chartN.xml correspondientes."""
    with zipfile.ZipFile(template_path) as tpl:
        tpl_names = set(tpl.namelist())
        with zipfile.ZipFile(BytesIO(xlsm_bytes)) as out:
            out_names = set(out.namelist())
            faltantes = {n for n in tpl_names - out_names
                         if n.startswith("xl/drawings/") or n.startswith("xl/charts/_rels/")}
            if not faltantes:
                return xlsm_bytes

            content_types_tpl = tpl.read("[Content_Types].xml").decode("utf-8")
            content_types_out = out.read("[Content_Types].xml").decode("utf-8")
            for nombre in faltantes:
                if not nombre.endswith(".xml"):
                    continue  # los .rels usan el Default de extensión "rels", no necesitan Override
                parte = "/" + nombre
                m = re.search(r'<Override PartName="' + re.escape(parte) + r'"[^>]*?/>', content_types_tpl)
                if m and m.group(0) not in content_types_out:
                    content_types_out = content_types_out.replace("</Types>", m.group(0) + "</Types>")

            # Para cada chartN.xml.rels restaurado, extraemos el rId de su relación
            # "chartUserShapes" y lo reinsertamos dentro del chartN.xml de salida (que openpyxl
            # ya escribió, pero sin esa referencia).
            R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
            charts_a_parchar = {}  # "xl/charts/chartN.xml" -> rId
            for nombre in faltantes:
                m = re.match(r"xl/charts/_rels/(chart\d+)\.xml\.rels$", nombre)
                if not m:
                    continue
                rels_xml = tpl.read(nombre).decode("utf-8")
                rm = re.search(
                    r'<Relationship[^>]*Type="[^"]*chartUserShapes"[^>]*Id="([^"]+)"'
                    r'|<Relationship[^>]*Id="([^"]+)"[^>]*Type="[^"]*chartUserShapes"',
                    rels_xml,
                )
                if rm:
                    rid = rm.group(1) or rm.group(2)
                    charts_a_parchar[f"xl/charts/{m.group(1)}.xml"] = rid

            chart_bytes_parchados = {}
            for chart_nombre, rid in charts_a_parchar.items():
                if chart_nombre not in out_names:
                    continue
                chart_xml = out.read(chart_nombre).decode("utf-8")
                if "userShapes" in chart_xml:
                    continue  # ya la tiene, nada que hacer
                # La etiqueta raíz puede no declarar el namespace "r:" (openpyxl no lo usa en
                # el cuerpo del chart), así que lo agregamos si hace falta.
                if 'xmlns:r=' not in chart_xml.split(">", 1)[0]:
                    chart_xml = chart_xml.replace(
                        "<chartSpace ", f'<chartSpace xmlns:r="{R_NS}" ', 1
                    )
                userShapes_tag = f'<userShapes r:id="{rid}"/>'
                chart_xml = chart_xml.replace("</chartSpace>", userShapes_tag + "</chartSpace>")
                chart_bytes_parchados[chart_nombre] = chart_xml.encode("utf-8")

            bio = BytesIO()
            with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
                for item in out.infolist():
                    if item.filename == "[Content_Types].xml":
                        z.writestr(item, content_types_out)
                    elif item.filename in chart_bytes_parchados:
                        z.writestr(item, chart_bytes_parchados[item.filename])
                    else:
                        z.writestr(item, out.read(item.filename))
                for nombre in faltantes:
                    z.writestr(nombre, tpl.read(nombre))
            bio.seek(0)
            return bio.getvalue()


def _restaurar_imagenes_perdidas(xlsx_bytes, template_path):
    """openpyxl vacía las <xdr:pic> (imágenes incrustadas — logo, firmas, diagramas de
    referencia de la hoja GUIA) de cualquier drawingN.xml al volver a guardar el archivo: el
    archivo drawingN.xml sigue existiendo y las formas/gráficos de adentro se conservan, pero
    las imágenes puntuales se pierden en blanco porque openpyxl no sabe reconstruir esa parte
    del XML al reescribirlo (a diferencia de _reparar_graficos_perdidos, que arregla archivos
    que openpyxl bota por completo, acá el archivo sigue estando pero vacío de imágenes).

    Restaurar solo el drawingN.xml (y su .rels) NO basta: openpyxl vuelve a numerar y a mezclar
    TODOS los archivos xl/media/imageN.png al guardar, así que el "image7.png" que deja openpyxl
    casi nunca es el mismo contenido que el "image7.png" de la plantilla, aunque el nombre
    coincida — restaurar solo el drawing dejaba el nombre correcto apuntando a la imagen
    equivocada (una firma se veía reemplazada por la etiqueta de otra parte de la hoja). Por eso
    también se restauran, con su contenido original, los xl/media/*.png que el drawing
    restaurado referencia — así drawing + rels + imagen quedan siempre como un trío consistente
    sacado íntegro de la plantilla, sin importar qué numeración haya usado openpyxl."""
    with zipfile.ZipFile(template_path) as tpl:
        tpl_names = set(tpl.namelist())
        drawings = {n for n in tpl_names if re.match(r"xl/drawings/drawing\d+\.xml$", n)}
        parches = {}
        with zipfile.ZipFile(BytesIO(xlsx_bytes)) as out:
            out_names = set(out.namelist())
            for nombre in drawings:
                tpl_xml = tpl.read(nombre).decode("utf-8")
                if "<xdr:pic>" not in tpl_xml or nombre not in out_names:
                    continue  # sin imágenes en la plantilla, o el archivo falta entero (lo arregla _reparar_graficos_perdidos)
                out_xml = out.read(nombre).decode("utf-8")
                if out_xml.count("<xdr:pic>") >= tpl_xml.count("<xdr:pic>"):
                    continue  # openpyxl sí las conservó, nada que restaurar
                parches[nombre] = tpl.read(nombre)
                rels_nombre = nombre.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
                if rels_nombre not in tpl_names:
                    continue
                rels_bytes = tpl.read(rels_nombre)
                parches[rels_nombre] = rels_bytes
                for target in re.findall(r'Target="(\.\./media/[^"]+)"', rels_bytes.decode("utf-8")):
                    media_nombre = "xl/" + target[len("../"):]
                    if media_nombre in tpl_names:
                        parches[media_nombre] = tpl.read(media_nombre)

            if not parches:
                return xlsx_bytes

            bio = BytesIO()
            with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
                for item in out.infolist():
                    if item.filename in parches:
                        z.writestr(item, parches.pop(item.filename))
                    else:
                        z.writestr(item, out.read(item.filename))
                for nombre, data in parches.items():
                    z.writestr(nombre, data)  # por si openpyxl nunca escribió ese archivo
            bio.seek(0)
            return bio.getvalue()


def _generar_excel_clasificacion(codigo, perf_codigo, muestra, project, gran_data=None, lim_data=None, observaciones_ensayo=""):
    """Granulometría y Límites de Atterberg comparten la misma plantilla y hoja ("GUIA") —
    por muestra van juntos en un solo archivo, sin importar si se descarga desde el ensayo de
    Granulometría o desde el de Límites."""
    wb = load_workbook(TEMPLATE_GRANULOMETRIA, keep_vba=True)
    ws = wb["GUIA"]
    _llenar_encabezado_informe(ws, codigo, perf_codigo, muestra, project, observaciones_ensayo, perf_numero_cell="E12")

    if gran_data is not None:
        masa_inicial_seca = to_float(gran_data.get("masa_inicial_seca"))
        if masa_inicial_seca is not None:
            ws["D17"] = masa_inicial_seca
        # Si solo se solicitó Pasa 200 (sin Granulometría), `gran_data` no trae los tamices en
        # absoluto (nunca pasó por el formulario de Granulometría) — ahí se dejan las celdas tal
        # cual trae la plantilla, para no simular una curva granulométrica falsa ("pasa 100%" en
        # todos los tamices). Pero si el tamiz SÍ se digitó (el ensayo de Granulometría existe y
        # su formulario ya se abrió), un tamiz que quedó en blanco significa "no quedó nada
        # retenido ahí" y se escribe como 0, igual que en la vista de solo lectura de la app.
        for key, _label, _apert, cell in SIEVES:
            if key not in gran_data:
                continue
            valor = to_float(gran_data.get(key))
            ws[cell] = valor if valor is not None else 0

    if lim_data is not None:
        _escribir_limites(ws, lim_data)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    data = _reparar_graficos_perdidos(bio.getvalue(), TEMPLATE_GRANULOMETRIA)
    return _restaurar_imagenes_perdidas(data, TEMPLATE_GRANULOMETRIA)


def generar_excel_granulometria(codigo, perf_codigo, muestra, project, data, observaciones_ensayo=""):
    lim_assay = get_assay(muestra["id_unico"], "limites")
    lim_data = lim_assay.get("data", {}) if lim_assay else None
    return _generar_excel_clasificacion(codigo, perf_codigo, muestra, project, gran_data=data, lim_data=lim_data,
                                         observaciones_ensayo=observaciones_ensayo)


def generar_excel_pasa200(codigo, perf_codigo, muestra, project, data, observaciones_ensayo=""):
    """Pasa 200 usa la misma plantilla de Granulometría. `data` ya viene resuelto por
    render_assay_form: si la muestra también tiene Granulometría, es el mismo diccionario de
    esa muestra (con tamices incluidos si se digitaron); si no, son solo los datos propios de
    Pasa 200."""
    lim_assay = get_assay(muestra["id_unico"], "limites")
    lim_data = lim_assay.get("data", {}) if lim_assay else None
    return _generar_excel_clasificacion(codigo, perf_codigo, muestra, project, gran_data=data, lim_data=lim_data,
                                         observaciones_ensayo=observaciones_ensayo)


def generar_excel_humedad(codigo, perf_codigo, muestra, project, data, observaciones_ensayo=""):
    wb = load_workbook(TEMPLATE_HUMEDAD)
    ws = wb["GUIA"]

    _llenar_encabezado_informe(ws, codigo, perf_codigo, muestra, project, observaciones_ensayo)
    ws["I19"] = data.get("hum_recipiente", "")
    ws["I20"] = to_float(data.get("hum_masa_humedo_mas_recipiente"))
    ws["I21"] = to_float(data.get("hum_seco_mas_recipiente"))
    ws["I22"] = to_float(data.get("hum_masa_recipiente"))
    # I23 (masa seca) e I24 (% humedad) son fórmulas de la propia plantilla; no se tocan.

    metodo = data.get("hum_metodo", "")
    ws["C28"] = "MÉTODO A" if metodo == "Método A" else ("MÉTODO B" if metodo == "Método B" else "")
    temp_horno = data.get("hum_temp_horno", "")
    ws["E28"] = "110°C" if "110" in temp_horno else ("60°C" if "60" in temp_horno else "")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return _restaurar_imagenes_perdidas(bio.getvalue(), TEMPLATE_HUMEDAD)


def generar_excel_limites(codigo, perf_codigo, muestra, project, data, observaciones_ensayo=""):
    gran_assay = get_assay(muestra["id_unico"], "granulometria")
    gran_data = gran_assay.get("data", {}) if gran_assay else None
    return _generar_excel_clasificacion(codigo, perf_codigo, muestra, project, gran_data=gran_data, lim_data=data,
                                         observaciones_ensayo=observaciones_ensayo)


def _llenar_encabezado_masa_unitaria(ws, codigo, perf_codigo, muestra, project, observaciones_ensayo=""):
    """La plantilla de Peso Unitario Parafinado (GDA-FLC-004) usa una distribución de celdas
    de encabezado propia, distinta a la de Granulometría/Límites/Humedad."""
    ws["C6"] = project.get("cliente", "") if project else ""  # Cliente
    ws["C7"] = project["nombre"] if project else codigo  # Proyecto
    ws["C8"] = project.get("correo_cliente", "") if project else ""  # Correo electrónico
    ws["C9"] = project.get("localizacion", "") if project else ""  # Localización
    ws["C10"] = project.get("muestra_tomada_por", "") if project else ""  # Muestra tomada por
    ws["I6"] = _fecha_ddmmaaaa(project.get("fecha_recepcion", "")) if project else ""  # Fecha de recepción
    ws["I7"] = _fecha_ddmmaaaa(project.get("fecha_ejecucion", "")) if project else ""  # Fecha de ejecución
    ws["I8"] = _fecha_ddmmaaaa(project.get("fecha_emision", "")) if project else ""  # Fecha de emisión
    ws["J9"] = project.get("numero", "") if project else ""  # Código interno — número (I9 ya trae "GDA")
    ws["K9"] = project.get("anio", "") if project else ""  # Código interno — año

    perf = get_perforacion(codigo, perf_codigo)
    ws["C12"] = TIPO_PERFORACION_EXCEL.get(perf["tipo"], "") if perf else ""  # Tipo de perforación (lista desplegable)
    ws["D12"] = perf["consecutivo"] if perf else ""  # Número de perforación
    ws["F12"] = muestra["numero"]  # Muestra No.
    ws["I12"] = to_float(muestra.get("profundidad_de"))
    ws["K12"] = to_float(muestra.get("profundidad_hasta"))
    ws["C13"] = descripcion_visual_para_excel(muestra) or observaciones_ensayo or f"Tipo de muestra: {muestra.get('tipo_muestra','')}"


def generar_excel_masa_unitaria(codigo, perf_codigo, muestra, project, data, observaciones_ensayo=""):
    wb = load_workbook(TEMPLATE_MASA_UNITARIA)
    ws = wb["GUIA"]
    _llenar_encabezado_masa_unitaria(ws, codigo, perf_codigo, muestra, project, observaciones_ensayo)

    ws["E20"] = to_float(data.get("mu_peso_aire"))  # B = masa en el aire
    ws["F20"] = to_float(data.get("mu_peso_aire_par"))  # C = masa en el aire parafinado
    ws["G20"] = to_float(data.get("mu_peso_agua_par"))  # D = masa parafinada sumergida
    # La densidad de la parafina (L24) no se digita en la app — se deja el 0.86 por defecto
    # que ya trae la plantilla.

    # La humedad (G28) la necesita la fórmula de "densidad seca" pero esta plantilla no la
    # digita — se toma del ensayo de Humedad de la misma muestra, igual que Granulometría y
    # Límites comparten datos entre sí.
    hum_assay = get_assay(muestra["id_unico"], "humedad")
    humedad_pct = calcular_humedad_pct(hum_assay.get("data", {})) if hum_assay else None
    if humedad_pct is not None:
        ws["G28"] = humedad_pct
    # A (masa de la cuerda, D20) y temperatura del agua no tienen celda equivalente en esta
    # plantilla — quedan para completar manualmente en el Excel.

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return _restaurar_imagenes_perdidas(bio.getvalue(), TEMPLATE_MASA_UNITARIA)


def _llenar_encabezado_cbr(ws, codigo, perf_codigo, muestra, project, observaciones_ensayo=""):
    """La plantilla oficial de CBR (GDA-FLC-013) usa una distribución de encabezado propia,
    distinta a la de Granulometría/Límites/Humedad/Masa Unitaria: etiquetas en B/H, valores en
    C/J en vez de D/K (ver hoja "GUIA" de GDA-FLC-013 CBR INALTERADO.xlsx)."""
    ws["C6"] = project.get("cliente", "") if project else ""  # Cliente
    ws["C7"] = project["nombre"] if project else codigo  # Proyecto
    ws["C8"] = project.get("correo_cliente", "") if project else ""  # Correo electrónico
    ws["C9"] = project.get("localizacion", "") if project else ""  # Localización
    ws["C10"] = project.get("muestra_tomada_por", "") if project else ""  # Muestra tomada por
    ws["J6"] = _fecha_ddmmaaaa(project.get("fecha_recepcion", "")) if project else ""  # Fecha de recepción
    ws["J7"] = _fecha_ddmmaaaa(project.get("fecha_ejecucion", "")) if project else ""  # Fecha de ejecución
    ws["J8"] = _fecha_ddmmaaaa(project.get("fecha_emision", "")) if project else ""  # Fecha de emisión
    ws["K9"] = project.get("numero", "") if project else ""  # Código interno — número (J9 ya trae "GDA")
    ws["L9"] = project.get("anio", "") if project else ""  # Código interno — año

    perf = get_perforacion(codigo, perf_codigo)
    ws["D12"] = TIPO_PERFORACION_EXCEL.get(perf["tipo"], "") if perf else ""  # Tipo de perforación (lista desplegable)
    ws["F12"] = perf["consecutivo"] if perf else ""  # Número de perforación
    ws["H12"] = muestra["numero"]  # Muestra No.
    prof_de, prof_hasta = to_float(muestra.get("profundidad_de")), to_float(muestra.get("profundidad_hasta"))
    ws["K12"] = f"{prof_de}-{prof_hasta}" if prof_de is not None and prof_hasta is not None else ""
    ws["C13"] = descripcion_visual_para_excel(muestra) or observaciones_ensayo or f"Tipo de muestra: {muestra.get('tipo_muestra','')}"


def generar_excel_cbr(codigo, perf_codigo, muestra, project, data, observaciones_ensayo=""):
    """CBR (GDA-FLC-013, INV E-148-13). La plantilla real trae ~20 hojas ocultas más (proyectos
    anteriores, cada uno una copia de "GUIA" ya diligenciada) que no tienen nada que ver con esta
    muestra — se descartan todas menos "GUIA" antes de guardar, para no mandar datos de otros
    clientes en cada descarga ni inflar el archivo."""
    wb = load_workbook(TEMPLATE_CBR)
    for nombre in list(wb.sheetnames):
        if nombre != "GUIA":
            del wb[nombre]
    ws = wb["GUIA"]
    _llenar_encabezado_cbr(ws, codigo, perf_codigo, muestra, project, observaciones_ensayo)

    ws["D20"] = data.get("cbr_molde", "")
    ws["D21"] = to_float(data.get("cbr_diametro"))
    ws["D22"] = to_float(data.get("cbr_altura"))
    ws["D24"] = to_float(data.get("cbr_masa_molde"))
    ws["D23"] = to_float(data.get("cbr_masa_muestra_molde_antes"))
    ws["E23"] = to_float(data.get("cbr_masa_muestra_molde_despues"))

    # Humedad antes de inmersión: no se digita en el formulario de CBR, se comparte con el
    # ensayo de Contenido de Humedad de la misma muestra (ver render_cbr_form) — se copia acá
    # porque la plantilla real necesita el valor en su propia celda, no una referencia cruzada
    # a otro archivo.
    hum_assay = get_assay(muestra["id_unico"], "humedad")
    hum_data = hum_assay.get("data", {}) if hum_assay else {}
    ws["D30"] = hum_data.get("hum_recipiente", "")
    ws["D31"] = to_float(hum_data.get("hum_masa_humedo_mas_recipiente"))
    ws["D32"] = to_float(hum_data.get("hum_seco_mas_recipiente"))
    ws["D33"] = to_float(hum_data.get("hum_masa_recipiente"))

    ws["E30"] = data.get("cbr_desp_recipiente", "")
    ws["E31"] = to_float(data.get("cbr_desp_masa_humedo"))
    ws["E32"] = to_float(data.get("cbr_desp_masa_seco"))
    ws["E33"] = to_float(data.get("cbr_desp_masa_recipiente"))

    ws["D37"] = to_float(data.get("cbr_exp_lectura_inicial"))
    ws["D38"] = to_float(data.get("cbr_exp_lectura_final"))

    # Pesas de sobrecarga y tiempo de inmersión: la plantilla ya trae un valor por defecto
    # (4554 g / 4 días) — solo se pisa si el laboratorista digitó algo distinto.
    pesas_antes = to_float(data.get("cbr_pesas_antes"))
    if pesas_antes is not None:
        ws["I39"] = pesas_antes
    pesas_despues = to_float(data.get("cbr_pesas_despues"))
    if pesas_despues is not None:
        ws["I40"] = pesas_despues
    tiempo_antes = to_float(data.get("cbr_tiempo_inmersion_antes"))
    if tiempo_antes is not None:
        ws["J39"] = tiempo_antes
    tiempo_despues = to_float(data.get("cbr_tiempo_inmersion_despues"))
    if tiempo_despues is not None:
        ws["J40"] = tiempo_despues

    # Tabla de penetración: fila 22 es la profundidad "0" (fuerza 0 fija, no se digita); las 12
    # profundidades reales (CBR_PENETRACION_FILAS) caen en las filas 23 a 34, en el mismo orden.
    for i in range(1, len(CBR_PENETRACION_FILAS) + 1):
        fila = 22 + i
        antes = to_float(data.get(f"cbr_pen_antes_{i}"))
        if antes is not None:
            ws[f"I{fila}"] = antes
        despues = to_float(data.get(f"cbr_pen_despues_{i}"))
        if despues is not None:
            ws[f"K{fila}"] = despues

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return _restaurar_imagenes_perdidas(bio.getvalue(), TEMPLATE_CBR)


# ════════════════════════════════════════════════════════════════════
# FORMULARIOS DE ENSAYO (solo captura de datos, sin cálculos)
# ════════════════════════════════════════════════════════════════════
def render_equipo(data, prefix, equipo_list=None):
    lista = equipo_list or EQUIPO_LIST
    with st.container(border=True):
        st.markdown(card_header_html("construction", "Equipos Utilizados"), unsafe_allow_html=True)
        seleccionados = set(data.get(f"{prefix}_equipos", []))
        cols = st.columns(2)
        nuevos = []
        for i, equipo in enumerate(lista):
            with cols[i % 2]:
                if st.checkbox(equipo, value=equipo in seleccionados, key=f"{prefix}_equipo_{i}"):
                    nuevos.append(equipo)
        data[f"{prefix}_equipos"] = nuevos


def render_norma_selector(assay_type, data, key_prefix):
    with st.container(border=True):
        st.markdown(card_header_html("rule", "Norma a Utilizar"), unsafe_allow_html=True)
        options = NORMAS_ENSAYO[assay_type]
        current = data.get(f"{key_prefix}_norma", "")
        idx = options.index(current) if current in options else 0
        choice = st.selectbox("Norma", options, index=idx, key=f"norma_{key_prefix}", label_visibility="collapsed")
        data[f"{key_prefix}_norma"] = choice


PASA_200_FILAS = [
    ("p200_recipiente", "Recipiente No."),
    ("p200_seco_mas_recipiente", "Masa suelo seco + recipiente"),
    ("p200_seco_14h", "Masa suelo seco (14 hrs)"),
    ("p200_seco_15h", "Masa suelo seco (15 hrs)"),
    ("p200_seco_16h", "Masa suelo seco (16 hrs)"),
    ("p200_masa_recipiente", "Masa del recipiente"),
]

# ════════════════════════════════════════════════════════════════════
# CAMPOS REQUERIDOS POR ENSAYO — antes de "Enviar a revisión" se valida que estén digitados; si
# falta alguno, no se manda y ese campo se resalta en rojo (ver render_assay_form/campos_faltantes).
# Los 14h/15h/16h de Pasa 200 y Humedad NO están acá: son solo lecturas de verificación que se
# autocompletan de la lectura base, no algo que el laboratorista tenga que digitar aparte. Igual
# los tamices de Granulometría: un tamiz vacío es un dato válido (nada quedó retenido ahí), no un
# dato faltante — ver el "0 en vez de guion" ya implementado para esa tabla.
# ════════════════════════════════════════════════════════════════════
CAMPOS_REQUERIDOS_PASA200 = [
    (f"p200_{campo}_{suf}", f"{label} ({'antes' if suf == 'antes' else 'después'} del lavado)")
    for campo, label in (("recipiente", "Recipiente No."), ("seco_mas_recipiente", "Masa suelo seco + recipiente"),
                          ("masa_recipiente", "Masa del recipiente"))
    for suf in ("antes", "despues")
]
CAMPOS_REQUERIDOS_HUMEDAD = [
    ("hum_recipiente", "Recipiente No."),
    ("hum_masa_recipiente", "Masa del recipiente (g)"),
    ("hum_masa_humedo_mas_recipiente", "Masa suelo húmedo + recipiente (g)"),
    ("hum_seco_mas_recipiente", "Masa suelo seco + recipiente (g)"),
]
CAMPOS_REQUERIDOS_LIMITES = (
    [(f"lim_ll_{campo}_{i}", f"{label} — Límite Líquido, ensayo {i}")
     for campo, label in (("recipiente", "Recipiente No."), ("golpes", "No. de golpes"),
                           ("humedo", "Masa húmedo + recipiente"), ("seco", "Masa seco + recipiente"),
                           ("recip_masa", "Masa recipiente"))
     for i in range(1, LIMITE_LIQUIDO_N + 1)]
    + [(f"lim_lp_{campo}_{i}", f"{label} — Límite Plástico, ensayo {i}")
       for campo, label in (("recipiente", "Recipiente No."), ("humedo", "Masa húmedo + recipiente"),
                             ("seco", "Masa seco + recipiente"), ("recip_masa", "Masa recipiente"))
       for i in range(1, LIMITE_PLASTICO_N + 1)]
)
CAMPOS_REQUERIDOS_MASA_UNITARIA = [
    ("mu_peso_aire", "Masa en el aire (g)"),
    ("mu_peso_agua_par", "Masa en el agua parafinado (g)"),
    ("mu_peso_aire_par", "Masa en el aire parafinado (g)"),
    ("mu_temp_agua", "Temperatura del agua (°C)"),
]
# Reconstruido a partir de la plantilla oficial real (GDA-FLC-013 CBR INALTERADO.xlsx, hoja
# "GUIA", INV E-148-13) — la primera versión de este formulario se armó a partir de una captura
# de pantalla que resultó no coincidir con el formato oficial (traía "No. de golpes"/"No. de
# capas", que no existen en la plantilla real, y una humedad "después de inmersión" con lecturas
# de verificación a 16/17/18/19 horas que tampoco están ahí). Ver TEMPLATE_CBR/generar_excel_cbr
# para el mapeo exacto celda por celda.
# La "Humedad antes de inmersión" NO está acá ni se digita en el formulario del CBR: se comparte
# con el ensayo de Contenido de Humedad de la misma muestra (ver render_cbr_form) y se copia sola
# a la plantilla al exportar — si falta, se marca como faltante allá, no acá. La tabla de
# penetración (24 lecturas de fuerza) y las "Condiciones del Ensayo" (pesas/tiempo de inmersión,
# que ya traen un valor por defecto en la plantilla) tampoco son obligatorias.
CAMPOS_REQUERIDOS_CBR = [
    ("cbr_molde", "Molde No."),
    ("cbr_diametro", "Diámetro de la muestra (cm)"),
    ("cbr_altura", "Altura de la muestra (cm)"),
    ("cbr_masa_molde", "Masa molde (g)"),
    ("cbr_masa_muestra_molde_antes", "Masa de la muestra + molde (g) — antes de inmersión"),
    ("cbr_masa_muestra_molde_despues", "Masa de la muestra + molde (g) — después de inmersión"),
    ("cbr_desp_recipiente", "Recipiente — humedad después de inmersión"),
    ("cbr_desp_masa_humedo", "Peso recipiente + suelo húmedo (g) — después de inmersión"),
    ("cbr_desp_masa_seco", "Peso recipiente + suelo seco (g) — después de inmersión"),
    ("cbr_desp_masa_recipiente", "Peso recipiente (g) — después de inmersión"),
    ("cbr_exp_lectura_inicial", "Lectura inicial (in) — expansión"),
    ("cbr_exp_lectura_final", "Lectura final (in) — expansión"),
]
# Las 12 profundidades de penetración estándar de INV E-148 (pulgadas), con su equivalente en mm
# tal como aparece impreso en la plantilla — no se digitan, son fijas; lo que se digita es la
# Fuerza (kN) leída en cada una, antes y después de inmersión (ver render_cbr_form).
CBR_PENETRACION_FILAS = [
    ("0.005", "0.127"), ("0.025", "0.635"), ("0.05", "1.27"), ("0.075", "1.905"),
    ("0.1", "2.54"), ("0.125", "3.175"), ("0.15", "3.81"), ("0.175", "4.445"),
    ("0.2", "5.08"), ("0.3", "7.62"), ("0.4", "10.16"), ("0.5", "12.7"),
]
CAMPOS_REQUERIDOS_POR_TIPO = {
    "humedad": CAMPOS_REQUERIDOS_HUMEDAD,
    "pasa200": CAMPOS_REQUERIDOS_PASA200,
    "granulometria": CAMPOS_REQUERIDOS_PASA200,  # comparte los mismos campos de Pasa 200 (embebido y "Requerido")
    "limites": CAMPOS_REQUERIDOS_LIMITES,
    "masa-unitaria": CAMPOS_REQUERIDOS_MASA_UNITARIA,
    "cbr": CAMPOS_REQUERIDOS_CBR,
}


def campos_faltantes(tipo, data):
    """Campos requeridos que todavía están vacíos para este tipo de ensayo — lista de
    (clave, etiqueta) en el mismo orden en que se digitan en el formulario."""
    return [(key, label) for key, label in CAMPOS_REQUERIDOS_POR_TIPO.get(tipo, [])
            if not str(data.get(key, "")).strip()]


def render_pasa200_section(data, assay_id, requerido=True):
    """Campos de "Determinación Pasa No. 200". Se usa tanto embebido dentro del formulario de
    Granulometría como en el formulario del ensayo "Pasa 200" independiente — en ambos casos
    `data` puede terminar siendo el mismo diccionario compartido (ver render_assay_form), así
    que lo que se digite en cualquiera de las dos pantallas se refleja en la otra."""
    badge = '<span class="badge badge-warning">Requerido</span>' if requerido else ""
    with st.container(border=True):
        st.markdown(card_header_html("water_drop", "Determinación Pasa No. 200", badge), unsafe_allow_html=True)
        head = st.columns([2.2, 1, 1])
        head[1].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Antes del lavado (g)</div>', unsafe_allow_html=True)
        head[2].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Después del lavado (g)</div>', unsafe_allow_html=True)
        for key, label in PASA_200_FILAS:
            row = st.columns([2.2, 1, 1])
            row[0].markdown(f'<div style="padding-top:8px;">{label}</div>', unsafe_allow_html=True)
            for suffix, col in (("antes", row[1]), ("despues", row[2])):
                field_key = f"{key}_{suffix}"
                widget_key = f"{field_key}_{assay_id}"
                # No se pasa `value=` junto con un key que también se controla por session_state
                # (el autocompletado de abajo lo hace) — Streamlit no permite mezclar ambos.
                if widget_key not in st.session_state:
                    st.session_state[widget_key] = data.get(field_key, "")
                data[field_key] = col.text_input(
                    f"{label} {suffix}", key=widget_key, label_visibility="collapsed", placeholder="0.00")
            if key == "p200_seco_mas_recipiente":
                # Autocompleta las 3 lecturas de horas con esta masa apenas se digita, pero si el
                # laboratorista ya las cambió a mano, no se vuelven a pisar en el siguiente rerun —
                # solo se repite el autocompletado cuando el valor de origen vuelve a cambiar.
                for suffix in ("antes", "despues"):
                    src_key = f"{key}_{suffix}"
                    current_val = data[src_key]
                    lastsync_key = f"{src_key}_lastsync"
                    if data.get(lastsync_key) != current_val:
                        for hkey in ("p200_seco_14h", "p200_seco_15h", "p200_seco_16h"):
                            st.session_state[f"{hkey}_{suffix}_{assay_id}"] = current_val
                            data[f"{hkey}_{suffix}"] = current_val
                        data[lastsync_key] = current_val

        # La plantilla de Excel necesita la masa inicial seca (neta, sin el recipiente) para calcular
        # el % que pasa cada tamiz. Se deriva de las lecturas "antes del lavado" ya digitadas arriba
        # (no es un campo nuevo en la interfaz, solo cómo se arma el dato para el Excel).
        masa_seco_mas_recip = to_float(data.get("p200_seco_mas_recipiente_antes"))
        masa_recip = to_float(data.get("p200_masa_recipiente_antes"))
        data["masa_inicial_seca"] = (masa_seco_mas_recip - masa_recip) if (masa_seco_mas_recip is not None and masa_recip is not None) else ""


def render_pasa200_form(data, assay_id):
    st.info("Estos datos se guardan tal cual y se llevan a la plantilla oficial de Excel de Granulometría — "
            "si la muestra también tiene Granulometría, los dos ensayos comparten los mismos datos.")
    render_norma_selector("granulometria", data, "gran")
    render_equipo(data, "gran", EQUIPO_GRANULOMETRIA)
    render_pasa200_section(data, assay_id, requerido=False)


def render_granulometria_form(data, assay_id):
    st.info("Estos datos se guardan tal cual y se llevan a la plantilla oficial de Excel — los cálculos y la clasificación USCS los hace el Excel, no la app.")

    render_norma_selector("granulometria", data, "gran")
    render_equipo(data, "gran", EQUIPO_GRANULOMETRIA)

    render_pasa200_section(data, assay_id)

    with st.container(border=True):
        st.markdown(card_header_html("grid_view", "Granulometría (Masa de Suelo Retenido)"), unsafe_allow_html=True)
        # Campos de texto libre (no st.data_editor con NumberColumn) porque el editor de tabla
        # de Streamlit borra el punto decimal apenas se escribe más de un dígito después de él
        # — el mismo patrón confiable que ya se usa en Pasa No. 200 y en Humedad.
        head = st.columns([1.2, 1, 1.4])
        head[0].markdown('<div class="cell-muted" style="font-weight:700;">Tamiz</div>', unsafe_allow_html=True)
        head[1].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Abertura (mm)</div>', unsafe_allow_html=True)
        head[2].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Retenido (g)</div>', unsafe_allow_html=True)
        for key, label, apert, _cell in SIEVES:
            row = st.columns([1.2, 1, 1.4])
            row[0].markdown(f'<div style="padding-top:8px;">{label}</div>', unsafe_allow_html=True)
            row[1].markdown(f'<div style="padding-top:8px;text-align:center;">{apert}</div>', unsafe_allow_html=True)
            widget_key = f"retenido_{key}_{assay_id}"
            if widget_key not in st.session_state:
                # Ensayos guardados antes de este cambio pueden tener el valor como float
                # (venía del st.data_editor con NumberColumn) — text_input necesita un str.
                raw = data.get(key, "")
                st.session_state[widget_key] = "" if raw in (None, "") else str(raw)
            data[key] = row[2].text_input(f"Retenido {label}", key=widget_key, label_visibility="collapsed", placeholder="0.00")
        st.caption("El % retenido y la clasificación USCS se calculan en la plantilla de Excel, no aquí.")


def render_humedad_form(data, assay_id):
    st.info("Estos datos se guardan tal cual y se llevan a la plantilla oficial de Excel — el % de humedad lo calcula el Excel, no la app.")

    render_norma_selector("humedad", data, "hum")
    render_equipo(data, "hum", EQUIPO_HUMEDAD)

    with st.container(border=True):
        st.markdown(card_header_html("science", "Determinación de Humedad"), unsafe_allow_html=True)

        def _campo(key, label, placeholder="0.00"):
            row = st.columns([2.2, 1])
            row[0].markdown(f'<div style="padding-top:8px;">{label}</div>', unsafe_allow_html=True)
            data[key] = row[1].text_input(label, value=data.get(key, ""), key=f"{key}_{assay_id}",
                                           label_visibility="collapsed", placeholder=placeholder)

        _campo("hum_recipiente", "Recipiente no.", placeholder="839")
        _campo("hum_masa_recipiente", "Masa del recipiente (g)")
        _campo("hum_masa_humedo_mas_recipiente", "Masa suelo húmedo + recipiente (g)")

        # "Masa suelo seco + recipiente" se autocompleta en las 3 lecturas de horas apenas se
        # digita, igual que en Pasa No. 200 de Granulometría — si el laboratorista cambia una
        # lectura a mano, ya no se vuelve a pisar hasta que el valor de origen vuelva a cambiar.
        src_key = "hum_seco_mas_recipiente"
        src_widget_key = f"{src_key}_{assay_id}"
        if src_widget_key not in st.session_state:
            st.session_state[src_widget_key] = data.get(src_key, "")
        row = st.columns([2.2, 1])
        row[0].markdown('<div style="padding-top:8px;">Masa suelo seco + recipiente (g)</div>', unsafe_allow_html=True)
        data[src_key] = row[1].text_input("Masa suelo seco + recipiente (g)", key=src_widget_key,
                                           label_visibility="collapsed", placeholder="0.00")
        current_val = data[src_key]
        lastsync_key = f"{src_key}_lastsync"
        if data.get(lastsync_key) != current_val:
            for hkey in ("hum_seco_14h", "hum_seco_15h", "hum_seco_16h"):
                st.session_state[f"{hkey}_{assay_id}"] = current_val
                data[hkey] = current_val
            data[lastsync_key] = current_val
        for hkey, hlabel in (("hum_seco_14h", "Masa suelo seco + recipiente (g) (14 hrs)"),
                              ("hum_seco_15h", "Masa suelo seco + recipiente (g) (15 hrs)"),
                              ("hum_seco_16h", "Masa suelo seco + recipiente (g) (16 hrs)")):
            hwidget_key = f"{hkey}_{assay_id}"
            if hwidget_key not in st.session_state:
                st.session_state[hwidget_key] = data.get(hkey, "")
            hrow = st.columns([2.2, 1])
            hrow[0].markdown(f'<div style="padding-top:8px;">{hlabel}</div>', unsafe_allow_html=True)
            data[hkey] = hrow[1].text_input(hlabel, key=hwidget_key, label_visibility="collapsed", placeholder="0.00")

        st.caption("La masa del agua y la masa de suelo seco se calculan solas (restando la masa del recipiente) — no se digitan aquí.")

    with st.container(border=True):
        st.markdown(card_header_html("local_fire_department", "Datos del Laboratorio"), unsafe_allow_html=True)
        temp_actual = data.get("hum_temp_horno", "110 ± 5 °C")
        opciones_temp = ["110 ± 5 °C", "60 °C"]
        idx = opciones_temp.index(temp_actual) if temp_actual in opciones_temp else 0
        c1, c2 = st.columns(2)
        with c1:
            data["hum_temp_horno"] = st.selectbox("Temperatura Horno", opciones_temp, index=idx, key=f"hum_temp_horno_{assay_id}")
        with c2:
            metodo_actual = data.get("hum_metodo", METODO_HUMEDAD[0])
            midx = METODO_HUMEDAD.index(metodo_actual) if metodo_actual in METODO_HUMEDAD else 0
            data["hum_metodo"] = st.selectbox("Método del Ensayo", METODO_HUMEDAD, index=midx, key=f"hum_metodo_{assay_id}")


def render_masa_unitaria_form(data, assay_id):
    st.info("Estos datos se guardan tal cual, sin calcular el peso unitario dentro de la app.")
    c1, c2 = st.columns(2)
    with c1:
        data["mu_peso_aire"] = st.text_input("Masa en el aire (g)", value=data.get("mu_peso_aire", ""),
                                              key=f"mu_peso_aire_{assay_id}", placeholder="245.80")
        data["mu_peso_agua_par"] = st.text_input("Masa en el agua parafinado (g)", value=data.get("mu_peso_agua_par", ""),
                                                  key=f"mu_peso_agua_par_{assay_id}", placeholder="138.20")
    with c2:
        data["mu_peso_aire_par"] = st.text_input("Masa en el aire parafinado (g)", value=data.get("mu_peso_aire_par", ""),
                                                  key=f"mu_peso_aire_par_{assay_id}", placeholder="258.30")
        data["mu_temp_agua"] = st.text_input("Temperatura del agua (°C)", value=data.get("mu_temp_agua", ""),
                                              key=f"mu_temp_agua_{assay_id}", placeholder="22.0")

    render_equipo(data, "mu", EQUIPO_MASA_UNITARIA)
    render_norma_selector("masa-unitaria", data, "mu")


def render_cbr_form(data, assay_id, muestra_id):
    st.info("Estos datos se guardan tal cual y se llevan a la plantilla oficial de Excel — el CBR a 0.1\" y 0.2\" "
            "de penetración, igual que el resto de valores calculados, los saca el Excel, no la app.")

    def _campo(key, label, placeholder="0.00"):
        row = st.columns([2.2, 1])
        row[0].markdown(f'<div style="padding-top:8px;">{label}</div>', unsafe_allow_html=True)
        data[key] = row[1].text_input(label, value=data.get(key, ""), key=f"{key}_{assay_id}",
                                       label_visibility="collapsed", placeholder=placeholder)

    def _campo_antes_despues(key_base, label):
        """Un mismo dato con lectura antes Y después de inmersión (ej. masa de la muestra +
        molde, que cambia porque la muestra absorbe agua) — dos casillas lado a lado."""
        row = st.columns([2, 1, 1])
        row[0].markdown(f'<div style="padding-top:8px;">{label}</div>', unsafe_allow_html=True)
        data[f"{key_base}_antes"] = row[1].text_input(f"{label} (antes)", value=data.get(f"{key_base}_antes", ""),
                                                        key=f"{key_base}_antes_{assay_id}", label_visibility="collapsed",
                                                        placeholder="Antes")
        data[f"{key_base}_despues"] = row[2].text_input(f"{label} (después)", value=data.get(f"{key_base}_despues", ""),
                                                          key=f"{key_base}_despues_{assay_id}", label_visibility="collapsed",
                                                          placeholder="Después")

    with st.container(border=True):
        st.markdown(card_header_html("science", "Datos Iniciales"), unsafe_allow_html=True)
        st.caption("Molde, diámetro, altura y masa del molde son los mismos antes y después de inmersión — "
                   "solo la masa de la muestra + molde cambia (la muestra absorbe agua).")
        _campo("cbr_molde", "Molde No.", placeholder="1")
        _campo("cbr_diametro", "Diámetro de la muestra (cm)")
        _campo("cbr_altura", "Altura de la muestra (cm)")
        _campo("cbr_masa_molde", "Masa molde (g)")
        head = st.columns([2, 1, 1])
        head[1].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Antes</div>', unsafe_allow_html=True)
        head[2].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Después</div>', unsafe_allow_html=True)
        _campo_antes_despues("cbr_masa_muestra_molde", "Masa de la muestra + molde (g)")

    with st.container(border=True):
        st.markdown(card_header_html("water_drop", "Humedad antes de inmersión"), unsafe_allow_html=True)
        st.caption("Se comparte con el ensayo de Contenido de Humedad de esta muestra — no es un dato "
                   "propio del CBR. Si falta o está mal, corrígelo desde ese ensayo, no desde aquí.")
        hum_assay = get_assay(muestra_id, "humedad")
        hum_data = hum_assay.get("data", {}) if hum_assay else {}
        if hum_data.get("hum_recipiente") or hum_data.get("hum_seco_mas_recipiente"):
            rows = [
                ("Recipiente no.", hum_data.get("hum_recipiente")),
                ("Peso recipiente + suelo húmedo (g)", hum_data.get("hum_masa_humedo_mas_recipiente")),
                ("Peso recipiente + suelo seco (g)", hum_data.get("hum_seco_mas_recipiente")),
                ("Peso recipiente (g)", hum_data.get("hum_masa_recipiente")),
                ("Humedad (%)", fmt_num(calcular_humedad_pct(hum_data), decimals=2)),
            ]
            st.markdown(param_table_html(rows), unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="display:flex;align-items:center;gap:6px;color:{NEUTRAL};font-style:italic;">'
                         f'{icon("visibility_off", size=16)} El ensayo de Contenido de Humedad de esta muestra '
                         f'todavía no tiene datos</div>', unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown(card_header_html("water_drop", "Humedad después de inmersión"), unsafe_allow_html=True)
        _campo("cbr_desp_recipiente", "Recipiente", placeholder="839")
        _campo("cbr_desp_masa_humedo", "Peso recipiente + suelo húmedo (g)")
        _campo("cbr_desp_masa_seco", "Peso recipiente + suelo seco (g)")
        _campo("cbr_desp_masa_recipiente", "Peso recipiente (g)")

    with st.container(border=True):
        st.markdown(card_header_html("straighten", "Datos de Expansión"), unsafe_allow_html=True)
        _campo("cbr_exp_lectura_inicial", "Lectura inicial (in)")
        _campo("cbr_exp_lectura_final", "Lectura final (in)")

    with st.container(border=True):
        st.markdown(card_header_html("tune", "Condiciones del Ensayo"), unsafe_allow_html=True)
        st.caption("La plantilla ya trae un valor por defecto (4554 g / 4 días) — solo digita algo aquí si es distinto.")
        head = st.columns([2, 1, 1])
        head[1].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Antes</div>', unsafe_allow_html=True)
        head[2].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Después</div>', unsafe_allow_html=True)
        _campo_antes_despues("cbr_pesas", "Pesas de sobrecarga (g)")
        _campo_antes_despues("cbr_tiempo_inmersion", "Tiempo de inmersión (días)")

    with st.container(border=True):
        st.markdown(card_header_html("show_chart", "Penetración"), unsafe_allow_html=True)
        st.caption("Fuerza (kN) leída en cada profundidad — el esfuerzo (MPa) y el CBR a 0.1\"/0.2\" los calcula el Excel.")
        head = st.columns([1.2, 1, 1])
        head[0].markdown('<div class="cell-muted" style="font-weight:700;">Profundidad</div>', unsafe_allow_html=True)
        head[1].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Fuerza antes (kN)</div>', unsafe_allow_html=True)
        head[2].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Fuerza después (kN)</div>', unsafe_allow_html=True)
        for i, (pulg, mm) in enumerate(CBR_PENETRACION_FILAS, start=1):
            row = st.columns([1.2, 1, 1])
            row[0].markdown(f'<div style="padding-top:8px;">{pulg}" ({mm} mm)</div>', unsafe_allow_html=True)
            data[f"cbr_pen_antes_{i}"] = row[1].text_input(f"Fuerza antes {pulg}in", value=data.get(f"cbr_pen_antes_{i}", ""),
                                                             key=f"cbr_pen_antes_{i}_{assay_id}", label_visibility="collapsed",
                                                             placeholder="kN")
            data[f"cbr_pen_despues_{i}"] = row[2].text_input(f"Fuerza después {pulg}in", value=data.get(f"cbr_pen_despues_{i}", ""),
                                                               key=f"cbr_pen_despues_{i}_{assay_id}", label_visibility="collapsed",
                                                               placeholder="kN")

    render_equipo(data, "cbr", EQUIPO_CBR)
    render_norma_selector("cbr", data, "cbr")


def render_limites_form(data, assay_id):
    st.info("Estos datos se guardan tal cual y se llevan a la plantilla oficial de Excel — el Límite Líquido, el Límite Plástico y el Índice de Plasticidad los calcula el Excel, no la app.")

    with st.container(border=True):
        st.markdown(card_header_html("info", "Información de Ensayo"), unsafe_allow_html=True)
        metodo_actual = data.get("lim_metodo", METODO_HUMEDAD[0])
        midx = METODO_HUMEDAD.index(metodo_actual) if metodo_actual in METODO_HUMEDAD else 0
        data["lim_metodo"] = st.radio("Método de Ensayo", METODO_HUMEDAD, index=midx, horizontal=True, key=f"lim_metodo_{assay_id}")

    def _tabla_limite(icono, titulo, filas, n):
        with st.container(border=True):
            st.markdown(card_header_html(icono, titulo), unsafe_allow_html=True)
            head = st.columns([2] + [1] * n)
            head[0].markdown('<div class="cell-muted" style="font-weight:700;">Parámetro</div>', unsafe_allow_html=True)
            for i in range(n):
                head[i + 1].markdown(f'<div class="cell-muted" style="text-align:center;font-weight:700;">Ensayo {i + 1}</div>', unsafe_allow_html=True)
            for key, label, _cells in filas:
                row = st.columns([2] + [1] * n)
                row[0].markdown(f'<div style="padding-top:8px;">{label}</div>', unsafe_allow_html=True)
                for i in range(n):
                    field_key = f"{key}_{i + 1}"
                    widget_key = f"{field_key}_{assay_id}"
                    if widget_key not in st.session_state:
                        raw = data.get(field_key, "")
                        st.session_state[widget_key] = "" if raw in (None, "") else str(raw)
                    data[field_key] = row[i + 1].text_input(f"{label} {i + 1}", key=widget_key, label_visibility="collapsed", placeholder="0.00")

                if key.endswith("_seco"):
                    # "Masa suelo seco + rec." se autocompleta en las lecturas de 14/15/16 horas
                    # de cada columna, igual que en Humedad y Pasa No. 200 — si el laboratorista
                    # cambia una lectura a mano, no se vuelve a pisar hasta que el valor de
                    # origen vuelva a cambiar.
                    for i in range(n):
                        field_key = f"{key}_{i + 1}"
                        current_val = data[field_key]
                        lastsync_key = f"{field_key}_lastsync"
                        if data.get(lastsync_key) != current_val:
                            for suffix in ("14h", "15h", "16h"):
                                hkey = f"{key}_{suffix}_{i + 1}"
                                st.session_state[f"{hkey}_{assay_id}"] = current_val
                                data[hkey] = current_val
                            data[lastsync_key] = current_val

    _tabla_limite("water_drop", "Límite Líquido (INV. 125 - 13)", LIMITE_LIQUIDO_FILAS, LIMITE_LIQUIDO_N)
    _tabla_limite("gesture", "Límite Plástico (INV. 126 - 13)", LIMITE_PLASTICO_FILAS, LIMITE_PLASTICO_N)

    render_equipo(data, "lim", EQUIPO_LIMITES)


def render_read_only_summary(tipo, data, laboratorista="—", muestra_id=None):
    """Vista de solo lectura ('Resultados de Ensayo') — la misma para el Jefe (siempre) y para
    el laboratorista cuando el proyecto ya fue ejecutado. Sin casillas de digitación, solo tarjetas
    y tablas con los datos ya registrados."""
    if tipo == "granulometria":
        # Toda orden de granulometría incluye el Pasa No. 200 — se muestra tal cual se digitó,
        # con sus dos columnas (antes/después del lavado), igual que en el formulario editable.
        with st.container(border=True):
            st.markdown(card_header_html("water_drop", "Determinación Pasa No. 200"), unsafe_allow_html=True)
            pasa200_rows = [(label, data.get(f"{key}_antes"), data.get(f"{key}_despues")) for key, label in PASA_200_FILAS]
            st.markdown(param_table_3col_html(pasa200_rows), unsafe_allow_html=True)
        # "Masa inicial seca" no se muestra en la app (ni aquí ni en el formulario editable) —
        # se deriva solo al momento de generar el Excel (ver generar_excel_granulometria), tal
        # como se llena a mano en la plantilla física: masa suelo seco + recipiente, menos recipiente.
        with st.container(border=True):
            st.markdown(card_header_html("grid_view", "Granulometría (Masa de Suelo Retenido)"), unsafe_allow_html=True)
            # Tamiz sin digitar = no se pesó nada retenido ahí, no un dato faltante -> se muestra 0.
            sieve_rows = [(label, data.get(key) if data.get(key) not in (None, "") else 0) for key, label, _apert, _cell in SIEVES]
            st.markdown(param_table_html(sieve_rows, header_left="TAMIZ", header_right="RETENIDO (g)"), unsafe_allow_html=True)
        equipos, norma = data.get("gran_equipos", []), data.get("gran_norma", "—")
    elif tipo == "pasa200":
        with st.container(border=True):
            st.markdown(card_header_html("water_drop", "Determinación Pasa No. 200"), unsafe_allow_html=True)
            pasa200_rows = [(label, data.get(f"{key}_antes"), data.get(f"{key}_despues")) for key, label in PASA_200_FILAS]
            st.markdown(param_table_3col_html(pasa200_rows), unsafe_allow_html=True)
        equipos, norma = data.get("gran_equipos", []), data.get("gran_norma", "—")
    elif tipo == "humedad":
        masa_humedo = to_float(data.get("hum_masa_humedo_mas_recipiente"))
        masa_seco = to_float(data.get("hum_seco_mas_recipiente"))
        masa_recip = to_float(data.get("hum_masa_recipiente"))
        masa_agua = (masa_humedo - masa_seco) if (masa_humedo is not None and masa_seco is not None) else None
        masa_suelo_seco = (masa_seco - masa_recip) if (masa_seco is not None and masa_recip is not None) else None
        humedad_pct = calcular_humedad_pct(data)
        rows = [
            ("Recipiente no.", data.get("hum_recipiente")),
            ("Masa del recipiente (g)", data.get("hum_masa_recipiente")),
            ("Masa suelo húmedo + recipiente (g)", data.get("hum_masa_humedo_mas_recipiente")),
            ("Masa suelo seco + recipiente (g) (14 hrs)", data.get("hum_seco_14h")),
            ("Masa suelo seco + recipiente (g) (15 hrs)", data.get("hum_seco_15h")),
            ("Masa suelo seco + recipiente (g) (16 hrs)", data.get("hum_seco_16h")),
            ("Masa del agua (g)", fmt_num(masa_agua)),
            ("Masa suelo seco (g)", fmt_num(masa_suelo_seco)),
            ("Humedad (%)", fmt_num(humedad_pct, decimals=2)),
        ]
        with st.container(border=True):
            st.markdown(card_header_html("science", "Parámetros Registrados"), unsafe_allow_html=True)
            st.markdown(param_table_html(rows), unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(card_header_html("local_fire_department", "Datos del Laboratorio"), unsafe_allow_html=True)
            lab_rows = [
                ("Temperatura Horno", data.get("hum_temp_horno")),
                ("Método del Ensayo", data.get("hum_metodo")),
                ("Laboratorista", laboratorista),
            ]
            st.markdown(param_table_html(lab_rows, header_left="DATO", header_right="VALOR"), unsafe_allow_html=True)
        equipos, norma = data.get("hum_equipos", []), data.get("hum_norma", "—")
    elif tipo == "limites":
        with st.container(border=True):
            st.markdown(card_header_html("water_drop", "Límite Líquido (INV. 125 - 13)"), unsafe_allow_html=True)
            headers = ["PARÁMETRO"] + [f"ENSAYO {i}" for i in range(1, LIMITE_LIQUIDO_N + 1)]
            ll_rows = [(label, *[data.get(f"{key}_{i}") for i in range(1, LIMITE_LIQUIDO_N + 1)]) for key, label, _c in LIMITE_LIQUIDO_FILAS]
            st.markdown(param_table_ncol_html(headers, ll_rows), unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(card_header_html("gesture", "Límite Plástico (INV. 126 - 13)"), unsafe_allow_html=True)
            headers = ["PARÁMETRO"] + [f"ENSAYO {i}" for i in range(1, LIMITE_PLASTICO_N + 1)]
            lp_rows = [(label, *[data.get(f"{key}_{i}") for i in range(1, LIMITE_PLASTICO_N + 1)]) for key, label, _c in LIMITE_PLASTICO_FILAS]
            st.markdown(param_table_ncol_html(headers, lp_rows), unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(card_header_html("info", "Información de Ensayo"), unsafe_allow_html=True)
            st.markdown(param_table_html([("Método de Ensayo", data.get("lim_metodo"))], header_left="DATO", header_right="VALOR"), unsafe_allow_html=True)
        equipos, norma = data.get("lim_equipos", []), "INV. E-125-13 / INV. E-126-13"
    elif tipo == "masa-unitaria":
        rows = [("Masa en el aire (g)", data.get("mu_peso_aire")), ("Masa en el aire parafinado (g)", data.get("mu_peso_aire_par")),
                ("Masa en el agua parafinado (g)", data.get("mu_peso_agua_par")), ("Temperatura del agua (°C)", data.get("mu_temp_agua"))]
        with st.container(border=True):
            st.markdown(card_header_html("science", "Parámetros Registrados"), unsafe_allow_html=True)
            st.markdown(param_table_html(rows), unsafe_allow_html=True)
        equipos, norma = data.get("mu_equipos", []), data.get("mu_norma", "—")
    else:  # "cbr"
        with st.container(border=True):
            st.markdown(card_header_html("science", "Datos Iniciales"), unsafe_allow_html=True)
            rows = [
                ("Molde No.", data.get("cbr_molde")), ("Diámetro de la muestra (cm)", data.get("cbr_diametro")),
                ("Altura de la muestra (cm)", data.get("cbr_altura")), ("Masa molde (g)", data.get("cbr_masa_molde")),
                ("Masa de la muestra + molde (g) — antes", data.get("cbr_masa_muestra_molde_antes")),
                ("Masa de la muestra + molde (g) — después", data.get("cbr_masa_muestra_molde_despues")),
            ]
            st.markdown(param_table_html(rows), unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(card_header_html("water_drop", "Humedad antes de inmersión"), unsafe_allow_html=True)
            st.caption("Compartida con el ensayo de Contenido de Humedad de esta muestra.")
            hum_data = (get_assay(muestra_id, "humedad") or {}).get("data", {}) if muestra_id else {}
            if hum_data.get("hum_recipiente") or hum_data.get("hum_seco_mas_recipiente"):
                rows = [
                    ("Recipiente no.", hum_data.get("hum_recipiente")),
                    ("Peso recipiente + suelo húmedo (g)", hum_data.get("hum_masa_humedo_mas_recipiente")),
                    ("Peso recipiente + suelo seco (g)", hum_data.get("hum_seco_mas_recipiente")),
                    ("Peso recipiente (g)", hum_data.get("hum_masa_recipiente")),
                    ("Humedad (%)", fmt_num(calcular_humedad_pct(hum_data), decimals=2)),
                ]
                st.markdown(param_table_html(rows), unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="color:{NEUTRAL};font-style:italic;">— sin datos —</div>', unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(card_header_html("water_drop", "Humedad después de inmersión"), unsafe_allow_html=True)
            rows = [
                ("Recipiente", data.get("cbr_desp_recipiente")),
                ("Peso recipiente + suelo húmedo (g)", data.get("cbr_desp_masa_humedo")),
                ("Peso recipiente + suelo seco (g)", data.get("cbr_desp_masa_seco")),
                ("Peso recipiente (g)", data.get("cbr_desp_masa_recipiente")),
            ]
            st.markdown(param_table_html(rows), unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(card_header_html("straighten", "Datos de Expansión"), unsafe_allow_html=True)
            rows = [
                ("Lectura inicial (in)", data.get("cbr_exp_lectura_inicial")),
                ("Lectura final (in)", data.get("cbr_exp_lectura_final")),
            ]
            st.markdown(param_table_html(rows), unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(card_header_html("tune", "Condiciones del Ensayo"), unsafe_allow_html=True)
            rows = [
                ("Pesas de sobrecarga (g) — antes", data.get("cbr_pesas_antes") or "4554 (por defecto)"),
                ("Pesas de sobrecarga (g) — después", data.get("cbr_pesas_despues") or "4554 (por defecto)"),
                ("Tiempo de inmersión (días) — antes", data.get("cbr_tiempo_inmersion_antes") or "4 (por defecto)"),
                ("Tiempo de inmersión (días) — después", data.get("cbr_tiempo_inmersion_despues") or "4 (por defecto)"),
            ]
            st.markdown(param_table_html(rows), unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(card_header_html("show_chart", "Penetración"), unsafe_allow_html=True)
            headers = ["PROFUNDIDAD (in)", "FUERZA ANTES (kN)", "FUERZA DESPUÉS (kN)"]
            pen_rows = [(pulg, data.get(f"cbr_pen_antes_{i}"), data.get(f"cbr_pen_despues_{i}"))
                        for i, (pulg, _mm) in enumerate(CBR_PENETRACION_FILAS, start=1)]
            st.markdown(param_table_ncol_html(headers, pen_rows), unsafe_allow_html=True)
        equipos, norma = data.get("cbr_equipos", []), data.get("cbr_norma", "—")

    with st.container(border=True):
        st.markdown(card_header_html("rule", "Norma Aplicada"), unsafe_allow_html=True)
        st.markdown(f'<div style="font-weight:600;">{html.escape(norma or "—")}</div>', unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown(card_header_html("construction", "Equipos Utilizados"), unsafe_allow_html=True)
        st.markdown(equipos_readonly_html(equipos), unsafe_allow_html=True)


def render_assay_form():
    assay_id = st.session_state.selected_assay_id
    assay = next((a for a in st.session_state.assays if a["id"] == assay_id), None)
    if not assay:
        navigate("muestra-detail")
        return

    codigo, perf_codigo, muestra_id = assay["codigo_interno"], assay["perforacion_codigo"], assay["muestra_id"]
    project = get_project(codigo)
    muestra = get_muestra(codigo, perf_codigo, muestra_id)
    es_jefe = st.session_state.role == "jefe"
    es_supervisor = st.session_state.role in ("jefe", "ingeniero")
    # El Jefe y el Director Técnico solo consultan los ensayos — quien digita los datos de laboratorio es el laboratorista.
    read_only = es_supervisor or (st.session_state.role == "laboratorista" and project_status(codigo) == "ejecutado")

    if st.button("← Atrás"):
        go_back(fallback="muestra-detail")

    st.markdown(f"## {ASSAY_LABELS[assay['tipo']]}")
    st.caption("Resultados de Ensayo" if read_only else "Registro de Ensayo")
    st.markdown(f'<div style="margin-bottom:10px;">{status_badge_html(assay["status"])}&nbsp;&nbsp;'
                f'<span class="timestamp-caption">{icon("history", size=13)} Última actualización: {format_dt(assay["lastModified"])}'
                + (f' · {html.escape(assay["laboratorist"])}' if assay.get("laboratorist") else "") + '</span></div>',
                unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown(card_header_html("info", "Información General"), unsafe_allow_html=True)
        g1, g2 = st.columns(2)
        g1.markdown(f'<div class="cell-muted">Proyecto</div><div style="font-weight:600;">{html.escape(codigo)}</div>', unsafe_allow_html=True)
        g2.markdown(f'<div class="cell-muted">Sondeo</div><div style="font-weight:600;">{html.escape(perf_codigo)}</div>', unsafe_allow_html=True)
        g3, g4 = st.columns(2)
        g3.markdown(f'<div class="cell-muted" style="margin-top:12px;">Muestra</div>'
                    f'<div style="font-weight:600;">M-{html.escape(str(muestra["numero"])) if muestra else "—"}</div>', unsafe_allow_html=True)
        profundidad_txt = f'{muestra["profundidad_de"]:.2f}m - {muestra["profundidad_hasta"]:.2f}m' if muestra else "—"
        g4.markdown(f'<div class="cell-muted" style="margin-top:12px;">Profundidad</div><div style="font-weight:600;">{profundidad_txt}</div>',
                    unsafe_allow_html=True)
        if muestra is not None:
            st.markdown(f'<div class="cell-muted" style="margin-top:12px;">Descripción visual de la muestra</div>'
                        f'<div style="font-weight:600;">{html.escape(descripcion_visual_para_excel(muestra) or "— (el laboratorista aún no la digita) —")}</div>',
                        unsafe_allow_html=True)
            # Ver descripcion_visual_calculada: la misma frase pero con el tipo de suelo que de
            # verdad salió en la clasificación USCS, en vez del que se eligió a ojo — debajo de
            # la inicial, no en su lugar.
            descripcion_calc = descripcion_visual_calculada(muestra)
            if descripcion_calc:
                st.markdown(f'<div class="cell-muted" style="margin-top:10px;">Según la clasificación USCS calculada</div>'
                            f'<div style="font-weight:600;color:{PRIMARY};">{html.escape(descripcion_calc)}</div>',
                            unsafe_allow_html=True)

    if muestra is not None:
        with st.container(border=True):
            st.markdown(card_header_html("thermostat", "Condición del Ensayo"), unsafe_allow_html=True)
            with st.expander("Ver temperatura y humedad", icon=":material/thermostat:",
                              expanded=bool(muestra.get("cond_inicial_temp") or muestra.get("cond_inicial_hum"))):
                st.caption("Se digita una sola vez por muestra: la inicial al empezar el ensayo y la final al terminarlo. Se comparte entre todos los ensayos de esta muestra.")
                if read_only:
                    st.markdown(condicion_table_html(muestra), unsafe_allow_html=True)
                else:
                    head = st.columns([1.4, 1, 1])
                    head[1].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Temperatura °C</div>', unsafe_allow_html=True)
                    head[2].markdown('<div class="cell-muted" style="text-align:center;font-weight:700;">Humedad %</div>', unsafe_allow_html=True)
                    for cond_key, cond_label in (("inicial", "Inicial"), ("final", "Final")):
                        row = st.columns([1.4, 1, 1])
                        row[0].markdown(f'<div style="padding-top:8px;">{cond_label}</div>', unsafe_allow_html=True)
                        nuevo_temp = row[1].text_input(
                            f"Temperatura {cond_label}", value=muestra.get(f"cond_{cond_key}_temp", ""),
                            key=f"cond_{cond_key}_temp_{muestra_id}", label_visibility="collapsed", placeholder="0.0")
                        nuevo_hum = row[2].text_input(
                            f"Humedad {cond_label}", value=muestra.get(f"cond_{cond_key}_hum", ""),
                            key=f"cond_{cond_key}_hum_{muestra_id}", label_visibility="collapsed", placeholder="0")
                        cambios_cond = {}
                        if nuevo_temp != muestra.get(f"cond_{cond_key}_temp", ""):
                            cambios_cond[f"cond_{cond_key}_temp"] = nuevo_temp
                        if nuevo_hum != muestra.get(f"cond_{cond_key}_hum", ""):
                            cambios_cond[f"cond_{cond_key}_hum"] = nuevo_hum
                        if cambios_cond:
                            db.update_muestra(muestra["id"], **cambios_cond)
                            muestra.update(cambios_cond)

    # Pasa 200 comparte plantilla y datos con Granulometría: si la muestra también tiene un
    # ensayo de Granulometría, "Pasa 200" lee y escribe directamente sobre ESE diccionario de
    # datos (no el suyo propio), así que lo digitado en cualquiera de las dos pantallas se ve
    # reflejado en la otra. Si no hay Granulometría, Pasa 200 usa sus propios datos, igual que
    # cualquier otro ensayo independiente.
    pasa200_gran_sibling = get_assay(muestra_id, "granulometria") if assay["tipo"] == "pasa200" else None
    data = dict(pasa200_gran_sibling["data"]) if pasa200_gran_sibling else dict(assay.get("data", {}))

    if read_only:
        if es_supervisor:
            st.info("Estás viendo el ensayo en modo consulta — solo el laboratorista puede digitar estos datos.")
        else:
            st.info("Este proyecto ya fue ejecutado. Estás en modo consulta — no puedes editar estos datos.")
        render_read_only_summary(assay["tipo"], data, assay.get("laboratorist") or "—", muestra_id=muestra_id)
        with st.container(border=True):
            st.markdown(card_header_html("notes", "Observaciones"), unsafe_allow_html=True)
            st.markdown(f'<div>{html.escape(assay.get("observations") or "—")}</div>', unsafe_allow_html=True)
        if assay["tipo"] != "humedad":
            with st.container(border=True):
                st.markdown(card_header_html("person", "Laboratorista"), unsafe_allow_html=True)
                st.markdown(f'<div style="font-weight:600;">{html.escape(assay.get("laboratorist") or "—")}</div>', unsafe_allow_html=True)
        if es_jefe and assay["status"] == "finalizado":
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Habilitar edición para el laboratorista", icon=":material/lock_open:", use_container_width=True):
                db.update_assay_data(assay["id"], status="en-proceso")
                st.success("Ensayo habilitado — el laboratorista ya puede volver a digitar los datos.")
                st.rerun()
    else:
        intento_incompleto_key = f"_intento_incompleto_{assay_id}"
        if st.session_state.get(intento_incompleto_key):
            faltantes_actuales = campos_faltantes(assay["tipo"], data)
            if faltantes_actuales:
                css_reglas = "".join(
                    f'.st-key-{key}_{assay_id} input {{ border: 2px solid #d32f2f !important; '
                    f'background-color: #fdecea !important; }}\n'
                    for key, _ in faltantes_actuales
                )
                st.markdown(f"<style>{css_reglas}</style>", unsafe_allow_html=True)
                st.error("⚠️ Faltan datos por digitar antes de poder enviar este ensayo a revisión "
                          "(resaltados en rojo abajo):\n\n"
                          + "\n".join(f"- {label}" for _, label in faltantes_actuales))
            else:
                st.session_state.pop(intento_incompleto_key, None)

        if assay["tipo"] == "granulometria":
            render_granulometria_form(data, assay_id)
        elif assay["tipo"] == "pasa200":
            render_pasa200_form(data, assay_id)
        elif assay["tipo"] == "humedad":
            render_humedad_form(data, assay_id)
        elif assay["tipo"] == "limites":
            render_limites_form(data, assay_id)
        elif assay["tipo"] == "masa-unitaria":
            render_masa_unitaria_form(data, assay_id)
        elif assay["tipo"] == "cbr":
            render_cbr_form(data, assay_id, muestra_id)

        with st.expander("Observaciones (opcional)", icon=":material/notes:", expanded=bool(assay.get("observations"))):
            observations = st.text_area("Observaciones", value=assay.get("observations", ""), label_visibility="collapsed",
                                         placeholder="Observaciones generales del ensayo, en caso de que se requiera…")

        st.markdown('<div class="section-title">Laboratorista</div>', unsafe_allow_html=True)
        laboratorist = st.text_input("Laboratorista", value=assay.get("laboratorist", ""), label_visibility="collapsed", placeholder="Nombre completo")

        # Autoguardado: si el laboratorista digita y se le olvida darle a "Guardar borrador"
        # antes de salir, los datos no se pierden — se persisten solos en cada rerun (cada vez
        # que se completa un campo), sin necesidad de un clic explícito.
        if (data != assay.get("data", {}) or observations != assay.get("observations", "")
                or laboratorist != assay.get("laboratorist", "")):
            nuevo_status = "en-proceso" if assay["status"] == "sin-iniciar" else assay["status"]
            db.update_assay_data(assay["id"], data=data, observations=observations, laboratorist=laboratorist, status=nuevo_status)
            if pasa200_gran_sibling:
                db.update_assay_shared_data(muestra["id"], ["granulometria", "pasa200"], data)
            assay["data"] = data
            assay["observations"] = observations
            assay["laboratorist"] = laboratorist
            assay["status"] = nuevo_status
        st.markdown(f'<div class="timestamp-caption">{icon("cloud_done", size=13)} Los cambios se guardan automáticamente mientras digitas.</div>',
                    unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Guardar borrador", use_container_width=True, icon=":material/save:"):
                db.update_assay_data(assay["id"], data=data, observations=observations, laboratorist=laboratorist, status="en-proceso")
                if pasa200_gran_sibling:
                    db.update_assay_shared_data(muestra["id"], ["granulometria", "pasa200"], data)
                navigate("muestra-detail")
        with col2:
            if st.button("Enviar a revisión", type="primary", use_container_width=True, icon=":material/send:"):
                faltantes = campos_faltantes(assay["tipo"], data)
                if faltantes:
                    st.session_state[intento_incompleto_key] = True
                    st.rerun()
                else:
                    st.session_state.pop(intento_incompleto_key, None)
                    ya_estaba_finalizado = assay["status"] == "finalizado"
                    db.update_assay_data(assay["id"], data=data, observations=observations, laboratorist=laboratorist, status="finalizado")
                    assay.update(data=data, observations=observations, laboratorist=laboratorist, status="finalizado")
                    if pasa200_gran_sibling:
                        db.update_assay_shared_data(muestra["id"], ["granulometria", "pasa200"], data)
                    if muestra:
                        actor_lab = f"{laboratorist} (Laboratorista)" if laboratorist else "Laboratorista"
                        add_historial(assay, "Enviado a Revisión", actor_lab, icono="science", tono="primary")
                        # Cada ensayo que el laboratorista termina se avisa al Jefe (nunca al Director Técnico,
                        # que solo entra en juego cuando el Jefe confirma ese ensayo individual).
                        if not ya_estaba_finalizado:
                            add_notification("jefe", f"El laboratorista terminó {ASSAY_LABELS[assay['tipo']]} de la Muestra "
                                                      f"{muestra['numero']} de {codigo}.", codigo, perf_codigo, muestra_id)
                            # Si este era el último ensayo pendiente, la muestra completa su ciclo en el
                            # laboratorio (semáforo en rojo) — se avisa aparte para que la confirme.
                            if compute_muestra_estado(muestra) == "finalizado":
                                add_notification("jefe", f"La Muestra {muestra['numero']} de {codigo} ya completó todos sus "
                                                          f"ensayos — está lista para tu confirmación.", codigo, perf_codigo, muestra_id)
                    navigate("muestra-detail")

    if es_supervisor and assay["tipo"] == "granulometria" and muestra:
        st.markdown("---")
        st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
        excel_bytes = generar_excel_granulometria(codigo, perf_codigo, muestra, project, data, assay.get("observations", ""))
        st.download_button(
            "Descargar Excel (Granulometría y Límites de Atterberg — mismo archivo por muestra)", icon=":material/download:",
            data=excel_bytes, file_name=f"Clasificacion_de_suelos_{muestra['id_unico']}.xlsm",
            mime="application/vnd.ms-excel.sheet.macroEnabled.12", use_container_width=True,
        )

    if es_supervisor and assay["tipo"] == "pasa200" and muestra:
        st.markdown("---")
        st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
        excel_bytes = generar_excel_pasa200(codigo, perf_codigo, muestra, project, data, assay.get("observations", ""))
        st.download_button(
            "Descargar Excel (Granulometría — mismo archivo por muestra)", icon=":material/download:",
            data=excel_bytes, file_name=f"Clasificacion_de_suelos_{muestra['id_unico']}.xlsm",
            mime="application/vnd.ms-excel.sheet.macroEnabled.12", use_container_width=True,
        )

    if es_supervisor and assay["tipo"] == "humedad" and muestra:
        st.markdown("---")
        st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
        excel_bytes = generar_excel_humedad(codigo, perf_codigo, muestra, project, data, assay.get("observations", ""))
        st.download_button(
            "Descargar Excel (plantilla oficial de Humedad)", icon=":material/download:",
            data=excel_bytes, file_name=f"Humedad_{muestra['id_unico']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True,
        )

    if es_supervisor and assay["tipo"] == "limites" and muestra:
        st.markdown("---")
        st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
        excel_bytes = generar_excel_limites(codigo, perf_codigo, muestra, project, data, assay.get("observations", ""))
        st.download_button(
            "Descargar Excel (Granulometría y Límites de Atterberg — mismo archivo por muestra)", icon=":material/download:",
            data=excel_bytes, file_name=f"Clasificacion_de_suelos_{muestra['id_unico']}.xlsm",
            mime="application/vnd.ms-excel.sheet.macroEnabled.12", use_container_width=True,
        )

    if es_supervisor and assay["tipo"] == "masa-unitaria" and muestra:
        st.markdown("---")
        st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
        excel_bytes = generar_excel_masa_unitaria(codigo, perf_codigo, muestra, project, data, assay.get("observations", ""))
        st.download_button(
            "Descargar Excel (plantilla oficial de Peso Unitario Parafinado)", icon=":material/download:",
            data=excel_bytes, file_name=f"Peso_unitario_parafinado_{muestra['id_unico']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True,
        )

    if es_supervisor and assay["tipo"] == "cbr" and muestra:
        st.markdown("---")
        st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
        excel_bytes = generar_excel_cbr(codigo, perf_codigo, muestra, project, data, assay.get("observations", ""))
        st.download_button(
            "Descargar Excel (plantilla oficial de CBR)", icon=":material/download:",
            data=excel_bytes, file_name=f"CBR_{muestra['id_unico']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True,
        )


# ════════════════════════════════════════════════════════════════════
# CONTINUAR / BUSCAR
# ════════════════════════════════════════════════════════════════════
def render_continue():
    if st.button("← Atrás"):
        go_back()
    st.markdown("## Continuar ensayo")
    in_progress = [a for a in st.session_state.assays if a["status"] == "en-proceso"]
    if not in_progress:
        st.info("No hay ensayos en proceso.")
    for a in in_progress:
        with st.container(border=True):
            cols = st.columns([3, 2, 2, 1])
            cols[0].markdown(f"**{a['codigo_interno']}**")
            cols[1].markdown(f"{a['perforacion_codigo']} · Muestra {a['muestra_numero']}")
            cols[2].markdown(ASSAY_LABELS[a["tipo"]])
            with cols[3]:
                if st.button("Continuar", key=f"cont_{a['id']}", use_container_width=True):
                    st.session_state.selected_codigo = a["codigo_interno"]
                    st.session_state.selected_perforacion = a["perforacion_codigo"]
                    st.session_state.selected_muestra_id = a["muestra_id"]
                    st.session_state.selected_assay_id = a["id"]
                    st.session_state.read_only_view = False
                    navigate("assay-form")


SEARCH_PAGE_SIZE = 8


def render_search():
    if st.button("← Atrás"):
        go_back()
    st.markdown("## Buscar ensayos")

    codes = [p["codigo_interno"] for p in st.session_state.projects]
    if not codes:
        st.info("Todavía no hay proyectos.")
        return

    with st.container(border=True):
        st.markdown(card_header_html("filter_list", "Filtros de Búsqueda"), unsafe_allow_html=True)
        default_idx = codes.index(st.session_state.selected_codigo) if st.session_state.selected_codigo in codes else 0
        codigo = st.selectbox("Proyecto", codes, index=default_idx)

        perforaciones = st.session_state.perforaciones.get(codigo, [])
        perf_options = ["(todas)"] + [p["codigo"] for p in perforaciones]
        perf_choice = st.selectbox("Perforación", perf_options)

        perfs_to_show = perforaciones if perf_choice == "(todas)" else [p for p in perforaciones if p["codigo"] == perf_choice]
        # El código del proyecto ya se eligió arriba — en este desplegable solo hace falta la
        # perforación y el número de muestra, no el id_unico completo repitiendo el proyecto.
        muestra_label_by_id = {
            m["id_unico"]: f"{perf['codigo']} · Muestra {m['numero']}"
            for perf in perfs_to_show for m in st.session_state.muestras.get(f"{codigo}::{perf['codigo']}", [])
        }
        muestra_options = ["(todas)"] + list(muestra_label_by_id.keys())
        muestra_choice = st.selectbox("Muestra", muestra_options,
                                       format_func=lambda v: "(todas)" if v == "(todas)" else muestra_label_by_id[v])

        f_type = st.selectbox("Tipo de ensayo", ["(todos)"] + list(ASSAY_LABELS.values()))

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Aplicar filtros", type="primary", use_container_width=True, icon=":material/search:"):
            st.session_state["search_page"] = 0

    if not perforaciones:
        st.info("Este proyecto todavía no tiene perforaciones. Ve a la Bitácora para agregarlas.")
        return

    project = get_project(codigo)
    rows = []
    for perf in perfs_to_show:
        muestras = st.session_state.muestras.get(f"{codigo}::{perf['codigo']}", [])
        for m in muestras:
            if muestra_choice != "(todas)" and m["id_unico"] != muestra_choice:
                continue
            solicitados = [e for e, v in m["ensayos"].items() if v and e in BITACORA_ENSAYOS]
            if f_type != "(todos)":
                solicitados = [e for e in solicitados if ASSAY_LABELS.get(SUPPORTED_ASSAY_MAP.get(e), None) == f_type]
            for ensayo_label in solicitados:
                tipo_interno = SUPPORTED_ASSAY_MAP.get(ensayo_label)
                if not tipo_interno:
                    continue
                rows.append((perf, m, ensayo_label, tipo_interno))

    with st.container(border=True):
        col_ratios = [1.7, 2.2, 1.8, 1.3, 0.6]
        headers = st.columns(col_ratios)
        for col, label in zip(headers, ["ID ensayo", "Proyecto", "Tipo / Muestra", "Estado", ""]):
            col.markdown(f'<div class="assigned-th">{label}</div>', unsafe_allow_html=True)

        if not rows:
            st.info("No se encontraron ensayos con esos filtros.")
        else:
            total = len(rows)
            total_pages = max(1, (total + SEARCH_PAGE_SIZE - 1) // SEARCH_PAGE_SIZE)
            page = min(st.session_state.get("search_page", 0), total_pages - 1)
            st.session_state["search_page"] = page
            start = page * SEARCH_PAGE_SIZE
            for i, (perf, m, ensayo_label, tipo_interno) in enumerate(rows[start:start + SEARCH_PAGE_SIZE]):
                if i:
                    st.markdown(f'<hr style="margin:8px 0;border-color:{BORDER};">', unsafe_allow_html=True)
                existing = get_assay(m["id_unico"], tipo_interno)
                status = existing["status"] if existing else "sin-iniciar"
                ensayo_id = f'{codigo}-{perf["codigo"]}-M{m["numero"]}'
                cols = st.columns(col_ratios, vertical_alignment="center")
                cols[0].markdown(f'<span class="cell-id">{html.escape(ensayo_id)}</span>', unsafe_allow_html=True)
                cols[1].markdown(f'<div class="cell-title">{html.escape(project["nombre"] if project else codigo)}</div>'
                                  f'<div class="cell-sub">{html.escape(codigo)}</div>', unsafe_allow_html=True)
                cols[2].markdown(f'<div class="cell-title">{html.escape(ensayo_label)}</div>'
                                  f'<div class="cell-sub">Muestra {html.escape(str(m["numero"]))}</div>', unsafe_allow_html=True)
                with cols[3]:
                    st.markdown(f'<div style="text-align:center;">{status_circle_html(status, size=16)}</div>', unsafe_allow_html=True)
                with cols[4]:
                    if st.button("", key=f"search_open_{m['id_unico']}_{tipo_interno}", icon=":material/chevron_right:",
                                 use_container_width=True, help="Abrir"):
                        if existing:
                            st.session_state.selected_assay_id = existing["id"]
                        else:
                            nuevo = db.create_assay(m["id"], tipo_interno)
                            st.session_state.selected_assay_id = nuevo["id"]
                        st.session_state.selected_codigo = codigo
                        st.session_state.selected_perforacion = perf["codigo"]
                        st.session_state.selected_muestra_id = m["id_unico"]
                        st.session_state.selected_assay_type = tipo_interno
                        navigate("assay-form")

            st.markdown("<hr style='margin:8px 0;'>", unsafe_allow_html=True)
            f1, f2, f3 = st.columns([3, 1, 1])
            f1.caption(f"Mostrando {len(rows[start:start + SEARCH_PAGE_SIZE])} de {total} resultado(s)")
            with f2:
                if st.button("", key="search_prev", icon=":material/chevron_left:", use_container_width=True, disabled=page == 0):
                    st.session_state["search_page"] = page - 1
                    st.rerun()
            with f3:
                if st.button("", key="search_next", icon=":material/chevron_right:", use_container_width=True,
                             disabled=page >= total_pages - 1):
                    st.session_state["search_page"] = page + 1
                    st.rerun()


# ════════════════════════════════════════════════════════════════════
# ENRUTADOR PRINCIPAL
# ════════════════════════════════════════════════════════════════════
# El borrado de la cookie de sesión al cerrar sesión tiene el mismo problema de timing que
# escribirla al iniciar sesión (ver _set_session_cookie/_pending_cookie_tokens): si se llama justo
# antes de un st.rerun(), el componente no alcanza a mandarle la orden de borrado al navegador
# antes de que el rerun reemplace la página. Por eso se difiere igual, un rerun después — y como
# el logout deja st.session_state.role en None, este chequeo va ANTES del if/else de abajo, no
# adentro del "else" (que solo corre con sesión iniciada).
if st.session_state.pop("_pending_logout_cookie_clear", False):
    _clear_session_cookie()

if st.session_state.role is None:
    render_login()
else:
    if st.session_state.get("_pending_cookie_tokens"):
        tokens = st.session_state.pop("_pending_cookie_tokens")
        _set_session_cookie(tokens["access_token"], tokens["refresh_token"])
    if "_pending_remember_user" in st.session_state:
        codigo_a_recordar = st.session_state.pop("_pending_remember_user")
        if codigo_a_recordar:
            _set_remember_user_cookie(codigo_a_recordar)
        else:
            _clear_remember_user_cookie()
    if st.session_state.pop("_pending_history_push", False):
        _push_history_entry()
    _load_data()
    # Si algo de lo anterior refrescó el token (ver _tokens_rotated en db._refresh_if_needed
    # y en el restore de cookie de init_state), la cookie del navegador queda con un
    # refresh_token ya rotado/vencido si no se vuelve a guardar aquí con el vigente.
    if st.session_state.pop("_tokens_rotated", False):
        tokens_vigentes = db.get_session_tokens()
        if tokens_vigentes:
            _set_session_cookie(tokens_vigentes["access_token"], tokens_vigentes["refresh_token"])
    render_topbar()
    SCREENS = {
        "home": render_home, "new-project": render_new_project, "project-detail": render_project_detail,
        "edit-project": render_edit_project,
        "perforacion-detail": render_perforacion_detail, "muestra-detail": render_muestra_detail,
        "bitacora": render_bitacora, "assay-form": render_assay_form,
        "continue": render_continue, "search": render_search,
        "projects-active": render_projects_active, "projects-done": render_projects_done,
    }
    SCREENS.get(st.session_state.screen, render_home)()
    render_bottomnav()
